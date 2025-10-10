import os
import re
import calendar
import logging
from io import BytesIO
from datetime import datetime

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from dotenv import load_dotenv
import ibm_boto3
from ibm_botocore.client import Config

# -----------------------------------------------------------------------------
# CONFIG / CONSTANTS
# -----------------------------------------------------------------------------
load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

COS_API_KEY    = os.getenv("COS_API_KEY")
COS_CRN        = os.getenv("COS_SERVICE_INSTANCE_CRN")
COS_ENDPOINT   = os.getenv("COS_ENDPOINT")
BUCKET         = os.getenv("COS_BUCKET_NAME")
KRA_KEY        = os.getenv("KRA_FILE_PATH")

# Dynamic tracker paths - will be set by get_latest_tracker_paths()
T6_TRACKER_KEY = None
T5_TRACKER_KEY = None
T7_TRACKER_KEY = None
GREEN3_TRACKER_KEY = None

# Dynamic months - will be set based on tracker date
MONTHS = []
MONTHS_DATA = []  # Store (month_num, year) tuples
TRACKER_DATE = None
PROCESSING_MONTHS = []  # Months to actually process (previous months)

GREEN_HEX = "FF92D050"

TOWER6_ROWS = [4, 5, 6, 7, 9, 10, 14, 15, 16, 17, 19, 20]
TOWER6_COLS = ['FK', 'FM', 'FO', 'FQ', 'FS', 'FU', 'FW', 'FY', 'GA', 'GC', 'GE', 'GG', 'GI', 'GK']

# Dynamic target cells - will be updated based on months
T5_TARGET_CELLS = {}
T7_TARGET_CELLS = {}

# HARDCODED VALUES FOR T7 EL-FIRST FIX - will be updated dynamically
T7_HARDCODED_VALUES = {}

T5_ACTIVITIES = ["Installation of Rear & Front balcony UPVC Windows", "EL-Second Fix", "Gypsum board false ceiling", "Paint 1st Coat"]
T7_ACTIVITIES = ["El- First Fix", "Floor Tiling", "False Ceiling Framing", "C-Stone flooring"]

def init_cos():
    return ibm_boto3.client(
        "s3",
        ibm_api_key_id=COS_API_KEY,
        ibm_service_instance_id=COS_CRN,
        config=Config(signature_version="oauth"),
        endpoint_url=COS_ENDPOINT,
    )

def list_files_in_folder(cos, folder_prefix):
    """List all files in a specific folder (prefix) in the COS bucket"""
    try:
        response = cos.list_objects_v2(Bucket=BUCKET, Prefix=folder_prefix)
        files = []
        if 'Contents' in response:
            for obj in response['Contents']:
                # Only include actual files, not folder markers
                if not obj['Key'].endswith('/'):
                    files.append(obj['Key'])
        return files
    except Exception as e:
        logger.error(f"Error listing files in folder {folder_prefix}: {e}")
        return []

def extract_date_from_filename(filename):
    """Extract date from filename in format (dd-mm-yyyy)"""
    # Pattern to match date in parentheses like (01-07-2025)
    pattern = r'\((\d{2}-\d{2}-\d{4})\)'
    match = re.search(pattern, filename)
    if match:
        date_str = match.group(1)
        try:
            # Parse the date
            return datetime.strptime(date_str, '%d-%m-%Y')
        except ValueError:
            logger.warning(f"Could not parse date {date_str} from filename {filename}")
            return None
    return None

def get_month_name(month_num):
    """Convert month number to month name"""
    months = {
        1: "January", 2: "February", 3: "March", 4: "April",
        5: "May", 6: "June", 7: "July", 8: "August", 
        9: "September", 10: "October", 11: "November", 12: "December"
    }
    return months.get(month_num, "Unknown")

def setup_dynamic_months_and_targets(tracker_date):
    """Setup months and target cells based on tracker date with dynamic logic and year tracking."""
    global MONTHS, MONTHS_DATA, TRACKER_DATE, PROCESSING_MONTHS, T5_TARGET_CELLS, T7_TARGET_CELLS, T7_HARDCODED_VALUES

    TRACKER_DATE = tracker_date
    tracker_month = tracker_date.month
    tracker_year = tracker_date.year

    logger.info(f"=== SETTING UP DYNAMIC MONTHS WITH YEAR TRACKING ===")
    logger.info(f"Tracker date: {tracker_date.strftime('%d-%m-%Y')} (Month: {tracker_month}, Year: {tracker_year})")

    # ===================== DYNAMIC MONTH LOGIC =====================
    months_list = [
        'January', 'February', 'March', 'April', 'May', 'June',
        'July', 'August', 'September', 'October', 'November', 'December'
    ]

    # ✅ Rule 1: September tracker → June, July, August (same year)
    if tracker_month == 9:
        MONTHS_DATA = [(6, tracker_year), (7, tracker_year), (8, tracker_year)]
        logger.info("Special September rule applied → June–July–August included")

    # ✅ Rule 2: Otherwise → previous, current, next (year-aware)
    else:
        from datetime import timedelta

        prev_month_date = tracker_date.replace(day=1) - timedelta(days=1)
        next_month_year = tracker_year + (1 if tracker_month == 12 else 0)
        next_month_num = 1 if tracker_month == 12 else tracker_month + 1
        next_month_date = tracker_date.replace(year=next_month_year, month=next_month_num, day=1)

        MONTHS_DATA = [
            (prev_month_date.month, prev_month_date.year),
            (tracker_month, tracker_year),
            (next_month_date.month, next_month_date.year)
        ]
        logger.info("Standard dynamic rule applied → previous, current, next months")

    # Build labeled month list
    MONTHS = [f"{months_list[m - 1]}" for m, y in MONTHS_DATA]
    MONTHS_WITH_YEAR = [f"{months_list[m - 1]} {y}" for m, y in MONTHS_DATA]

    TARGET_MONTH = MONTHS_WITH_YEAR[-1]

    logger.info(f"Generated MONTHS: {MONTHS}")
    logger.info(f"With years: {MONTHS_WITH_YEAR}")
    logger.info(f"Target month: {TARGET_MONTH}")

    # ===================== PROCESSING MONTH LOGIC =====================
    PROCESSING_MONTHS = []
    for month_name, (month_num, year) in zip(MONTHS, MONTHS_DATA):
        if (year < tracker_year) or (year == tracker_year and month_num < tracker_month):
            PROCESSING_MONTHS.append(month_name)
            logger.info(f"Including {month_name} {year} for processing (before tracker date)")

    # Ensure at least one month
    if not PROCESSING_MONTHS and MONTHS:
        PROCESSING_MONTHS = [MONTHS[-1]]
        logger.info(f"No processing months found, defaulting to last month: {MONTHS[-1]}")

    logger.info(f"PROCESSING_MONTHS: {PROCESSING_MONTHS}")

    # ===================== T5 TARGET CELLS =====================
    T5_TARGET_CELLS = {}
    base_rows = [23, 24, 25, 26]  # Row numbers for each activity
    base_cols = ['D', 'E', 'F']   # Columns for the 3 months
    units = ['Flats', 'Flats', 'Flats', 'Modules']

    for i, activity in enumerate(T5_ACTIVITIES):
        T5_TARGET_CELLS[activity] = {}
        for j, month in enumerate(MONTHS):
            if j < len(base_cols):
                cell = f"{base_cols[j]}{base_rows[i]}"
                T5_TARGET_CELLS[activity][month] = (cell, units[i])

    # ===================== T7 TARGET CELLS =====================
    T7_TARGET_CELLS = {}
    base_rows = [30, 31, 32, 33]  # Row numbers for each activity
    units = ['Flats', 'Flats', 'Flats', 'Modules']

    for i, activity in enumerate(T7_ACTIVITIES):
        T7_TARGET_CELLS[activity] = {}
        for j, month in enumerate(MONTHS):
            if j < len(base_cols):
                cell = f"{base_cols[j]}{base_rows[i]}"
                T7_TARGET_CELLS[activity][month] = (cell, units[i])

    T7_HARDCODED_VALUES = {}

    logger.info(f"T5_TARGET_CELLS keys: {list(T5_TARGET_CELLS.keys())}")
    logger.info(f"T7_TARGET_CELLS keys: {list(T7_TARGET_CELLS.keys())}")

    # Return for reference (optional)
    return MONTHS, TARGET_MONTH, TRACKER_DATE, MONTHS_DATA, PROCESSING_MONTHS

def get_latest_tracker_paths(cos):
    """Get the latest tracker file paths for all Veridia trackers"""
    global T6_TRACKER_KEY, T5_TRACKER_KEY, T7_TRACKER_KEY, GREEN3_TRACKER_KEY
    
    logger.info("=== FINDING LATEST VERIDIA TRACKER FILES ===")
    
    # List all files in Veridia folder
    veridia_files = list_files_in_folder(cos, "Veridia/")
    logger.info(f"Found {len(veridia_files)} files in Veridia folder")
    
    # Define tracker patterns and their corresponding variable names
    tracker_patterns = {
        'T6_TRACKER': r'Structure Work Tracker.*\.xlsx$',
        'T5_TRACKER': r'Tower 5 Finishing Tracker.*\.xlsx$', 
        'T7_TRACKER': r'Tower 7 Finishing Tracker.*\.xlsx$',
        'GREEN3_TRACKER': r'External Development Green 3 Tracker.*\.xlsx$'
    }
    
    latest_trackers = {}
    latest_date = None
    
    for tracker_type, pattern in tracker_patterns.items():
        logger.info(f"\n--- Looking for {tracker_type} files ---")
        matching_files = []
        
        for file_path in veridia_files:
            filename = os.path.basename(file_path)
            if re.search(pattern, filename, re.IGNORECASE):
                file_date = extract_date_from_filename(filename)
                if file_date:
                    matching_files.append((file_path, file_date))
                    logger.info(f"Found: {filename} with date {file_date.strftime('%d-%m-%Y')}")
                    
                    # Track the latest date across all trackers
                    if latest_date is None or file_date > latest_date:
                        latest_date = file_date
                else:
                    logger.warning(f"Found matching file but no date: {filename}")
        
        if matching_files:
            # Sort by date and get the latest
            latest_file = max(matching_files, key=lambda x: x[1])
            latest_trackers[tracker_type] = latest_file[0]
            logger.info(f"✅ Latest {tracker_type}: {latest_file[0]} ({latest_file[1].strftime('%d-%m-%Y')})")
        else:
            logger.error(f"❌ No {tracker_type} files found!")
            latest_trackers[tracker_type] = None
    
    # Set the global variables
    T6_TRACKER_KEY = latest_trackers.get('T6_TRACKER')
    T5_TRACKER_KEY = latest_trackers.get('T5_TRACKER')
    T7_TRACKER_KEY = latest_trackers.get('T7_TRACKER')
    GREEN3_TRACKER_KEY = latest_trackers.get('GREEN3_TRACKER')
    
    logger.info(f"\n=== FINAL TRACKER PATHS ===")
    logger.info(f"T6_TRACKER_KEY: {T6_TRACKER_KEY}")
    logger.info(f"T5_TRACKER_KEY: {T5_TRACKER_KEY}")
    logger.info(f"T7_TRACKER_KEY: {T7_TRACKER_KEY}")
    logger.info(f"GREEN3_TRACKER_KEY: {GREEN3_TRACKER_KEY}")
    logger.info(f"Latest tracker date found: {latest_date.strftime('%d-%m-%Y') if latest_date else 'None'}")
    
    # Setup dynamic months and targets based on the latest tracker date
    if latest_date:
        setup_dynamic_months_and_targets(latest_date)
    else:
        logger.error("No valid tracker date found - using default setup")
        # Fallback to current date minus 1 month
        fallback_date = datetime.now()
        setup_dynamic_months_and_targets(fallback_date)
    
    # Verify all trackers were found
    missing_trackers = [k for k, v in latest_trackers.items() if v is None]
    if missing_trackers:
        logger.error(f"⚠️  Missing trackers: {missing_trackers}")
        raise Exception(f"Could not find latest tracker files for: {missing_trackers}")
    
    return latest_trackers

def download_file_bytes(cos, key):
    if not key:
        raise ValueError("File key cannot be None or empty")
    obj = cos.get_object(Bucket=BUCKET, Key=key)
    return obj["Body"].read()

def extract_number(cell_value):
    if not cell_value or cell_value == "-":
        return 0.0
    match = re.search(r"(\d+)", str(cell_value))
    return float(match.group(1)) if match else 0.0

def get_previous_months():
    """Return months that should be processed based on tracker date"""
    return PROCESSING_MONTHS

def get_slab_targets_fixed_cells(cos):
    """Get slab targets dynamically based on current months"""
    raw = download_file_bytes(cos, KRA_KEY)
    wb = load_workbook(filename=BytesIO(raw), data_only=True)
    sheet = wb["VeridiaTargets Till August 2025"]
    
    # Map months to columns - extended mapping
    month_to_col = {
        "January": "I18", "February": "J18", "March": "K18",
        "April": "L18", "May": "M18", "June": "B18",
        "July": "C18", "August": "D18", "September": "E18",
        "October": "F18", "November": "G18", "December": "H18"
    }
    
    targets = {}
    for month in MONTHS:
        if month in month_to_col:
            targets[month] = extract_number(sheet[month_to_col[month]].value)
        else:
            logger.warning(f"No target cell defined for month {month}, using 0")
            targets[month] = 0.0
    
    logger.info(f"Loaded T6 slab targets: {targets}")
    return targets

def count_tower6_completed(wb):
    """Count completed Tower 6 slabs by month, considering year from MONTHS_DATA"""
    counts = {m: 0 for m in MONTHS}
    sheet = wb["Revised baseline with 60d NGT"]
    
    for row in TOWER6_ROWS:
        for col in TOWER6_COLS:
            cell = sheet[f"{col}{row}"]
            val = cell.value
            cell_date = None
            if isinstance(val, datetime):
                cell_date = val
            elif isinstance(val, str):
                try:
                    cell_date = datetime.strptime(val, "%Y-%m-%d")
                except Exception:
                    continue
            
            if cell_date:
                cell_month_name = cell_date.strftime("%B")
                cell_year = cell_date.year
                
                # Match against MONTHS_DATA to ensure year is correct
                for i, (month_num, target_year) in enumerate(MONTHS_DATA):
                    month_name = MONTHS[i]
                    if cell_month_name == month_name and cell_year == target_year:
                        fill = cell.fill
                        if fill.fill_type == "solid" and fill.start_color:
                            if fill.start_color.rgb == GREEN_HEX:
                                counts[month_name] += 1
                                logger.debug(f"Counted slab at {col}{row}: {cell_date.strftime('%B %Y')}")
                        break
    
    logger.info(f"Tower 6 completed counts by month: {counts}")
    return counts

def build_t6_milestone_dataframe(targets, completed):
    tracker_year = TRACKER_DATE.year
    prev_months = [m for (m, y) in zip(MONTHS, [y for _, y in MONTHS_DATA]) if y == tracker_year]
    
    total_milestones = 1
    weightage = round(100 / total_milestones, 2) if total_milestones else 0

    def pct(m):
        """Calculate % only for tracker year months with cumulative logic"""
        if m in prev_months:
            # Only sum up to the current month within tracker year
            month_idx = MONTHS.index(m)
            cumulative_done = 0
            cumulative_target = 0
            
            for i in range(len(MONTHS)):
                if i <= month_idx and MONTHS_DATA[i][1] == tracker_year:
                    cumulative_done += int(completed.get(MONTHS[i], 0))
                    cumulative_target += int(targets.get(MONTHS[i], 0))
            
            if cumulative_target == 0:
                return "0.0%"
            pct_done = min(round((cumulative_done / cumulative_target) * 100, 2), 100)
            return f"{pct_done}%"
        return ""

    # Build target text (only tracker year)
    target_parts = []
    total_target = 0
    for i, (month_num, year) in enumerate(MONTHS_DATA):
        if year == tracker_year:
            month = MONTHS[i]
            month_target = int(targets.get(month, 0))
            total_target += month_target
            if month_target > 0:
                target_parts.append(f"{month_target} Slabs-{month}")

    target_text = f"{total_target} Slabs ({', '.join(target_parts)})" if target_parts else "0 Slabs"

    row = {
        "Milestone": "Milestone-01",
        "Activity": "Slab Casting", 
        "Target Till August": target_text,
        "Weightage": weightage,
        "Weighted Delay against Targets": "",
        "Total achieved": "",
        f"Delay Reasons_{TRACKER_DATE.strftime('%B %Y')}": "",
    }

    # Percentage columns
    for month in MONTHS:
        row[f"% Work Done against Target-Till {month}"] = pct(month)

    # Target achieved columns
    for month in MONTHS:
        if month in prev_months:
            month_target = int(targets.get(month, 0))
            month_completed = int(completed.get(month, 0))
            row[f"Target achieved in {month}"] = f"{month_completed} slab cast out of {month_target} planned"
        else:
            row[f"Target achieved in {month}"] = ""

    # Weighted delay = only for last tracker year month
    if prev_months:
        try:
            last_processed_month = prev_months[-1]
            pct_str = pct(last_processed_month).replace("%", "")
            if pct_str:
                pct_val = float(pct_str)
                row["Weighted Delay against Targets"] = f"{round((pct_val * weightage) / 100, 2)}%"
        except Exception:
            pass

    # Column order
    all_cols = ["Milestone", "Activity", "Target Till August"]
    for month in MONTHS:
        all_cols.append(f"% Work Done against Target-Till {month}")
    all_cols.extend(["Weightage", "Weighted Delay against Targets"])
    for month in MONTHS:
        all_cols.append(f"Target achieved in {month}")
    all_cols.extend(["Total achieved", f"Delay Reasons_{TRACKER_DATE.strftime('%B %Y')}"])

    final_df = pd.DataFrame(columns=all_cols)
    final_df.loc[0] = row
    return final_df

def count_completed_activities_by_module_and_month(wb, sheet_name, activity_mapping):
    """Count activities considering year from MONTHS_DATA"""
    sheet = wb[sheet_name]
    activity_counts = {}
    
    for activity in activity_mapping.keys():
        activity_counts[activity] = {month: 0 for month in MONTHS}
    
    actual_finish_col = None
    activity_name_col = None
    
    # Find the columns for Actual Finish and Activity
    for row in sheet.iter_rows(min_row=1, max_row=10):
        for cell in row:
            if cell.value:
                if "Actual Finish" in str(cell.value):
                    actual_finish_col = cell.column
                if "Activity" in str(cell.value) or "Task" in str(cell.value):
                    activity_name_col = cell.column
        if actual_finish_col:
            break
    
    if not actual_finish_col:
        return activity_counts
    
    if not activity_name_col:
        activity_name_col = 6
    
    logger.info(f"Processing sheet: {sheet_name}")
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        try:
            activity_cell = row[activity_name_col - 1] if len(row) >= activity_name_col else None
            if not activity_cell or not activity_cell.value:
                continue
                
            activity_name = str(activity_cell.value).strip()
            mapped_activity = None
            
            # Check all activities in the mapping
            for standard_name, variations in activity_mapping.items():
                if standard_name == "El- First Fix":
                    activity_lower = activity_name.lower().strip()
                    if (activity_name == "EL-First Fix" or
                        activity_name == "El- First Fix" or
                        activity_name == "EL- First Fix" or
                        activity_name == "EL First Fix" or
                        activity_name == "El-First Fix" or
                        activity_name == "Electrical First Fix" or
                        activity_lower == "el-first fix" or
                        activity_lower == "el- first fix" or
                        activity_lower == "el first fix"):
                        mapped_activity = standard_name
                        break
                        
                elif standard_name == "Installation of Rear & Front balcony UPVC Windows":
                    if (activity_name == standard_name or 
                        activity_name == "Installation of Rear &amp; Front balcony UPVC Windows" or
                        activity_name == "Installation of Rear and Front balcony UPVC Windows" or
                        activity_name == "Installation of Rear & Front Balcony UPVC Windows" or
                        activity_name == "Installation of rear & front balcony UPVC Windows"):
                        mapped_activity = standard_name
                        break
                        
                else:
                    if activity_name in variations or activity_name.lower() in [v.lower() for v in variations]:
                        mapped_activity = standard_name
                        break
            
            if not mapped_activity:
                continue
            
            # Check actual finish date
            actual_finish_cell = row[actual_finish_col - 1] if len(row) >= actual_finish_col else None
            if not actual_finish_cell or not actual_finish_cell.value:
                continue
            
            actual_finish_date = None
            if isinstance(actual_finish_cell.value, datetime):
                actual_finish_date = actual_finish_cell.value
            elif isinstance(actual_finish_cell.value, str):
                try:
                    for date_format in ["%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y"]:
                        try:
                            actual_finish_date = datetime.strptime(actual_finish_cell.value, date_format)
                            break
                        except ValueError:
                            continue
                except Exception:
                    continue
            
            if actual_finish_date:
                cell_month_name = actual_finish_date.strftime("%B")
                cell_year = actual_finish_date.year
                
                # Match against MONTHS_DATA to ensure year is correct
                for i, (month_num, target_year) in enumerate(MONTHS_DATA):
                    month_name = MONTHS[i]
                    if cell_month_name == month_name and cell_year == target_year:
                        activity_counts[mapped_activity][month_name] += 1
                        logger.debug(f"Counted {mapped_activity} at row {row_idx}: {actual_finish_date.strftime('%B %Y')}")
                        break
                    
        except Exception as e:
            logger.warning(f"Error processing row {row_idx} in sheet {sheet_name}: {e}")
            continue
    
    return activity_counts

def get_t5_targets_and_progress(cos):
    raw = download_file_bytes(cos, KRA_KEY)
    wb_kra = load_workbook(filename=BytesIO(raw), data_only=True)
    sheet_kra = wb_kra["VeridiaTargets Till August 2025"]

    t5_targets = {}
    for activity in T5_ACTIVITIES:
        t5_targets[activity] = {}
        for month in MONTHS:
            if month in T5_TARGET_CELLS[activity]:
                cell, unit = T5_TARGET_CELLS[activity][month]
                val = extract_number(sheet_kra[cell].value)
                t5_targets[activity][month] = (val, unit)
            else:
                t5_targets[activity][month] = (0.0, "Flats")

    raw_tracker = download_file_bytes(cos, T5_TRACKER_KEY)
    wb_tracker = load_workbook(filename=BytesIO(raw_tracker), data_only=True)

    t5_activity_mapping = {
        "Installation of Rear & Front balcony UPVC Windows": [
            "Installation of Rear & Front balcony UPVC Windows",
            "Installation of Rear &amp; Front balcony UPVC Windows",
            "Installation of Rear and Front balcony UPVC Windows"
        ],
        "EL-Second Fix": [
            "EL-Second Fix",
            "EL Second Fix",
            "Electrical Second Fix",
            "EL- Second Fix"
        ],
        "Gypsum board false ceiling": [
            "Gypsum board false ceiling",
            "Gypsum False Ceiling",
            "False Ceiling Gypsum"
        ],
        "Paint 1st Coat": [
            "Paint 1st Coat",
            "Painting First Coat",
            "Paint First Coat",
            "1st Coat Paint"
        ]
    }

    required_t5_sheets = ["M7 T5", "M6 T5", "M5 T5", "M4 T5", "M3 T5", "M2 T5"]
    t5_sheet_names = []
    available_sheets = wb_tracker.sheetnames
    
    for required_sheet in required_t5_sheets:
        if required_sheet in available_sheets:
            t5_sheet_names.append(required_sheet)
    
    if not t5_sheet_names:
        activity_counts = {}
        for activity in T5_ACTIVITIES:
            activity_counts[activity] = {month: 0 for month in MONTHS}
    else:
        activity_counts = {}
        for activity in T5_ACTIVITIES:
            activity_counts[activity] = {month: 0 for month in MONTHS}

        for sheet_name in t5_sheet_names:
            sheet_counts = count_completed_activities_by_module_and_month(
                wb_tracker, sheet_name, t5_activity_mapping
            )
            
            for activity in T5_ACTIVITIES:
                for month in MONTHS:
                    activity_counts[activity][month] += sheet_counts[activity][month]

    prev_months = get_previous_months()
    progress_data = []
    total_milestones = len(T5_ACTIVITIES)
    weightage = round(100 / total_milestones, 2) if total_milestones else 0
    
    # Get tracker year for filtering
    tracker_year = TRACKER_DATE.year

    for i, activity in enumerate(T5_ACTIVITIES):
        row = {
            "Milestone": f"Milestone-{i+1:02d}",
            "Activity": activity,
            "Weightage": weightage,
            "Weighted Delay against Targets": "",
            "Total achieved": "",
            f"Delay Reasons_{TRACKER_DATE.strftime('%B %Y')}": "",
        }
        
        # Add percentage and target achieved columns dynamically
        for m in MONTHS:
            if m in prev_months:
                # Calculate cumulative only for tracker year months
                month_idx = MONTHS.index(m)
                count_cumulative = 0
                target_cumulative = 0
                
                for i_month in range(len(MONTHS)):
                    if i_month <= month_idx and MONTHS_DATA[i_month][1] == tracker_year:
                        count_cumulative += activity_counts[activity][MONTHS[i_month]]
                        target_cumulative += t5_targets[activity][MONTHS[i_month]][0]

                if target_cumulative == 0:
                    pct_done = 100.0
                else:
                    pct_done = min(round((count_cumulative / target_cumulative) * 100, 2), 100)

                row[f"% Work Done against Target-Till {m}"] = f"{pct_done}%"
                
                month_target, month_unit = t5_targets[activity][m]
                count_in_month = activity_counts[activity][m]
                row[f"Target achieved in {m}"] = f"{count_in_month} {month_unit} out of {int(month_target)} planned"
            else:
                row[f"% Work Done against Target-Till {m}"] = ""
                row[f"Target achieved in {m}"] = ""

        # Calculate weighted delay for last processed month
        if prev_months:
            last_month = prev_months[-1]
            pct_str = row.get(f"% Work Done against Target-Till {last_month}", "0%").replace("%", "")
            try:
                pct_val = float(pct_str)
                row["Weighted Delay against Targets"] = f"{round((pct_val * weightage) / 100, 2)}%"
            except ValueError:
                row["Weighted Delay against Targets"] = ""

        # Build target text dynamically (only tracker year)
        total_target = 0
        target_parts = []
        for idx, (month_num, year) in enumerate(MONTHS_DATA):
            if year == tracker_year:
                month = MONTHS[idx]
                month_target = int(t5_targets[activity][month][0])
                total_target += month_target
                if month_target > 0:
                    unit = t5_targets[activity][month][1]
                    target_parts.append(f"{month_target} {unit}-{month}")
        
        unit = t5_targets[activity][MONTHS[0]][1] if total_target > 0 else ""
        row["Target Till August"] = f"{int(total_target)} {unit} ({', '.join(target_parts)})" if target_parts else f"0 {unit}"
        
        progress_data.append(row)

    # Create dynamic column list
    all_cols = ["Milestone", "Activity", "Target Till August"]
    for month in MONTHS:
        all_cols.append(f"% Work Done against Target-Till {month}")
    all_cols.extend(["Weightage", "Weighted Delay against Targets"])
    for month in MONTHS:
        all_cols.append(f"Target achieved in {month}")
    all_cols.extend(["Total achieved", "Delay Reasons_June 2025"])

    df_t5 = pd.DataFrame(progress_data, columns=all_cols)
    return df_t5

def get_t7_targets_and_progress(cos):
    logger.info("=== STARTING T7 PROCESSING WITH DYNAMIC MONTHS ===")
    logger.info(f"Processing months: {PROCESSING_MONTHS}")
    logger.info(f"All months: {MONTHS}")
    
    raw = download_file_bytes(cos, KRA_KEY)
    wb_kra = load_workbook(filename=BytesIO(raw), data_only=True)
    sheet_kra = wb_kra["VeridiaTargets Till August 2025"]

    t7_targets = {}
    for activity in T7_ACTIVITIES:
        t7_targets[activity] = {}
        for month in MONTHS:
            if month in T7_TARGET_CELLS[activity]:
                cell, unit = T7_TARGET_CELLS[activity][month]
                val = extract_number(sheet_kra[cell].value)
                t7_targets[activity][month] = (val, unit)
            else:
                t7_targets[activity][month] = (0.0, "Flats")

    # Override targets with hardcoded values if available
    for activity, month_data in T7_HARDCODED_VALUES.items():
        if activity in t7_targets:
            for month, values in month_data.items():
                if month in t7_targets[activity]:
                    hardcoded_target = values["target_count"]
                    t7_targets[activity][month] = (hardcoded_target, "Flats")
                    logger.info(f"OVERRIDDEN T7 target for {activity} {month}: {hardcoded_target} Flats")

    raw_tracker = download_file_bytes(cos, T7_TRACKER_KEY)
    wb_tracker = load_workbook(filename=BytesIO(raw_tracker), data_only=True)

    available_sheets = wb_tracker.sheetnames
    logger.info(f"Available sheets in T7 tracker: {available_sheets}")

    t7_activity_mapping = {
        "El- First Fix": [
            "EL-First Fix",
            "El- First Fix",
            "EL- First Fix", 
            "EL First Fix",
            "El-First Fix",
            "Electrical First Fix",
            "el-first fix",
            "el- first fix"
        ],
        "Floor Tiling": [
            "Floor Tiling",
            "Flooring Tiling",
            "Tile Flooring",
            "floor tiling"
        ],
        "False Ceiling Framing": [
            "False Ceiling Framing",
            "Ceiling Framing",
            "False Ceiling Frame",
            "false ceiling framing"
        ],
        "C-Stone flooring": [
            "C-Stone flooring",
            "C Stone flooring",
            "C-Stone Flooring",
            "CStone flooring",
            "c-stone flooring"
        ]
    }

    required_t7_sheets = ["M7 T7", "M6 T7", "M5 T7", "M4 T7", "M3 T7", "M2 T7", "M1 T7"]
    
    actual_t7_sheets = []
    for sheet_name in available_sheets:
        if any(module in sheet_name.upper() for module in ['M1', 'M2', 'M3', 'M4', 'M5', 'M6', 'M7']):
            if 'T7' in sheet_name.upper():
                actual_t7_sheets.append(sheet_name)
    
    t7_sheet_names = actual_t7_sheets if actual_t7_sheets else []
    logger.info(f"Using T7 sheets: {t7_sheet_names}")
    
    if not t7_sheet_names:
        activity_counts = {}
        for activity in T7_ACTIVITIES:
            activity_counts[activity] = {month: 0 for month in MONTHS}
    else:
        activity_counts = {}
        for activity in T7_ACTIVITIES:
            activity_counts[activity] = {month: 0 for month in MONTHS}

        for sheet_name in t7_sheet_names:
            sheet_counts = count_completed_activities_by_module_and_month(
                wb_tracker, sheet_name, t7_activity_mapping
            )
            
            for activity in T7_ACTIVITIES:
                for month in MONTHS:
                    activity_counts[activity][month] += sheet_counts[activity][month]

    # Override activity counts with hardcoded values if available
    for activity, month_data in T7_HARDCODED_VALUES.items():
        if activity in activity_counts:
            for month, values in month_data.items():
                if month in activity_counts[activity]:
                    hardcoded_count = values["completed_count"]
                    activity_counts[activity][month] = hardcoded_count
                    logger.info(f"OVERRIDDEN T7 completed count for {activity} {month}: {hardcoded_count}")

    prev_months = get_previous_months()
    progress_data = []
    total_milestones = len(T7_ACTIVITIES)
    weightage = round(100 / total_milestones, 2) if total_milestones else 0
    
    # Get tracker year for filtering
    tracker_year = TRACKER_DATE.year

    for i, activity in enumerate(T7_ACTIVITIES):
        row = {
            "Milestone": f"Milestone-{i+1:02d}",
            "Activity": activity,
            "Weightage": weightage,
            "Weighted Delay against Targets": "",
            "Total achieved": "",
            f"Delay Reasons_{TRACKER_DATE.strftime('%B %Y')}": ""
        }
        
        # Add percentage and target achieved columns dynamically
        for m in MONTHS:
            if m in prev_months:
                # Calculate cumulative only for tracker year months
                month_idx = MONTHS.index(m)
                count_cumulative = 0
                target_cumulative = 0
                
                for i_month in range(len(MONTHS)):
                    if i_month <= month_idx and MONTHS_DATA[i_month][1] == tracker_year:
                        count_cumulative += activity_counts[activity][MONTHS[i_month]]
                        target_cumulative += t7_targets[activity][MONTHS[i_month]][0]

                # Use hardcoded percentage if available
                if activity in T7_HARDCODED_VALUES and m in T7_HARDCODED_VALUES[activity]:
                    pct_done = T7_HARDCODED_VALUES[activity][m]["percentage"]
                    logger.info(f"Using hardcoded percentage for {activity} {m}: {pct_done}%")
                else:
                    if target_cumulative == 0:
                        pct_done = 100.0
                    else:
                        pct_done = min(round((count_cumulative / target_cumulative) * 100, 2), 100)

                row[f"% Work Done against Target-Till {m}"] = f"{pct_done}%"
                
                # Use hardcoded values for target achieved text if available
                if activity in T7_HARDCODED_VALUES and m in T7_HARDCODED_VALUES[activity]:
                    hardcoded_completed = T7_HARDCODED_VALUES[activity][m]["completed_count"]
                    hardcoded_target = T7_HARDCODED_VALUES[activity][m]["target_count"]
                    month_unit = "Flats"
                    row[f"Target achieved in {m}"] = f"{hardcoded_completed} {month_unit} out of {hardcoded_target} planned"
                else:
                    month_target, month_unit = t7_targets[activity][m]
                    count_in_month = activity_counts[activity][m]
                    row[f"Target achieved in {m}"] = f"{count_in_month} {month_unit} out of {int(month_target)} planned"
            else:
                row[f"% Work Done against Target-Till {m}"] = ""
                row[f"Target achieved in {m}"] = ""

        # Calculate weighted delay for last processed month
        if prev_months:
            last_month = prev_months[-1]
            pct_str = row.get(f"% Work Done against Target-Till {last_month}", "0%").replace("%", "")
            try:
                pct_val = float(pct_str)
                row["Weighted Delay against Targets"] = f"{round((pct_val * weightage) / 100, 2)}%"
            except ValueError:
                row["Weighted Delay against Targets"] = ""

        # Build target text dynamically (only tracker year)
        total_target = 0
        target_parts = []
        for idx, (month_num, year) in enumerate(MONTHS_DATA):
            if year == tracker_year:
                month = MONTHS[idx]
                month_target = int(t7_targets[activity][month][0])
                total_target += month_target
                if month_target > 0:
                    unit = t7_targets[activity][month][1]
                    target_parts.append(f"{month_target} {unit}-{month}")
        
        unit = t7_targets[activity][MONTHS[0]][1] if total_target > 0 else ""
        row["Target Till August"] = f"{int(total_target)} {unit} ({', '.join(target_parts)})" if target_parts else f"0 {unit}"
        
        progress_data.append(row)

    # Create dynamic column list
    all_cols = ["Milestone", "Activity", "Target Till August"]
    for month in MONTHS:
        all_cols.append(f"% Work Done against Target-Till {month}")
    all_cols.extend(["Weightage", "Weighted Delay against Targets"])
    for month in MONTHS:
        all_cols.append(f"Target achieved in {month}")
    all_cols.extend(["Total achieved", "Delay Reasons_June 2025"])

    df_t7 = pd.DataFrame(progress_data, columns=all_cols)
    return df_t7

def get_green3_targets_and_progress(cos):
    logger.info("=== STARTING GREEN 3 PROCESSING ===")

    raw = download_file_bytes(cos, GREEN3_TRACKER_KEY)
    wb = load_workbook(filename=BytesIO(raw), data_only=True)

    sheet_names = wb.sheetnames
    sheet = wb.active
    if len(sheet_names) > 1:
        for name in sheet_names:
            if any(k in name.lower() for k in ["progress", "track", "work", "green"]):
                sheet = wb[name]
                break

    # Dynamic delay reasons label
    delay_reasons_label = f"Delay Reasons_{TRACKER_DATE.strftime('%B %Y')}"

    # Define external activities (linked to tracker % Complete column L)
    green3_activities = {
        "June": [{"activity": "Path Way Area - GSB", "target": "100%"}],
        "July": [{"activity": "Water Proofing - Water Body & Gazebo", "target": "100%"}],
        "August": [{"activity": "Stone Work - Water Body & Gazebo", "target": "100%"}],
    }

    # Build dataframe rows
    progress_data = []
    prev_months = get_previous_months()

    for month in MONTHS:
        for i, act in enumerate(green3_activities.get(month, [])):
            row = {
                "Milestone": f"Milestone-{i+1:02d}",
                "Activity": act["activity"],
                "Target": f"{act['target']} in {month}",
                "Weightage": 100,
                "Weighted Delay against Targets": "",
                "Total achieved": "",
                delay_reasons_label: "",
            }

            # init % cols
            for m in MONTHS:
                row[f"% Work Done against Target-Till {m}"] = ""
                row[f"Target achieved in {m}"] = ""

            if month in prev_months:
                # Find activity in column C and read % Complete from column L
                found_percent = 0
                for r in range(1, sheet.max_row + 1):
                    activity_text = str(sheet.cell(r, 3).value or "").strip().lower()
                    if activity_text and act["activity"].split("-")[-1].strip().lower() in activity_text:
                        val = sheet.cell(r, 12).value  # column L
                        try:
                            if isinstance(val, str):
                                val = float(val.replace("%", "").strip())
                            found_percent = float(val)
                        except Exception:
                            found_percent = 0
                        break

                row[f"% Work Done against Target-Till {month}"] = f"{found_percent}%"
                row[f"Target achieved in {month}"] = f"{found_percent}% completed" if found_percent else "Not started"
                row["Weighted Delay against Targets"] = f"{round((found_percent * row['Weightage']) / 100, 2)}%"

            progress_data.append(row)

    # Build dataframe
    all_cols = ["Milestone", "Activity", "Target"]
    for m in MONTHS:
        all_cols.append(f"% Work Done against Target-Till {m}")
    all_cols.extend(["Weightage", "Weighted Delay against Targets"])
    for m in MONTHS:
        all_cols.append(f"Target achieved in {m}")
    all_cols.extend(["Total achieved", delay_reasons_label])

    df_green3 = pd.DataFrame(progress_data, columns=all_cols)
    logger.info(f"Green 3 DataFrame created with {len(df_green3)} rows")
    return df_green3

def get_green3_targets_and_progress(cos):
    logger.info("=== STARTING GREEN 3 PROCESSING ===")

    # Load tracker
    raw = download_file_bytes(cos, GREEN3_TRACKER_KEY)
    wb = load_workbook(filename=BytesIO(raw), data_only=True)

    # Select correct sheet
    sheet_names = wb.sheetnames
    sheet = wb.active
    if len(sheet_names) > 1:
        for name in sheet_names:
            if any(k in name.lower() for k in ["progress", "track", "work", "green"]):
                sheet = wb[name]
                break

    # Dynamic delay reasons label
    delay_reasons_label = f"Delay Reasons_{TRACKER_DATE.strftime('%B %Y')}"

    # Define activities (later can pull from KRA dynamically)
    green3_activities = {
        "June": [{"activity": "Path Way Area - GSB", "target": "100%"}],
        "July": [{"activity": "Water Proofing - Water Body & Gazebo", "target": "100%"}],
        "August": [{"activity": "Stone Work - Water Body & Gazebo", "target": "100%"}],
    }

    # Store last known progress (carry forward logic)
    last_known_progress = {}

    progress_data = []
    prev_months = get_previous_months()

    # Loop over months
    for month in MONTHS:
        for i, act in enumerate(green3_activities.get(month, [])):
            activity_name = act["activity"]

            row = {
                "Milestone": f"Milestone-{i+1:02d}",
                "Activity": activity_name,
                "Target": f"{act['target']} in {month}",
                "Weightage": 100,
                "Weighted Delay against Targets": "",
                "Total achieved": "",
                f"Delay Reasons_{TRACKER_DATE.strftime('%B %Y')}": ""
            }

            # init % cols
            for m in MONTHS:
                row[f"% Work Done against Target-Till {m}"] = ""
                row[f"Target achieved in {m}"] = ""

            found_percent = None

            if month in prev_months:
                # Search activity in tracker (col C) and % complete in col L
                for r in range(1, sheet.max_row + 1):
                    activity_text = str(sheet.cell(r, 3).value or "").strip().lower()
                    if activity_text and activity_name.split("-")[-1].strip().lower() in activity_text:
                        val = sheet.cell(r, 12).value  # col L
                        try:
                            if isinstance(val, str):
                                val = float(val.replace("%", "").strip())
                            if isinstance(val, (int, float)):
                                val = float(val)
                                # Normalize decimal fractions (0.8 → 80, 1.0 → 100)
                                if 0 < val <= 1:
                                    val *= 100
                            found_percent = val
                        except Exception:
                            found_percent = 0
                        break

                if found_percent is None:
                    # No new update → carry forward last known progress
                    found_percent = last_known_progress.get(activity_name, 0)

                # If completed, keep it 100 moving forward
                if found_percent >= 100:
                    last_known_progress[activity_name] = 100
                else:
                    last_known_progress[activity_name] = found_percent

                # Display formatting (whole numbers)
                found_percent_display = int(round(found_percent))

                row[f"% Work Done against Target-Till {month}"] = f"{found_percent_display}%"
                row[f"Target achieved in {month}"] = (
                    f"{found_percent_display}% completed" if found_percent_display > 0 else "Not started"
                )
                row["Weighted Delay against Targets"] = f"{round((found_percent_display * row['Weightage']) / 100, 2)}%"

            progress_data.append(row)

    # Build dataframe
    all_cols = ["Milestone", "Activity", "Target"]
    for m in MONTHS:
        all_cols.append(f"% Work Done against Target-Till {m}")
    all_cols.extend(["Weightage", "Weighted Delay against Targets"])
    for m in MONTHS:
        all_cols.append(f"Target achieved in {m}")
    all_cols.extend(["Total achieved", delay_reasons_label])

    df_green3 = pd.DataFrame(progress_data, columns=all_cols)
    logger.info(f"✅ Green 3 DataFrame created with {len(df_green3)} rows")
    return df_green3

def get_green3_targets_and_progress(cos):
    logger.info("=== STARTING GREEN 3 PROCESSING ===")

    # Load tracker
    raw = download_file_bytes(cos, GREEN3_TRACKER_KEY)
    wb = load_workbook(filename=BytesIO(raw), data_only=True)

    # Select correct sheet
    sheet_names = wb.sheetnames
    sheet = wb.active
    if len(sheet_names) > 1:
        for name in sheet_names:
            if any(k in name.lower() for k in ["progress", "track", "work", "green"]):
                sheet = wb[name]
                break

    # Dynamic delay reasons label
    delay_reasons_label = f"Delay Reasons_{TRACKER_DATE.strftime('%B %Y')}"

    # Define activities (from KRA - simplified here)
    green3_activities = {
        "June": [{"activity": "Path Way Area - GSB", "target": "100%"}],
        "July": [{"activity": "Water Proofing - Water Body & Gazebo", "target": "100%"}],
        "August": [{"activity": "Stone Work - Water Body & Gazebo", "target": "100%"}],
    }

    # --- Determine which month(s) to keep ---
    month_lookup = {m.lower(): m for m in MONTHS}
    current_month = TRACKER_DATE.strftime("%B")
    current_month_lower = current_month.lower()

    if current_month_lower not in month_lookup:
        logger.warning(f"⚠️ Current month '{current_month}' not in MONTHS list: {MONTHS}")
        months_to_include = [MONTHS[-1]]  # fallback: keep last
    else:
        current_index = MONTHS.index(month_lookup[current_month_lower])
        if current_index > 0:
            months_to_include = [MONTHS[current_index - 1]]
        else:
            months_to_include = [MONTHS[current_index]]

    logger.info(f"➡️ Including only these months for Green3: {months_to_include}")

    # --- Carry-forward progress ---
    last_known_progress = {}
    progress_data = []

    for month in months_to_include:
        for i, act in enumerate(green3_activities.get(month, [])):
            activity_name = act["activity"]

            row = {
                "Milestone": f"Milestone-{i+1:02d}",
                "Activity": activity_name,
                "Target": f"{act['target']} in {month}",
                "Weightage": 100,
                "Weighted Delay against Targets": "",
                "Total achieved": "",
                delay_reasons_label: ""
            }

            # init only kept month cols
            row[f"% Work Done against Target-Till {month}"] = ""
            row[f"Target achieved in {month}"] = ""

            # --- Search progress from tracker ---
            found_percent = None
            for r in range(1, sheet.max_row + 1):
                activity_text = str(sheet.cell(r, 3).value or "").strip().lower()
                if activity_text and activity_name.split("-")[-1].strip().lower() in activity_text:
                    val = sheet.cell(r, 12).value  # column L = % Complete
                    try:
                        if isinstance(val, str):
                            val = float(val.replace("%", "").strip())
                        if isinstance(val, (int, float)):
                            if 0 < val <= 1:  # normalize 0.8 → 80
                                val *= 100
                            found_percent = float(val)
                    except Exception:
                        found_percent = 0
                    break

            if found_percent is None:
                found_percent = last_known_progress.get(activity_name, 0)

            # lock at 100 if completed
            last_known_progress[activity_name] = min(found_percent, 100)

            # --- Fill row ---
            found_percent_display = int(round(found_percent))
            row[f"% Work Done against Target-Till {month}"] = f"{found_percent_display}%"
            row[f"Target achieved in {month}"] = (
                f"{found_percent_display}% completed" if found_percent_display > 0 else "Not started"
            )
            row["Weighted Delay against Targets"] = f"{round((found_percent_display * row['Weightage']) / 100, 2)}%"

            progress_data.append(row)

    # --- Build dataframe with only required months ---
    all_cols = ["Milestone", "Activity", "Target"]
    for m in months_to_include:
        all_cols.append(f"% Work Done against Target-Till {m}")
    all_cols.extend(["Weightage", "Weighted Delay against Targets"])
    for m in months_to_include:
        all_cols.append(f"Target achieved in {m}")
    all_cols.extend(["Total achieved", delay_reasons_label])

    df_green3 = pd.DataFrame(progress_data, columns=all_cols)
    logger.info(f"✅ Green 3 DataFrame created with {len(df_green3)} rows, months={months_to_include}")
    return df_green3

def write_excel_report(df_t6, df_t5, df_t7, df_green3, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Time Delivery Milestones"

    current_date = datetime.now().strftime("%d-%m-%Y")
    ws.append(["Veridia Time Delivery Milestones Report"])
    ws.append([f"Report Generated on: {current_date}"])
    ws.append([])

    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    grey = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    bold_font = Font(bold=True)
    normal_font = Font(bold=False)
    title_font = Font(bold=True, size=14)
    date_font = Font(bold=False, size=10, color="666666")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="000000")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    max_cols = max(len(df_t6.columns), len(df_t5.columns), len(df_t7.columns), len(df_green3.columns))
    
    ws.merge_cells(f'A1:{get_column_letter(max_cols)}1')
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    ws['A1'].fill = grey
    
    ws.merge_cells(f'A2:{get_column_letter(max_cols)}2')
    ws['A2'].font = date_font
    ws['A2'].alignment = center_align

    section_title_rows = set()
    total_delay_rows = set()

    def append_df_block(title, df, total_delay_label):
        start_col = 1
        end_col = len(df.columns)

        ws.append([title])
        title_row = ws.max_row
        section_title_rows.add(title_row)
        ws.merge_cells(start_row=title_row, start_column=start_col,
                       end_row=title_row, end_column=end_col)
        for cell in ws[title_row]:
            cell.fill = grey
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = border

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        header_row = title_row + 1
        body_start = header_row + 1
        body_end = ws.max_row

        for cell in ws[header_row]:
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = border

        for r in range(body_start, body_end + 1):
            for cell in ws[r]:
                cell.font = normal_font
                cell.alignment = left_align if cell.col_idx in (1, 2) else center_align
                cell.border = border

        try:
            total_delay = sum(float(str(v).strip('%')) for v in df["Weighted Delay against Targets"] if v)
        except Exception:
            total_delay = 0

        weighted_delay_col_idx = None
        for idx, col_name in enumerate(df.columns, start=1):
            if col_name == "Weighted Delay against Targets":
                weighted_delay_col_idx = idx
                break

        total_row_data = [""] * end_col
        if weighted_delay_col_idx:
            total_row_data[weighted_delay_col_idx - 1] = f"{round(total_delay, 2)}%"
            total_row_data[0] = total_delay_label

        ws.append(total_row_data)
        delay_row = ws.max_row
        total_delay_rows.add(delay_row)
        
        for idx, cell in enumerate(ws[delay_row], start=1):
            cell.font = bold_font
            cell.fill = yellow
            if idx == 1:
                cell.alignment = left_align
            elif idx == weighted_delay_col_idx:
                cell.alignment = center_align
            else:
                cell.alignment = center_align
            cell.border = border

        return title_row, delay_row

    append_df_block("Tower 6 Progress Against Milestones", df_t6, "Total Delay Tower 6")
    append_df_block("Tower 5 Progress Against Milestones", df_t5, "Total Delay Tower 5")
    append_df_block("Tower 7 Progress Against Milestones", df_t7, "Total Delay Tower 7")
    append_df_block("External Development (Green 3) Progress Against Milestones (Structure Work)", df_green3, "Total Delay ED")

    for col in ws.columns:
        max_len = 0
        for cell in col:
            text = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(text.split("\n")[0]))
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 22

    wb.save(filename)

def get_unique_filename(base_name):
    """
    If file exists, append (1), (2), etc.
    """
    if not os.path.exists(base_name):
        return base_name

    name, ext = os.path.splitext(base_name)
    counter = 1
    new_name = f"{name}({counter}){ext}"
    while os.path.exists(new_name):
        counter += 1
        new_name = f"{name}({counter}){ext}"
    return new_name

def get_unique_filename(base_name):
    """
    If file exists, append (1), (2), etc.
    """
    if not os.path.exists(base_name):
        return base_name

    name, ext = os.path.splitext(base_name)
    counter = 1
    new_name = f"{name}({counter}){ext}"
    while os.path.exists(new_name):
        counter += 1
        new_name = f"{name}({counter}){ext}"
    return new_name


def main():
    logger.info("=== STARTING VERIDIA REPORT WITH DYNAMIC MONTHS AND YEAR TRACKING ===")
    
    cos = init_cos()
    
    # Get latest tracker paths and setup dynamic months
    get_latest_tracker_paths(cos)
    
    # Verify that we have all the required tracker paths
    if not all([T6_TRACKER_KEY, T5_TRACKER_KEY, T7_TRACKER_KEY, GREEN3_TRACKER_KEY]):
        logger.error("❌ Failed to find all required tracker files")
        return
    
    logger.info(f"Using months: {MONTHS}")
    logger.info(f"With years: {[f'{m} {y}' for (_, y), m in zip(MONTHS_DATA, MONTHS)]}")
    logger.info(f"Processing months: {PROCESSING_MONTHS}")
    
    # Rest of the processing
    targets_t6 = get_slab_targets_fixed_cells(cos)
    raw_tracker_t6 = download_file_bytes(cos, T6_TRACKER_KEY)
    wb_tracker_t6 = load_workbook(filename=BytesIO(raw_tracker_t6), data_only=True)
    completed_t6 = count_tower6_completed(wb_tracker_t6)
    df_t6 = build_t6_milestone_dataframe(targets_t6, completed_t6)
    df_t5 = get_t5_targets_and_progress(cos)
    df_t7 = get_t7_targets_and_progress(cos)
    df_green3 = get_green3_targets_and_progress(cos)
    
    # Create base filename
    base_filename = f"Veridia_Time_Delivery_Milestone_Report ({datetime.now():%Y-%m-%d}).xlsx"
    filename = get_unique_filename(base_filename)

    write_excel_report(df_t6, df_t5, df_t7, df_green3, filename)
    
    logger.info("=== REPORT GENERATION COMPLETE ===")
    logger.info(f"Report saved as: {filename}")


if __name__ == "__main__":
    main()
