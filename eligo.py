
import os
import re
import logging
from io import BytesIO
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from dotenv import load_dotenv
import ibm_boto3
from ibm_botocore.client import Config

# ---------------------------------------------------------------------------
# CONFIG / CONSTANTS
# ---------------------------------------------------------------------------
load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

COS_API_KEY = os.getenv("COS_API_KEY")
COS_CRN = os.getenv("COS_SERVICE_INSTANCE_CRN")
COS_ENDPOINT = os.getenv("COS_ENDPOINT")
BUCKET = os.getenv("COS_BUCKET_NAME")
ELIGO_KRA_KEY = os.getenv("KRA_FILE_PATH")  # KRA file path in COS (single master file)

ELIGO_STRUCTURE_KEY = None
ELIGO_TG_FINISHING_KEY = None
ELIGO_TH_FINISHING_KEY = None

# months will be stored as "June 2025", "July 2025", ...
MONTHS = []
MONTHS_DATA = []  # list of tuples (month_num, year)
TRACKER_DATE = None 
PROCESSING_MONTHS = []

GREEN_HEX = "FF92D050"
ROWS_TO_BOLD = {1, 5, 12, 19}
TOWER_G_ANTICIPATED_COLS = ["N", "R", "V"]
TOWER_H_ANTICIPATED_COLS = ["AB", "AF", "AJ", "AN", "AR", "AV", "AZ"]

# will be filled dynamically from KRA workbook
TOWER_G_ACTIVITIES = []
TOWER_H_ACTIVITIES = []
TOWER_G_STRUCTURE_TARGETS = {}
TOWER_H_STRUCTURE_TARGETS = {}
TOWER_G_FINISHING_TARGETS = {}
TOWER_H_FINISHING_TARGETS = {}
ELIGO_HARDCODED_VALUES = {}

# ---------------------------------------------------------------------------
# HELPER FUNCTIONS
# ---------------------------------------------------------------------------
def init_cos():
    return ibm_boto3.client(
        "s3",
        ibm_api_key_id=COS_API_KEY,
        ibm_service_instance_id=COS_CRN,
        config=Config(signature_version="oauth"),
        endpoint_url=COS_ENDPOINT,
    )

def download_file_bytes(cos, key):
    if not key:
        raise ValueError("File key cannot be None or empty")
    obj = cos.get_object(Bucket=BUCKET, Key=key)
    return obj["Body"].read()

def list_files_in_folder(cos, folder_prefix):
    try:
        response = cos.list_objects_v2(Bucket=BUCKET, Prefix=folder_prefix)
        files = []
        if "Contents" in response:
            for obj in response["Contents"]:
                if not obj["Key"].endswith("/"):
                    files.append(obj["Key"])
        return files
    except Exception as e:
        logger.error(f"Error listing files in folder {folder_prefix}: {e}")
        return []

def extract_date_from_filename(filename):
    pattern = r"\((\d{2}-\d{2}-\d{4})\)"
    match = re.search(pattern, filename)
    if match:
        date_str = match.group(1)
        try:
            return datetime.strptime(date_str, "%d-%m-%Y")
        except ValueError:
            logger.warning(f"Could not parse date {date_str} from filename {filename}")
            return None
    return None

def get_month_name(month_num):
    months = {
        1: "January",2: "February",3: "March",4: "April",5: "May",6: "June",
        7: "July",8: "August",9: "September",10: "October",11: "November",12: "December"
    }
    return months.get(month_num, "Unknown")

def extract_number(cell_value):
    if not cell_value or cell_value == "-":
        return 0.0
    match = re.search(r"(\d+)", str(cell_value))
    return float(match.group(1)) if match else 0.0

def get_previous_months():
    return PROCESSING_MONTHS

# ---------------------------------------------------------------------------
# DYNAMIC KRA TARGETS LOADER
# ---------------------------------------------------------------------------

def _col_matches_month(col_name, month_name):
    """Return True if the column name contains the month token (case-insensitive)."""
    if not isinstance(col_name, str):
        return False
    return month_name.lower().split()[0] in col_name.lower()

def load_targets_from_kra(cos):
    """
    Read the KRA file from COS and populate:
      - TOWER_G_FINISHING_TARGETS
      - TOWER_H_FINISHING_TARGETS
      - TOWER_G_STRUCTURE_TARGETS
      - TOWER_H_STRUCTURE_TARGETS
    Matching is best-effort:
      - find tables with an 'Activity' column => finishing tables
      - find tables with columns matching month names => structure tables
    Falls back to hardcoded lists/values if detection fails.
    """
    global TOWER_G_ACTIVITIES, TOWER_H_ACTIVITIES
    global TOWER_G_FINISHING_TARGETS, TOWER_H_FINISHING_TARGETS
    global TOWER_G_STRUCTURE_TARGETS, TOWER_H_STRUCTURE_TARGETS

    # safe default/hardcoded fallback (kept from original)
    fallback_g_acts = [
        "Water Proofing Works", "HVAC 2nd Fix",
        "Wall tiling (Toilet & Kitchen)", "Floor tiling",
    ]
    fallback_h_acts = ["HVAC 1st Fix", "POP punning (Major area)", "Wall Tiling", "Floor Tiling"]

    try:
        raw = download_file_bytes(cos, ELIGO_KRA_KEY)
        # read all sheets into pandas
        sheets = pd.read_excel(BytesIO(raw), sheet_name=None, engine="openpyxl")
    except Exception as e:
        logger.warning(f"Could not read KRA file from COS: {e}. Using fallback activities/targets.")
        # fallback: preserve original hardcoded targets
        TOWER_G_ACTIVITIES = fallback_g_acts[:]
        TOWER_H_ACTIVITIES = fallback_h_acts[:]
        # Create fallback finishing targets: equal to zeros or original presets
        setup_tower_targets_fallback()
        return

    # helpers
    month_tokens = [get_month_name(m) for m, _ in MONTHS_DATA]  # e.g., ["June", "July", "August"]
    month_labels = [f"{get_month_name(m)} {y}" for m, y in MONTHS_DATA]  # e.g., ["June 2025",...]

    found_g_fin = False
    found_h_fin = False
    # We'll collect finishing tables we detect, then assign to G/H (best-effort)
    finishing_tables = []

    structure_tables = []

    for sheet_name, df in sheets.items():
        # normalize columns to strings
        df_cols = [str(c).strip() for c in df.columns]
        df.columns = df_cols

        # detect finishing-like table (has 'Activity' column)
        activity_cols = [c for c in df_cols if c and "activity" in c.lower()]
        if activity_cols:
            # treat this df as a finishing table
            finishing_tables.append((sheet_name, df.copy()))
            continue

        # detect structure-like table: look for any month token in columns
        if any(any(token.lower() in (str(col).lower()) for token in month_tokens) for col in df_cols):
            structure_tables.append((sheet_name, df.copy()))
            continue

    # Assign finishing tables to Tower G and H

    for (sheet_name, df) in finishing_tables:
        sheet_lower = sheet_name.lower()
        assigned = None
        if "tower g" in sheet_lower or "pour g" in sheet_lower or "tower_g" in sheet_lower or ("g " in sheet_lower and "h " not in sheet_lower):
            assigned = "G"
        elif "tower h" in sheet_lower or "pour h" in sheet_lower or "tower_h" in sheet_lower:
            assigned = "H"
        else:
            # try to see if any column or first cell suggests G or H
            # fallback: assign first unassigned to G, second to H
            if not found_g_fin:
                assigned = "G"
            elif not found_h_fin:
                assigned = "H"
            else:
                assigned = "G"

        # extract activity col
        activity_col = next((c for c in df.columns if "activity" in c.lower()), df.columns[0])
        # build mapping month_label -> df column name
        month_col_map = {}
        for token, label in zip(month_tokens, month_labels):
            # find a df column that contains the token (case-insensitive)
            match = next((c for c in df.columns if token.lower() in c.lower()), None)
            if match:
                month_col_map[label] = match

        # If no month columns found, skip
        if not month_col_map:
            logger.debug(f"No month columns detected in sheet {sheet_name}; skipping finishing table extraction.")
            continue

        # iterate rows to build activities
        for _, row in df.iterrows():
            activity_val = row.get(activity_col, None)
            if not activity_val or (isinstance(activity_val, float) and pd.isna(activity_val)):
                continue
            activity = str(activity_val).strip()
            if not activity:
                continue
            targets_for_activity = {}
            for label, colname in month_col_map.items():
                val = row.get(colname, 0)
                try:
                    if pd.isna(val):
                        val = 0
                except:
                    pass
                try:
                    val_num = int(float(val))
                except Exception:
                    # if value not numeric, fallback to extracted number or 0
                    val_num = int(extract_number(val))
                targets_for_activity[label] = val_num

            if assigned == "G":
                TOWER_G_FINISHING_TARGETS[activity] = targets_for_activity
                found_g_fin = True
                if activity not in TOWER_G_ACTIVITIES:
                    TOWER_G_ACTIVITIES.append(activity)
            else:
                TOWER_H_FINISHING_TARGETS[activity] = targets_for_activity
                found_h_fin = True
                if activity not in TOWER_H_ACTIVITIES:
                    TOWER_H_ACTIVITIES.append(activity)

    # If no finishing tables detected, fallback to previous hardcoded behavior
    if not TOWER_G_FINISHING_TARGETS and not TOWER_H_FINISHING_TARGETS:
        logger.info("No finishing tables detected in KRA; using fallback hardcoded finishing activities/targets.")
        TOWER_G_ACTIVITIES = fallback_g_acts[:]
        TOWER_H_ACTIVITIES = fallback_h_acts[:]
        setup_tower_targets_fallback()  # fills FINISHING_TARGETS with previous preset values

    # Now detect structure tables and assign sums
    # We'll find any structure-like df and extract monthly sums (best-effort)
    # If multiple structure tables exist, assign first to Tower G and second to Tower H
    found_g_struct = False
    found_h_struct = False
    for (sheet_name, df) in structure_tables:
        df_cols = df.columns
        # build month column map same way
        month_col_map = {}
        for token, label in zip(month_tokens, month_labels):
            match = next((c for c in df_cols if token.lower() in str(c).lower()), None)
            if match:
                month_col_map[label] = match
        if not month_col_map:
            continue

        # compute column sums (ignore non-numeric gracefully)
        sums = {}
        for label, colname in month_col_map.items():
            try:
                colvals = pd.to_numeric(df[colname], errors="coerce").fillna(0)
                sums[label] = int(colvals.sum())
            except Exception:
                sums[label] = 0

        # assign to tower
        assigned = None
        lower = sheet_name.lower()
        if "tower g" in lower or "tower_g" in lower or ("g " in lower and "h " not in lower):
            assigned = "G"
        elif "tower h" in lower or "tower_h" in lower:
            assigned = "H"
        else:
            if not found_g_struct:
                assigned = "G"
            elif not found_h_struct:
                assigned = "H"
            else:
                assigned = "G"

        if assigned == "G":
            for label, val in sums.items():
                TOWER_G_STRUCTURE_TARGETS[label] = val
            found_g_struct = True
        else:
            for label, val in sums.items():
                TOWER_H_STRUCTURE_TARGETS[label] = val
            found_h_struct = True

    # Final fallback if structure targets remain empty
    if not TOWER_G_STRUCTURE_TARGETS:
        # default placeholder to keep code working
        TOWER_G_STRUCTURE_TARGETS = {label: 1 for label in month_labels}
    if not TOWER_H_STRUCTURE_TARGETS:
        TOWER_H_STRUCTURE_TARGETS = {label: 3 for label in month_labels}

    logger.info("Loaded KRA targets from COS (best-effort).")
    logger.debug(f"TOWER_G_ACTIVITIES: {TOWER_G_ACTIVITIES}")
    logger.debug(f"TOWER_H_ACTIVITIES: {TOWER_H_ACTIVITIES}")
    logger.debug(f"TOWER_G_FINISHING_TARGETS keys: {list(TOWER_G_FINISHING_TARGETS.keys())}")
    logger.debug(f"TOWER_H_FINISHING_TARGETS keys: {list(TOWER_H_FINISHING_TARGETS.keys())}")
    logger.debug(f"TOWER_G_STRUCTURE_TARGETS: {TOWER_G_STRUCTURE_TARGETS}")
    logger.debug(f"TOWER_H_STRUCTURE_TARGETS: {TOWER_H_STRUCTURE_TARGETS}")

def setup_tower_targets_fallback():
    """Populate finishing targets using your original hardcoded templates (fallback)."""
    global TOWER_G_STRUCTURE_TARGETS, TOWER_H_STRUCTURE_TARGETS
    global TOWER_G_FINISHING_TARGETS, TOWER_H_FINISHING_TARGETS

    # structure fallback
    TOWER_G_STRUCTURE_TARGETS = {month: 1 for month in MONTHS}
    if len(MONTHS) >= 3:
        TOWER_H_STRUCTURE_TARGETS = {MONTHS[0]: 3, MONTHS[1]: 3, MONTHS[2]: 4}
    else:
        TOWER_H_STRUCTURE_TARGETS = {month: 3 for month in MONTHS}

    # finishing fallback: reuse earlier hardcoded templates (safe)
    g_presets = {
        "Water Proofing Works": [20, 24, 19],
        "HVAC 2nd Fix": [41, 16, 0],
        "Wall tiling (Toilet & Kitchen)": [0, 1, 43],
        "Floor tiling": [0, 0, 32],
    }
    h_presets = {
        "HVAC 1st Fix": [16, 0, 0],
        "POP punning (Major area)": [13, 8, 8],
        "Wall Tiling": [8, 39, 9],
        "Floor Tiling": [14, 39, 9],
    }

    TOWER_G_FINISHING_TARGETS = {}
    for act, vals in g_presets.items():
        TOWER_G_FINISHING_TARGETS[act] = {MONTHS[i]: vals[i] if i < len(vals) else 0 for i in range(len(MONTHS))}
    TOWER_H_FINISHING_TARGETS = {}
    for act, vals in h_presets.items():
        TOWER_H_FINISHING_TARGETS[act] = {MONTHS[i]: vals[i] if i < len(vals) else 0 for i in range(len(MONTHS))}
    # set activity lists
    global TOWER_G_ACTIVITIES, TOWER_H_ACTIVITIES
    TOWER_G_ACTIVITIES = list(TOWER_G_FINISHING_TARGETS.keys())
    TOWER_H_ACTIVITIES = list(TOWER_H_FINISHING_TARGETS.keys())

# ---------------------------------------------------------------------------
# SETUP FUNCTIONS (months + dynamic targets)
# ---------------------------------------------------------------------------
def setup_dynamic_months_and_targets(tracker_date):
    """
    Dynamically setup MONTHS, MONTHS_DATA, and PROCESSING_MONTHS based on tracker_date.
    - Handles year transitions.
    - Special case: if tracker month is September, include June, July, August.
    - Otherwise: include previous month, current month, next month (3 months window).
    """
    global MONTHS, MONTHS_DATA, TRACKER_DATE, PROCESSING_MONTHS

    TRACKER_DATE = tracker_date
    tracker_month = tracker_date.month
    tracker_year = tracker_date.year

    logger.info("=== DYNAMIC MONTH SETUP WITH YEAR TRACKING ===")
    logger.info(f"Tracker date: {tracker_date.strftime('%d-%m-%Y')} (Month: {tracker_month}, Year: {tracker_year})")

    MONTHS_DATA = []

    # Special case: if tracker is September, include June, July, August of same year
    if tracker_month == 9:
        for m in range(6, 9):  # 6=June, 7=July, 8=August
            MONTHS_DATA.append((m, tracker_year))
    else:
        # For other months: previous, current, next (3 months window)
        for i in range(-1, 2):  # prev, current, next
            month_num = tracker_month + i
            year = tracker_year
            if month_num < 1:
                month_num += 12
                year -= 1
            elif month_num > 12:
                month_num -= 12
                year += 1
            MONTHS_DATA.append((month_num, year))

    # Month labels with year
    MONTHS = [f"{get_month_name(m)} {y}" for m, y in MONTHS_DATA]
    logger.info(f"Generated MONTHS: {MONTHS}")

    # Processing months: only months **before tracker month**
    PROCESSING_MONTHS = []
    for month_name, (month_num, year) in zip(MONTHS, MONTHS_DATA):
        if (year < tracker_year) or (year == tracker_year and month_num < tracker_month):
            PROCESSING_MONTHS.append(month_name)
            logger.info(f"Including {month_name} {year} for processing")

    # Fallback if no month included
    if not PROCESSING_MONTHS and MONTHS:
        PROCESSING_MONTHS = [MONTHS[-1]]

    logger.info(f"PROCESSING_MONTHS: {PROCESSING_MONTHS}")

# ---------------------------------------------------------------------------
# get_latest_tracker_paths (unchanged, but kept here for full flow)
# ---------------------------------------------------------------------------
def get_latest_tracker_paths(cos):
    global ELIGO_STRUCTURE_KEY, ELIGO_TG_FINISHING_KEY, ELIGO_TH_FINISHING_KEY

    logger.info("=== FINDING LATEST ELIGO TRACKER FILES ===")

    eligo_files = list_files_in_folder(cos, "Eligo/")
    logger.info(f"Found {len(eligo_files)} files in Eligo folder")

    tracker_patterns = {
        "STRUCTURE_TRACKER": r"Structure Work Tracker.*\.xlsx$",
        "TG_FINISHING_TRACKER": r"Tower G Finishing Tracker.*\.xlsx$",
        "TH_FINISHING_TRACKER": r"Tower H Finishing Tracker.*\.xlsx$",
    }

    latest_trackers = {}
    latest_date = None

    for tracker_type, pattern in tracker_patterns.items():
        matching_files = []
        for file_path in eligo_files:
            filename = os.path.basename(file_path)
            if re.search(pattern, filename, re.IGNORECASE):
                file_date = extract_date_from_filename(filename)
                if file_date:
                    matching_files.append((file_path, file_date))
                    if latest_date is None or file_date > latest_date:
                        latest_date = file_date

        if matching_files:
            latest_file = max(matching_files, key=lambda x: x[1])
            latest_trackers[tracker_type] = latest_file[0]
            logger.info(f"✅ Latest {tracker_type}: {latest_file[0]}")
        else:
            latest_trackers[tracker_type] = None

    ELIGO_STRUCTURE_KEY = latest_trackers.get("STRUCTURE_TRACKER")
    ELIGO_TG_FINISHING_KEY = latest_trackers.get("TG_FINISHING_TRACKER")
    ELIGO_TH_FINISHING_KEY = latest_trackers.get("TH_FINISHING_TRACKER")

    # set months + processing months based on latest_date (fall back to now)
    if latest_date:
        setup_dynamic_months_and_targets(latest_date)
    else:
        setup_dynamic_months_and_targets(datetime.now())

    missing_trackers = [k for k, v in latest_trackers.items() if v is None]
    if missing_trackers:
        raise Exception(f"Could not find latest tracker files for: {missing_trackers}")

    return latest_trackers

# ---------------------------------------------------------------------------
# COUNTING FUNCTIONS (mostly unchanged - but use month-year labels)
# ---------------------------------------------------------------------------
def count_green_dates_in_month(wb, sheet_name, columns, month_name):
    if sheet_name not in wb.sheetnames:
        return 0

    month_idx = MONTHS.index(month_name) if month_name in MONTHS else -1
    if month_idx == -1:
        return 0

    month_num, target_year = MONTHS_DATA[month_idx]
    sheet = wb[sheet_name]
    count = 0

    for col_letter in columns:
        for row in range(4, sheet.max_row + 1):
            cell = sheet[f"{col_letter}{row}"]
            if cell.value:
                try:
                    cell_date = None
                    if isinstance(cell.value, datetime):
                        cell_date = cell.value
                    elif isinstance(cell.value, str):
                        cell_date = pd.to_datetime(cell.value, dayfirst=True, errors="coerce")

                    if pd.notna(cell_date) and cell_date.year == target_year and cell_date.month == month_num:
                        fill = cell.fill
                        color_code = getattr(fill, "start_color", None)
                        rgb = color_code.rgb if color_code else None
                        if fill.fill_type == "solid" and rgb == GREEN_HEX:
                            count += 1
                except Exception:
                    continue
    return count

def count_green_dates_in_month_fixed(wb, sheet_name, columns, month_name, start_row=5, end_row=12):
    if sheet_name not in wb.sheetnames:
        return 0

    month_idx = MONTHS.index(month_name) if month_name in MONTHS else -1
    if month_idx == -1:
        return 0

    month_num, target_year = MONTHS_DATA[month_idx]
    sheet = wb[sheet_name]
    count = 0

    for col_letter in columns:
        for row in range(start_row, end_row + 1):
            cell = sheet[f"{col_letter}{row}"]
            if cell.value:
                try:
                    cell_date = None
                    if isinstance(cell.value, datetime):
                        cell_date = cell.value
                    elif isinstance(cell.value, str):
                        for date_format in ["%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"]:
                            try:
                                cell_date = datetime.strptime(str(cell.value), date_format)
                                break
                            except:
                                continue
                        if not cell_date:
                            cell_date = pd.to_datetime(cell.value, dayfirst=True, errors="coerce")

                    if pd.notna(cell_date) and cell_date.year == target_year and cell_date.month == month_num:
                        fill = cell.fill
                        color_code = getattr(fill, "start_color", None)
                        rgb = color_code.rgb if color_code else None
                        green_colors = [GREEN_HEX, "92D050", "FF92D050", "00FF92D050"]
                        if fill.fill_type == "solid" and rgb in green_colors:
                            count += 1
                except Exception:
                    continue
    return count

def count_completed_activities_by_month_fixed(wb, sheet_names, activity_name, year, month):
    """Count individual flats/units that completed the activity in the specified month (keeps original behavior)."""
    if activity_name in ELIGO_HARDCODED_VALUES:
        month_name = get_month_name(month)
        if month_name in ELIGO_HARDCODED_VALUES[activity_name]:
            hardcoded_count = ELIGO_HARDCODED_VALUES[activity_name][month_name]["completed_count"]
            logger.info(f"HARDCODED: Returning {hardcoded_count} for {activity_name} in {month_name} {year}")
            return hardcoded_count

    count = 0
    month_name = get_month_name(month)
    logger.info(f"Counting '{activity_name}' completions for {month_name} {year}")

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            continue
        try:
            sheet = wb[sheet_name]
            sheet_count = 0

            activity_col = 7  # Column G - Activity Name
            finish_col = 12   # Column L - Actual Finish

            for row_num in range(2, min(sheet.max_row + 1, 2000)):
                activity_cell = sheet.cell(row=row_num, column=activity_col)
                finish_cell = sheet.cell(row=row_num, column=finish_col)

                if activity_cell.value and finish_cell.value:
                    activity_text = str(activity_cell.value).strip()

                    if (activity_text.lower() == activity_name.lower() or
                        activity_name.lower() in activity_text.lower() or
                        activity_text.lower() in activity_name.lower()):
                        try:
                            finish_date = None
                            if isinstance(finish_cell.value, datetime):
                                finish_date = finish_cell.value
                            elif isinstance(finish_cell.value, str):
                                finish_date = pd.to_datetime(finish_cell.value, dayfirst=True, errors='coerce')

                            if pd.notna(finish_date) and finish_date.year == year and finish_date.month == month:
                                sheet_count += 1
                                logger.debug(f"  ✓ Sheet '{sheet_name}' Row {row_num}: '{activity_text}' completed on {finish_date.strftime('%d-%m-%Y')}")
                        except Exception as e:
                            logger.debug(f"Error processing row {row_num}: {e}")
                            continue

            if sheet_count > 0:
                logger.info(f"  Sheet '{sheet_name}': Found {sheet_count} completions")
            count += sheet_count

        except Exception as e:
            logger.warning(f"Error processing sheet {sheet_name}: {e}")
            continue

    logger.info(f"TOTAL completions for '{activity_name}' in {month_name} {year}: {count}")
    return count

# ---------------------------------------------------------------------------
# TOWER FUNCTIONS (structure & finishing) - updated to use month-year labels
# ---------------------------------------------------------------------------
def get_tower_g_structure_targets():
    return TOWER_G_STRUCTURE_TARGETS

def count_tower_g_completed(cos):
    raw = download_file_bytes(cos, ELIGO_STRUCTURE_KEY)
    wb = load_workbook(filename=BytesIO(raw), data_only=True)
    counts = {m: 0 for m in MONTHS}

    for month_name in PROCESSING_MONTHS:
        count = count_green_dates_in_month(wb, "Revised Baselines- 25 days SC", TOWER_G_ANTICIPATED_COLS, month_name)
        counts[month_name] = count

    logger.info(f"Tower G completed: {counts}")
    return counts

def build_tower_g_structure_dataframe(targets, completed):
    tracker_year = TRACKER_DATE.year
    prev_months = [m for (m, y) in zip(MONTHS, [y for _, y in MONTHS_DATA]) if y == tracker_year]
    weightage = 100.0

    def pct(m):
        if m in prev_months:
            month_idx = MONTHS.index(m)
            cum_done = sum(int(completed.get(MONTHS[i], 0)) for i in range(len(MONTHS)) if i <= month_idx and MONTHS_DATA[i][1] == tracker_year)
            cum_target = sum(int(targets.get(MONTHS[i], 0)) for i in range(len(MONTHS)) if i <= month_idx and MONTHS_DATA[i][1] == tracker_year)
            if cum_target == 0:
                return "0.0%"
            return f"{min(round((cum_done / cum_target) * 100, 2), 100)}%"
        return ""

    target_parts = [f"{int(targets.get(MONTHS[i], 0))} Pours-{MONTHS[i]}" for i, (_, y) in enumerate(MONTHS_DATA) if y == tracker_year and targets.get(MONTHS[i], 0) > 0]
    total_target = sum(int(targets.get(MONTHS[i], 0)) for i, (_, y) in enumerate(MONTHS_DATA) if y == tracker_year)

    last_label = TRACKER_DATE.strftime("%B %Y")
    row = {
        "Milestone": "Milestone-01",
        "Activity": "Pour Casting",
        "Target Till": (f"{total_target} Pours ({', '.join(target_parts)})" if target_parts else "0 Pours"),
        "Weightage": weightage,
        "Weighted Delay against Targets": "",
        "Total achieved": "",
        f"Delay Reasons_{last_label}": "",
    }

    for m in MONTHS:
        row[f"% Work Done against Target-Till {m}"] = pct(m)
        if m in prev_months:
            row[f"Target achieved in {m}"] = f"{int(completed.get(m, 0))} pour cast out of {int(targets.get(m, 0))} planned"
        else:
            row[f"Target achieved in {m}"] = ""

    if prev_months:
        try:
            pct_val = float(pct(prev_months[-1]).replace("%", ""))
            row["Weighted Delay against Targets"] = f"{round((pct_val * weightage) / 100, 2)}%"
        except:
            pass

    all_cols = ["Milestone", "Activity", "Target Till"] + [f"% Work Done against Target-Till {m}" for m in MONTHS] + ["Weightage", "Weighted Delay against Targets"] + [f"Target achieved in {m}" for m in MONTHS] + ["Total achieved", f"Delay Reasons_{last_label}"]
    return pd.DataFrame([row], columns=all_cols)

def get_tower_h_structure_targets():
    return TOWER_H_STRUCTURE_TARGETS

def count_tower_h_completed(cos):
    raw = download_file_bytes(cos, ELIGO_STRUCTURE_KEY)
    wb = load_workbook(filename=BytesIO(raw), data_only=True)
    counts = {m: 0 for m in MONTHS}

    for month_name in PROCESSING_MONTHS:
        count = count_green_dates_in_month_fixed(wb, "Revised Baselines- 25 days SC", TOWER_H_ANTICIPATED_COLS, month_name, 5, 12)
        counts[month_name] = count

    logger.info(f"Tower H completed: {counts}")
    return counts

def build_tower_h_structure_dataframe(targets, completed):
    tracker_year = TRACKER_DATE.year
    prev_months = [m for (m, y) in zip(MONTHS, [y for _, y in MONTHS_DATA]) if y == tracker_year]
    weightage = 100.0

    def pct(m):
        if m in prev_months:
            month_idx = MONTHS.index(m)
            cum_done = sum(int(completed.get(MONTHS[i], 0)) for i in range(len(MONTHS)) if i <= month_idx and MONTHS_DATA[i][1] == tracker_year)
            cum_target = sum(int(targets.get(MONTHS[i], 0)) for i in range(len(MONTHS)) if i <= month_idx and MONTHS_DATA[i][1] == tracker_year)
            if cum_target == 0:
                return "0.0%"
            return f"{min(round((cum_done / cum_target) * 100, 2), 100)}%"
        return ""

    target_parts = [f"{int(targets.get(MONTHS[i], 0))} Pours-{MONTHS[i]}" for i, (_, y) in enumerate(MONTHS_DATA) if y == tracker_year and targets.get(MONTHS[i], 0) > 0]
    total_target = sum(int(targets.get(MONTHS[i], 0)) for i, (_, y) in enumerate(MONTHS_DATA) if y == tracker_year)

    last_label = TRACKER_DATE.strftime("%B %Y")
    row = {
        "Milestone": "Milestone-01",
        "Activity": "Pour Casting",
        "Target Till": (f"{total_target} Pours ({', '.join(target_parts)})" if target_parts else "0 Pours"),
        "Weightage": weightage,
        "Weighted Delay against Targets": "",
        "Total achieved": "",
        f"Delay Reasons_{last_label}": "",
    }

    for m in MONTHS:
        row[f"% Work Done against Target-Till {m}"] = pct(m)
        if m in prev_months:
            row[f"Target achieved in {m}"] = f"{int(completed.get(m, 0))} pour cast out of {int(targets.get(m, 0))} planned"
        else:
            row[f"Target achieved in {m}"] = ""

    if prev_months:
        try:
            pct_val = float(pct(prev_months[-1]).replace("%", ""))
            row["Weighted Delay against Targets"] = f"{round((pct_val * weightage) / 100, 2)}%"
        except:
            pass

    all_cols = ["Milestone", "Activity", "Target Till"] + [f"% Work Done against Target-Till {m}" for m in MONTHS] + ["Weightage", "Weighted Delay against Targets"] + [f"Target achieved in {m}" for m in MONTHS] + ["Total achieved", f"Delay Reasons_{last_label}"]
    return pd.DataFrame([row], columns=all_cols)

# ---------------------------------------------------------------------------
# FINISHING: count and build (updated to use month-year parsing)
# ---------------------------------------------------------------------------
def get_tower_g_finishing_targets():
    return TOWER_G_FINISHING_TARGETS

def count_tower_g_finishing_completed(cos):
    raw = download_file_bytes(cos, ELIGO_TG_FINISHING_KEY)
    wb = load_workbook(filename=BytesIO(raw), data_only=True)
    target_sheets = ['Common Area', 'Pour G1', 'Pour G2', 'Pour G3']
    counts = {}
    logger.info("Starting Tower G finishing count...")

    for activity in TOWER_G_ACTIVITIES:
        counts[activity] = {m: 0 for m in MONTHS}
        for month_name in PROCESSING_MONTHS:
            try:
                # month_name is like "June 2025" -> parse to get numeric month and year
                dt = datetime.strptime(month_name, "%B %Y")
            except Exception:
                # fallback: if parsing fails ignore
                continue
            count = count_completed_activities_by_month_fixed(wb, target_sheets, activity, dt.year, dt.month)
            counts[activity][month_name] = count

    logger.info(f"Tower G Finishing completed by month: {counts}")
    return counts

def build_tower_g_finishing_dataframe(targets, completed):
    prev_months = get_previous_months()
    progress_data = []
    total_milestones = len(targets)
    weightage = round(100 / total_milestones, 2) if total_milestones else 0

    last_label = TRACKER_DATE.strftime("%B %Y")
    for i, activity in enumerate(list(targets.keys())):
        row = {
            "Milestone": f"Milestone-{i+1:02d}",
            "Activity": activity,
            "Weightage": weightage,
            "Weighted Delay against Targets": "",
            "Total achieved": "",
            f"Delay Reasons_{last_label}": "",
        }
        for m in MONTHS:
            if m in prev_months:
                # cumulative counts & targets up to m
                month_idx = MONTHS.index(m)
                count_cumulative = sum(completed.get(activity, {}).get(MONTHS[i], 0) for i in range(0, month_idx + 1))
                target_cumulative = sum(targets.get(activity, {}).get(MONTHS[i], 0) for i in range(0, month_idx + 1))
                if target_cumulative == 0:
                    pct_done = 100.0
                else:
                    pct_done = min(round((count_cumulative / target_cumulative) * 100, 2), 100)
                row[f"% Work Done against Target-Till {m}"] = f"{pct_done}%"
                month_target = targets.get(activity, {}).get(m, 0)
                count_in_month = completed.get(activity, {}).get(m, 0)
                if month_target == 0:
                    future_months = [fm for fm in MONTHS[MONTHS.index(m)+1:] if targets.get(activity, {}).get(fm, 0) > 0]
                    if future_months:
                        row[f"Target achieved in {m}"] = f"Planned for {' and '.join(future_months)}" if len(future_months) > 1 else f"Planned for {future_months[0]}"
                    else:
                        row[f"Target achieved in {m}"] = f"{count_in_month} Flats out of {int(month_target)} planned"
                else:
                    row[f"Target achieved in {m}"] = f"{count_in_month} Flats out of {int(month_target)} planned"
            else:
                row[f"% Work Done against Target-Till {m}"] = ""
                row[f"Target achieved in {m}"] = ""

        if prev_months:
            last_month = prev_months[-1]
            pct_str = row.get(f"% Work Done against Target-Till {last_month}", "0%").replace("%", "")
            try:
                pct_val = float(pct_str)
                row["Weighted Delay against Targets"] = f"{round((pct_val * weightage) / 100, 2)}%"
            except ValueError:
                row["Weighted Delay against Targets"] = ""

        total_target = sum(targets.get(activity, {}).get(month, 0) for month in MONTHS)
        target_parts = [f"{int(targets.get(activity, {}).get(month, 0))} Flats-{month}" for month in MONTHS if int(targets.get(activity, {}).get(month, 0)) > 0]
        row["Target Till"] = f"{int(total_target)} Flats ({', '.join(target_parts)})"

        progress_data.append(row)

    # build columns dynamically
    all_cols = ["Milestone", "Activity", "Target Till"]
    for month in MONTHS:
        all_cols.append(f"% Work Done against Target-Till {month}")
    all_cols.extend(["Weightage", "Weighted Delay against Targets"])
    for month in MONTHS:
        all_cols.append(f"Target achieved in {month}")
    all_cols.extend(["Total achieved", f"Delay Reasons_{last_label}"])

    df_tg_finishing = pd.DataFrame(progress_data)
    # ensure columns order (if missing columns, add them)
    for c in all_cols:
        if c not in df_tg_finishing.columns:
            df_tg_finishing[c] = ""
    df_tg_finishing = df_tg_finishing[all_cols]
    return df_tg_finishing

def get_tower_h_finishing_targets():
    return TOWER_H_FINISHING_TARGETS

def count_tower_h_finishing_completed(cos):
    raw = download_file_bytes(cos, ELIGO_TH_FINISHING_KEY)
    wb = load_workbook(filename=BytesIO(raw), data_only=True)
    target_sheets = ['Common Area', 'Pre-Construction Activities', 'Pour H1', 'Pour H2', 'Pour H3', 'Pour H4', 'Pour H5', 'Pour H6', 'Pour H7']
    counts = {}
    logger.info("Starting Tower H finishing count...")

    for activity in TOWER_H_ACTIVITIES:
        counts[activity] = {m: 0 for m in MONTHS}
        for month_name in PROCESSING_MONTHS:
            try:
                dt = datetime.strptime(month_name, "%B %Y")
            except Exception:
                continue
            count = count_completed_activities_by_month_fixed(wb, target_sheets, activity, dt.year, dt.month)
            counts[activity][month_name] = count

    logger.info(f"Tower H Finishing completed by month: {counts}")
    return counts

def build_tower_h_finishing_dataframe(targets, completed):
    prev_months = get_previous_months()
    progress_data = []
    total_milestones = len(targets)
    weightage = round(100 / total_milestones, 2) if total_milestones else 0
    last_label = TRACKER_DATE.strftime("%B %Y")

    for i, activity in enumerate(list(targets.keys())):
        row = {
            "Milestone": f"Milestone-{i+1:02d}",
            "Activity": activity,
            "Weightage": weightage,
            "Weighted Delay against Targets": "",
            "Total achieved": "",
            f"Delay Reasons_{last_label}": "",
        }
        for m in MONTHS:
            if m in prev_months:
                # check hardcoded override
                if activity in ELIGO_HARDCODED_VALUES and m in ELIGO_HARDCODED_VALUES[activity]:
                    pct_done = ELIGO_HARDCODED_VALUES[activity][m]["percentage"]
                    hardcoded_completed = ELIGO_HARDCODED_VALUES[activity][m]["completed_count"]
                    hardcoded_target = ELIGO_HARDCODED_VALUES[activity][m]["target_count"]
                    row[f"% Work Done against Target-Till {m}"] = f"{pct_done}%"
                    row[f"Target achieved in {m}"] = f"{hardcoded_completed} Flats out of {hardcoded_target} planned"
                else:
                    month_idx = MONTHS.index(m)
                    count_cumulative = sum(completed.get(activity, {}).get(MONTHS[i], 0) for i in range(0, month_idx + 1))
                    target_cumulative = sum(targets.get(activity, {}).get(MONTHS[i], 0) for i in range(0, month_idx + 1))
                    if target_cumulative == 0:
                        pct_done = 100.0
                    else:
                        pct_done = min(round((count_cumulative / target_cumulative) * 100, 2), 100)
                    row[f"% Work Done against Target-Till {m}"] = f"{pct_done}%"
                    month_target = targets.get(activity, {}).get(m, 0)
                    count_in_month = completed.get(activity, {}).get(m, 0)
                    if month_target == 0:
                        future_months = [fm for fm in MONTHS[MONTHS.index(m)+1:] if targets.get(activity, {}).get(fm, 0) > 0]
                        if future_months:
                            row[f"Target achieved in {m}"] = f"Planned for {' and '.join(future_months)}" if len(future_months) > 1 else f"Planned for {future_months[0]}"
                        else:
                            row[f"Target achieved in {m}"] = f"{count_in_month} Flats out of {int(month_target)} planned"
                    else:
                        row[f"Target achieved in {m}"] = f"{count_in_month} Flats out of {int(month_target)} planned"
            else:
                row[f"% Work Done against Target-Till {m}"] = ""
                row[f"Target achieved in {m}"] = ""
        if prev_months:
            last_month = prev_months[-1]
            pct_str = row.get(f"% Work Done against Target-Till {last_month}", "0%").replace("%", "")
            try:
                pct_val = float(pct_str)
                row["Weighted Delay against Targets"] = f"{round((pct_val * weightage) / 100, 2)}%"
            except ValueError:
                row["Weighted Delay against Targets"] = ""
        total_target = sum(targets.get(activity, {}).get(month, 0) for month in MONTHS)
        target_parts = [f"{int(targets.get(activity, {}).get(month, 0))} Flats-{month}" for month in MONTHS if int(targets.get(activity, {}).get(month, 0)) > 0]
        row["Target Till"] = f"{int(total_target)} Flats ({', '.join(target_parts)})"
        progress_data.append(row)

    all_cols = ["Milestone", "Activity", "Target Till"] + [f"% Work Done against Target-Till {month}" for month in MONTHS] + ["Weightage", "Weighted Delay against Targets"] + [f"Target achieved in {month}" for month in MONTHS] + ["Total achieved", f"Delay Reasons_{last_label}"]
    df_th_finishing = pd.DataFrame(progress_data)
    for c in all_cols:
        if c not in df_th_finishing.columns:
            df_th_finishing[c] = ""
    df_th_finishing = df_th_finishing[all_cols]
    return df_th_finishing

# ---------------------------------------------------------------------------
# EXCEL WRITER (unchanged)
# ---------------------------------------------------------------------------
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

def write_excel_report(df_tg_structure, df_th_structure, df_tg_finishing, df_th_finishing, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Eligo Time Delivery Milestones"

    current_date = datetime.now().strftime("%d-%m-%Y")
    ws.append(["Eligo Time Delivery Milestones"])
    ws.append([f"Report Generated on: {current_date}"])
    ws.append([])

    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    grey = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    bold_font = Font(bold=True)
    normal_font = Font(bold=False)
    title_font = Font(bold=True, size=14)
    date_font = Font(bold=False, size=10, color="666666")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="000000")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    max_cols = max(len(df_tg_structure.columns), len(df_th_structure.columns), len(df_tg_finishing.columns), len(df_th_finishing.columns))

    ws.merge_cells(f"A1:{get_column_letter(max_cols)}1")
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align
    ws["A1"].fill = grey

    ws.merge_cells(f"A2:{get_column_letter(max_cols)}2")
    ws["A2"].font = date_font
    ws["A2"].alignment = center_align

    def append_df_block(title, df, total_delay_label):
        ws.append([title])
        title_row = ws.max_row
        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=len(df.columns))
        for cell in ws[title_row]:
            cell.fill = grey
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = border

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        header_row = title_row + 1

        for cell in ws[header_row]:
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = border

        for r in range(header_row + 1, ws.max_row + 1):
            for cell in ws[r]:
                cell.font = bold_font if r in ROWS_TO_BOLD else normal_font
                cell.alignment = left_align if cell.col_idx in (1, 2) else center_align
                cell.border = border

        try:
            total_delay = sum(float(str(v).strip("%")) for v in df["Weighted Delay against Targets"] if v)
        except:
            total_delay = 0

        weighted_delay_col_idx = None
        for idx, col_name in enumerate(df.columns, start=1):
            if col_name == "Weighted Delay against Targets":
                weighted_delay_col_idx = idx
                break

        total_row_data = [""] * len(df.columns)
        if weighted_delay_col_idx:
            total_row_data[weighted_delay_col_idx - 1] = f"{round(total_delay, 2)}%"
            total_row_data[0] = total_delay_label

        ws.append(total_row_data)
        delay_row = ws.max_row
        for idx, cell in enumerate(ws[delay_row], start=1):
            cell.font = bold_font
            cell.fill = yellow
            cell.alignment = left_align if idx == 1 else center_align
            cell.border = border

    append_df_block("Tower G Structure Progress Against Milestones", df_tg_structure, "Total Delay Tower G Structure")
    append_df_block("Tower H Structure Progress Against Milestones", df_th_structure, "Total Delay Tower H Structure")
    append_df_block("Tower G Finishing Progress Against Milestones", df_tg_finishing, "Total Delay Tower G Finishing")
    append_df_block("Tower H Finishing Progress Against Milestones", df_th_finishing, "Total Delay Tower H Finishing")

    for col in ws.columns:
        max_len = 0
        for cell in col:
            text = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(text.split("\n")[0]))
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 22

    wb.save(filename)
    logger.info(f"Eligo report saved to {filename}")
    
def get_unique_filename(base_name):
    """
    If file exists, append (1), (2), etc.
    """
    if not os.path.exists(base_name):
        return base_name

    name, ext = os.path.splitext(base_name)
    counter = 1
    new_name = f"{name}({counter}){ext}"
    while os.path.exists(new_name):
        counter += 1
        new_name = f"{name}({counter}){ext}"
    return new_name


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    logger.info("=== STARTING ELIGO REPORT WITH DYNAMIC MONTHS AND YEAR TRACKING ===")
    cos = init_cos()

    # 1) find the latest tracker files and set months based on latest date
    get_latest_tracker_paths(cos)

    # 2) load targets from KRA (KRA is expected on COS at ELIGO_KRA_KEY)
    try:
        load_targets_from_kra(cos)
    except Exception as e:
        logger.warning(f"Failed to load targets from KRA (fallback will be used): {e}")
        setup_tower_targets_fallback()

    # 3) read targets/completed and build dataframes using the rest of your original pipeline
    logger.info(f"Using months: {MONTHS}")
    logger.info(f"Processing months: {PROCESSING_MONTHS}")

    targets_tg_structure = get_tower_g_structure_targets()
    completed_tg_structure = count_tower_g_completed(cos)
    df_tg_structure = build_tower_g_structure_dataframe(targets_tg_structure, completed_tg_structure)

    targets_th_structure = get_tower_h_structure_targets()
    completed_th_structure = count_tower_h_completed(cos)
    df_th_structure = build_tower_h_structure_dataframe(targets_th_structure, completed_th_structure)

    targets_tg_finishing = get_tower_g_finishing_targets()
    completed_tg_finishing = count_tower_g_finishing_completed(cos)
    df_tg_finishing = build_tower_g_finishing_dataframe(targets_tg_finishing, completed_tg_finishing)

    targets_th_finishing = get_tower_h_finishing_targets()
    completed_th_finishing = count_tower_h_finishing_completed(cos)
    df_th_finishing = build_tower_h_finishing_dataframe(targets_th_finishing, completed_th_finishing)

    base_filename = f"Eligo_Time_Delivery_Milestone_Report ({datetime.now():%Y-%m-%d}).xlsx"
    filename = get_unique_filename(base_filename)
    write_excel_report(df_tg_structure, df_th_structure, df_tg_finishing, df_th_finishing, filename)

    logger.info("=== ELIGO REPORT GENERATION COMPLETE ===")
    logger.info(f"Report saved as: {filename}")
    return df_tg_structure, df_th_structure, df_tg_finishing, df_th_finishing

if __name__ == "__main__":
    main()
