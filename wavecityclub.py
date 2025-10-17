# import os
# import re
# import logging
# from io import BytesIO
# from datetime import datetime
# from dateutil.relativedelta import relativedelta

# import pandas as pd
# from openpyxl import Workbook, load_workbook
# from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
# from openpyxl.utils import get_column_letter
# from openpyxl.utils.dataframe import dataframe_to_rows
# from dotenv import load_dotenv
# import ibm_boto3
# from ibm_botocore.client import Config

# # -----------------------------------------------------------------------------
# # CONFIG / CONSTANTS
# # -----------------------------------------------------------------------------
# load_dotenv()
# logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
# logger = logging.getLogger(__name__)

# # Validate required environment variables
# required = {
#     'COS_API_KEY': os.getenv('COS_API_KEY'),
#     'COS_SERVICE_INSTANCE_CRN': os.getenv('COS_SERVICE_INSTANCE_CRN'),
#     'COS_ENDPOINT': os.getenv('COS_ENDPOINT'),
#     'COS_BUCKET_NAME': os.getenv('COS_BUCKET_NAME'),
# }
# missing = [k for k, v in required.items() if not v]
# if missing:
#     logger.error(f"Missing required environment variables: {', '.join(missing)}")
#     raise SystemExit(1)

# COS_API_KEY     = required['COS_API_KEY']
# COS_CRN         = required['COS_SERVICE_INSTANCE_CRN']
# COS_ENDPOINT    = required['COS_ENDPOINT']
# BUCKET          = required['COS_BUCKET_NAME']  # projectreportnew - contains both KRA and trackers

# # Dynamic paths and configuration
# WCC_KRA_KEY = None  # Will be set by get_latest_kra_file()
# TRACKER_KEYS = {}  # Maps month name to tracker file key

# # Dynamic months and years
# MONTHS = []
# MONTH_YEARS = {}  # Maps month name to year
# TRACKER_DATE = None
# TARGET_END_MONTH = None
# TARGET_END_YEAR = None

# # Block mapping from KRA to tracker sheets
# BLOCK_MAPPING = {
#     'Block 1 (B1) Banquet Hall': 'B1 Banket Hall & Finedine ',
#     'Fine Dine': 'B1 Banket Hall & Finedine ',
#     'Block 5 (B5) Admin + Member Lounge+Creche+Av Room + Surveillance Room +Toilets': 'B5',
#     'Block 6 (B6) Toilets': 'B6',
#     'Block 7(B7) Indoor Sports': 'B7',
#     'Block 9 (B9) Spa & Saloon': 'B9',
#     'Block 8 (B8) Squash Court': 'B8',
#     'Block 2 & 3 (B2 & B3) Cafe & Bar': 'B2 & B3',
#     'Block 4 (B4) Indoor Swimming Pool Changing Room & Toilets': 'B4',
#     'Block 11 (B11) Guest House': 'B11',
#     'Block 10 (B10) Gym': 'B10'
# }

# # Special handling for blocks that need enhanced search within specific sheets
# SPECIAL_BLOCKS_ENHANCED_SEARCH = {
#     'Block 1 (B1) Banquet Hall': 'B1 Banket Hall & Finedine ',
#     'Fine Dine': 'B1 Banket Hall & Finedine '
# }

# # -----------------------------------------------------------------------------
# # COS HELPERS
# # -----------------------------------------------------------------------------

# def init_cos():
#     return ibm_boto3.client(
#         's3',
#         ibm_api_key_id=COS_API_KEY,
#         ibm_service_instance_id=COS_CRN,
#         config=Config(signature_version='oauth'),
#         endpoint_url=COS_ENDPOINT,
#     )

# def download_file_bytes(cos, key):
#     """Download file from bucket"""
#     if not key:
#         raise ValueError("File key cannot be None or empty")
#     obj = cos.get_object(Bucket=BUCKET, Key=key)
#     return obj['Body'].read()

# def list_files_in_folder(cos, folder_prefix=""):
#     """List all files in a specific folder (prefix) in the COS bucket"""
#     try:
#         if folder_prefix:
#             response = cos.list_objects_v2(Bucket=BUCKET, Prefix=folder_prefix)
#         else:
#             response = cos.list_objects_v2(Bucket=BUCKET)
        
#         files = []
#         if 'Contents' in response:
#             for obj in response['Contents']:
#                 if not obj['Key'].endswith('/'):
#                     files.append(obj['Key'])
#         return files
#     except Exception as e:
#         logger.error(f"Error listing files in bucket {BUCKET}, folder {folder_prefix}: {e}")
#         return []

# def extract_date_from_filename(filename):
#     """Extract date from filename in format (dd-mm-yyyy)"""
#     pattern = r'\((\d{2}-\d{2}-\d{4})\)'
#     match = re.search(pattern, filename)
#     if match:
#         date_str = match.group(1)
#         try:
#             return datetime.strptime(date_str, '%d-%m-%Y')
#         except ValueError:
#             logger.warning(f"Could not parse date {date_str} from filename {filename}")
#             return None
#     return None

# def extract_months_and_year_from_kra_filename(filename):
#     """
#     Extract months and year from KRA filename.
#     Example: 'KRA Milestones for June July August 2025' -> ['June', 'July', 'August'], 2025
#     """
#     # Pattern to match: KRA Milestones for [Month] [Month] [Month] [Year]
#     pattern = r'KRA Milestones for\s+(.*?)\s+(\d{4})'
#     match = re.search(pattern, filename, re.IGNORECASE)
    
#     if match:
#         months_str = match.group(1)
#         year = int(match.group(2))
        
#         # Extract month names
#         month_names = ["January", "February", "March", "April", "May", "June",
#                       "July", "August", "September", "October", "November", "December"]
        
#         found_months = []
#         for month in month_names:
#             if month in months_str:
#                 found_months.append(month)
        
#         logger.info(f"Extracted from '{filename}': Months={found_months}, Year={year}")
#         return found_months, year
    
#     return None, None

# def get_month_name(month_num):
#     """Convert month number to month name"""
#     months = {
#         1: "January", 2: "February", 3: "March", 4: "April",
#         5: "May", 6: "June", 7: "July", 8: "August", 
#         9: "September", 10: "October", 11: "November", 12: "December"
#     }
#     return months.get(month_num, "Unknown")

# def get_month_number(month_name):
#     """Convert month name to month number"""
#     months = {
#         "January": 1, "February": 2, "March": 3, "April": 4,
#         "May": 5, "June": 6, "July": 7, "August": 8, 
#         "September": 9, "October": 10, "November": 11, "December": 12
#     }
#     return months.get(month_name, 1)

# def get_latest_kra_file(cos):
#     """Get the latest KRA Milestones file from root of bucket"""
#     global WCC_KRA_KEY
    
#     logger.info("=== FINDING LATEST KRA MILESTONES FILE ===")
#     logger.info(f"Searching in bucket: {BUCKET} (root level)")
    
#     # List all files in the bucket root (no folder prefix)
#     all_files = list_files_in_folder(cos, "")
#     logger.info(f"Found {len(all_files)} total files in {BUCKET} bucket")
    
#     # Filter for KRA files at root level (not in subfolders)
#     kra_pattern = r'KRA Milestones for.*\.xlsx$'
    
#     matching_files = []
    
#     for file_path in all_files:
#         # Skip files in subfolders (contain /)
#         if '/' in file_path:
#             continue
            
#         filename = os.path.basename(file_path)
        
#         if re.search(kra_pattern, filename, re.IGNORECASE):
#             logger.info(f"Found KRA file: {filename}")
#             months, year = extract_months_and_year_from_kra_filename(filename)
            
#             if months and year:
#                 # Use the last month and year to create a comparable date
#                 last_month = months[-1]
#                 month_num = get_month_number(last_month)
#                 # Create a date object for comparison (use day=1)
#                 file_date = datetime(year, month_num, 1)
#                 matching_files.append((file_path, file_date, months, year))
#                 logger.info(f"  -> Months: {months}, Year: {year}")
#             else:
#                 logger.warning(f"Could not extract months/year from: {filename}")
    
#     if matching_files:
#         # Get the most recent one based on year and last month
#         latest_file = max(matching_files, key=lambda x: x[1])
#         WCC_KRA_KEY = latest_file[0]
#         logger.info(f"✅ Latest KRA file: {WCC_KRA_KEY}")
#         logger.info(f"   Months: {latest_file[2]}, Year: {latest_file[3]}")
#         return latest_file[1]  # Return the date
#     else:
#         logger.error(f"❌ No KRA Milestones files found in bucket {BUCKET} root!")
#         raise Exception("Could not find latest KRA Milestones file")

# def determine_quarter_months_from_kra(kra_date):
#     """
#     Determine quarter months based on KRA file date.
#     The KRA date represents the last month in the quarter.
#     """
#     month = kra_date.month
#     year = kra_date.year
    
#     # Determine the quarter based on the last month
#     if month in [6, 7, 8]:  # June, July, August quarter
#         return [
#             ("June", 6, year),
#             ("July", 7, year),
#             ("August", 8, year)
#         ]
#     elif month in [9, 10, 11]:  # September, October, November quarter
#         return [
#             ("September", 9, year),
#             ("October", 10, year),
#             ("November", 11, year)
#         ]
#     elif month in [12, 1, 2]:  # December, January, February quarter
#         # Handle year transition
#         dec_year = year if month == 12 else year - 1
#         jan_feb_year = year if month in [1, 2] else year + 1
#         return [
#             ("December", 12, dec_year),
#             ("January", 1, jan_feb_year),
#             ("February", 2, jan_feb_year)
#         ]
#     else:  # March, April, May quarter (months 3, 4, 5)
#         return [
#             ("March", 3, year),
#             ("April", 4, year),
#             ("May", 5, year)
#         ]

# def find_tracker_for_month(cos, month_name, month_num, year):
#     """
#     Find tracker file for a specific month in Wave City Club folder.
#     For a given month, we need tracker from the NEXT month.
#     E.g., for June data, we need tracker from July (DD-07-YYYY)
#     """
#     # Calculate the next month for tracker lookup
#     next_month = month_num + 1
#     tracker_year = year
    
#     if next_month > 12:
#         next_month = 1
#         tracker_year = year + 1
    
#     logger.info(f"Looking for tracker for {month_name} {year}: need tracker from month {next_month:02d}/{tracker_year}")
    
#     # List files in Wave City Club folder
#     wcc_files = list_files_in_folder(cos, "Wave City Club/")
#     tracker_pattern = r'Structure Work Tracker.*\.xlsx$'
    
#     matching_trackers = []
    
#     for file_path in wcc_files:
#         filename = os.path.basename(file_path)
#         if re.search(tracker_pattern, filename, re.IGNORECASE):
#             file_date = extract_date_from_filename(filename)
#             if file_date:
#                 # Check if this tracker is from the correct month and year
#                 if file_date.month == next_month and file_date.year == tracker_year:
#                     matching_trackers.append((file_path, file_date))
#                     logger.info(f"  Found matching tracker: {filename} ({file_date.strftime('%d-%m-%Y')})")
    
#     if matching_trackers:
#         # Get the latest tracker from that month (in case there are multiple)
#         latest_tracker = max(matching_trackers, key=lambda x: x[1])
#         logger.info(f"  ✅ Selected tracker for {month_name}: {os.path.basename(latest_tracker[0])}")
#         return latest_tracker[0]
#     else:
#         logger.warning(f"  ⚠️ No tracker found for {month_name} {year} (looking for {next_month:02d}/{tracker_year})")
#         return None

# def setup_quarterly_configuration(cos):
#     """Setup configuration based on latest KRA file"""
#     global MONTHS, MONTH_YEARS, TARGET_END_MONTH, TARGET_END_YEAR, TRACKER_KEYS
    
#     # Get latest KRA file from bucket root
#     kra_date = get_latest_kra_file(cos)
    
#     # Determine quarter months based on KRA
#     quarter_months = determine_quarter_months_from_kra(kra_date)
    
#     # Setup global variables
#     MONTHS = [f"{name} {yr}" for name, num, yr in quarter_months]
#     MONTH_YEARS = {f"{name} {yr}": yr for name, num, yr in quarter_months}
    
#     TARGET_END_MONTH = quarter_months[-1][0]
#     TARGET_END_YEAR = quarter_months[-1][2]
    
#     logger.info(f"=== QUARTERLY CONFIGURATION ===")
#     logger.info(f"Months: {MONTHS}")
#     logger.info(f"Target end: {TARGET_END_MONTH} {TARGET_END_YEAR}")
    
#     # Find trackers for each month in Wave City Club folder
#     TRACKER_KEYS = {}
#     for month_name, month_num, year in quarter_months:
#         tracker_key = find_tracker_for_month(cos, month_name, month_num, year)
#         month_label = f"{month_name} {year}"
#         TRACKER_KEYS[month_label] = tracker_key
#         if tracker_key:
#             logger.info(f"✅ Tracker mapped: {month_label} -> {os.path.basename(tracker_key)}")
#         else:
#             logger.warning(f"⚠️ No tracker for: {month_label}")
    
#     return quarter_months

# # -----------------------------------------------------------------------------
# # UTILITIES
# # -----------------------------------------------------------------------------

# def extract_percentage(cell_value):
#     """Extract percentage value from cell, handling different formats"""
#     if not cell_value or cell_value == '-':
#         return 0.0
    
#     if isinstance(cell_value, (int, float)):
#         if cell_value <= 1.0:
#             return cell_value * 100
#         return cell_value
    
#     val_str = str(cell_value).replace('%', '').strip()
#     try:
#         val = float(val_str)
#         if val <= 1.0:
#             return val * 100
#         return val
#     except ValueError:
#         numbers = re.findall(r'\d+\.?\d*', val_str)
#         if numbers:
#             val = float(numbers[0])
#             return val if val > 1.0 else val * 100
#         return 0.0

# def normalize_activity_name(activity):
#     """Normalize activity name for better matching"""
#     if not activity:
#         return ""
#     return str(activity).strip().lower()

# def activities_match(target_activity, tracker_activity):
#     """Enhanced matching with case-insensitive comparison"""
#     if not target_activity or not tracker_activity:
#         return False
    
#     target = str(target_activity).strip()
#     tracker = str(tracker_activity).strip()
    
#     if target == tracker:
#         return True
    
#     if target.lower() == tracker.lower():
#         logger.info(f"CASE-INSENSITIVE MATCH: '{target}' matches '{tracker}'")
#         return True
    
#     logger.debug(f"NO MATCH: Target='{target}' vs Tracker='{tracker}'")
#     return False

# # -----------------------------------------------------------------------------
# # DATA EXTRACTION FUNCTIONS
# # -----------------------------------------------------------------------------

# def detect_kra_column_mapping(sheet):
#     """Detect month-to-column mapping from KRA file headers"""
#     month_to_col = {}
    
#     for col_idx in range(2, 14):  # Columns B to M (2 to 13)
#         col_letter = get_column_letter(col_idx)
#         header_cell = sheet[f'{col_letter}1']
        
#         if header_cell.value:
#             header_text = str(header_cell.value).strip()
#             logger.info(f"KRA Header in column {col_letter}: '{header_text}'")
            
#             # Try to extract month name from header
#             for month in ["January", "February", "March", "April", "May", "June", 
#                          "July", "August", "September", "October", "November", "December"]:
#                 if month.lower() in header_text.lower():
#                     month_to_col[month] = col_letter
#                     logger.info(f"Mapped {month} -> Column {col_letter}")
#                     break
    
#     return month_to_col

# def get_wcc_targets_from_kra(cos):
#     """Extract targets from KRA file dynamically based on current months"""
#     # Download from bucket
#     raw = download_file_bytes(cos, WCC_KRA_KEY)
#     wb = load_workbook(filename=BytesIO(raw), data_only=True)
    
#     # Try to find the correct sheet
#     sheet = None
#     possible_sheet_names = ['Wave City Club targets till Aug', 'Sheet1', 'Targets', 'KRA', 'Wave City Club']
    
#     for sheet_name in possible_sheet_names:
#         if sheet_name in wb.sheetnames:
#             sheet = wb[sheet_name]
#             logger.info(f"Using sheet: {sheet_name}")
#             break
    
#     if sheet is None:
#         # Use first sheet as fallback
#         sheet = wb[wb.sheetnames[0]]
#         logger.info(f"Using first sheet: {wb.sheetnames[0]}")
    
#     targets = {}
#     logger.info("=== EXTRACTING TARGETS FROM KRA FILE ===")
#     logger.info(f"Looking for months: {MONTHS} with years: {MONTH_YEARS}")
    
#     # Detect column mapping from file
#     month_to_col = detect_kra_column_mapping(sheet)
    
#     if not month_to_col:
#         logger.warning("Could not detect column mapping, using default")
#         month_to_col = {
#             "June": "B", "July": "C", "August": "D",
#             "September": "E", "October": "F", "November": "G", "December": "H",
#             "January": "I", "February": "J", "March": "K", "April": "L", "May": "M"
#         }
    
#     # Read targets from the KRA file
#     for row_num in range(2, sheet.max_row + 1):
#         block_cell = sheet[f'A{row_num}']
        
#         if block_cell.value:
#             block_name = str(block_cell.value).strip()
#             month_activities = {}
            
#             for month_label in MONTHS:
#                 # Extract month name without year
#                 month_name = month_label.split()[0]
#                 col = month_to_col.get(month_name, "B")
#                 cell = sheet[f'{col}{row_num}']
#                 activity = str(cell.value or '').strip() if cell.value else ''
#                 month_activities[month_label] = activity
#                 logger.info(f"Row {row_num}, {month_label}: Block='{block_name}', Activity='{activity}'")
            
#             targets[block_name] = month_activities
    
#     logger.info(f"Extracted targets for {len(targets)} blocks")
#     return targets

# def find_activity_progress_in_sheet(sheet, target_activity, sheet_name, block_name=None):
#     """Find activity progress with enhanced search for special blocks"""
#     logger.info(f"=== Looking for '{target_activity}' in '{sheet_name}' for '{block_name}' ===")
    
#     # Return 100% when there's no target activity
#     if not target_activity or target_activity.strip() == '' or target_activity.lower() in ['no target', '-']:
#         logger.info(f"No target activity, returning 100%")
#         return 100.0
    
#     # Enhanced search for special blocks
#     if block_name in SPECIAL_BLOCKS_ENHANCED_SEARCH:
#         logger.info(f"SPECIAL CASE: {block_name} - enhanced search")
#         max_rows = min(sheet.max_row, 60)
        
#         # Collect all matches with their row numbers and progress values
#         all_matches = []
        
#         for row_num in range(1, max_rows + 1):
#             try:
#                 activity_cell = sheet[f'G{row_num}']
#                 if activity_cell.value:
#                     tracker_activity = str(activity_cell.value).strip()
                    
#                     if activities_match(target_activity, tracker_activity):
#                         progress_cell = sheet[f'AC{row_num}']
#                         ac_value = progress_cell.value
                        
#                         # Also check the block identifier in column A or B
#                         block_cell_a = sheet[f'A{row_num}']
#                         block_cell_b = sheet[f'B{row_num}']
                        
#                         block_identifier = None
#                         if block_cell_a.value:
#                             block_identifier = str(block_cell_a.value).strip()
#                         elif block_cell_b.value:
#                             block_identifier = str(block_cell_b.value).strip()
                        
#                         logger.info(f"Found match at row {row_num}: '{tracker_activity}', Block: '{block_identifier}', AC: {ac_value}")
                        
#                         if ac_value is not None:
#                             progress = extract_percentage(ac_value)
#                             all_matches.append({
#                                 'row': row_num,
#                                 'activity': tracker_activity,
#                                 'block': block_identifier,
#                                 'progress': progress
#                             })
#             except Exception as e:
#                 logger.debug(f"Error at row {row_num}: {e}")
#                 continue
        
#         # If we found multiple matches, try to find the best one
#         if all_matches:
#             logger.info(f"Found {len(all_matches)} matches for '{target_activity}'")
            
#             # For Fine Dine, look for the match that is NOT in B1 rows
#             # (B1 activities typically have "B1" in their block identifier)
#             if 'Fine Dine' in block_name:
#                 # Filter out B1 activities - look for rows that have "FI" or "Fine" or are NOT "B1" or "Bl"
#                 non_b1_matches = []
#                 for m in all_matches:
#                     if m['block']:
#                         block_upper = m['block'].upper()
#                         # Include if it contains FI or FINE, or if it doesn't contain B1/Bl
#                         if 'FI' in block_upper or 'FINE' in block_upper:
#                             non_b1_matches.append(m)
#                         elif 'B1' not in block_upper and 'BL' not in block_upper:
#                             non_b1_matches.append(m)
                
#                 if non_b1_matches:
#                     # Return the match with the highest progress
#                     best_match = max(non_b1_matches, key=lambda x: x['progress'])
#                     logger.info(f"✅ Selected Fine Dine match at row {best_match['row']}: {best_match['progress']}% (Block: {best_match['block']})")
#                     return best_match['progress']
#                 else:
#                     logger.warning(f"Could not find specific Fine Dine match, using first match")
            
#             # For Block 1 (B1) Banquet Hall, prefer B1 activities
#             if 'Block 1' in block_name or ('B1' in block_name and 'Fine' not in block_name):
#                 b1_matches = []
#                 for m in all_matches:
#                     if m['block']:
#                         block_upper = m['block'].upper()
#                         # Include if it contains B1 or Bl, but NOT FI or FINE
#                         if ('B1' in block_upper or 'BL' in block_upper) and 'FI' not in block_upper and 'FINE' not in block_upper:
#                             b1_matches.append(m)
                
#                 if b1_matches:
#                     best_match = max(b1_matches, key=lambda x: x['progress'])
#                     logger.info(f"✅ Selected B1 match at row {best_match['row']}: {best_match['progress']}% (Block: {best_match['block']})")
#                     return best_match['progress']
            
#             # If we couldn't determine which one to use, return the one with highest progress
#             best_match = max(all_matches, key=lambda x: x['progress'])
#             logger.info(f"Selected best match at row {best_match['row']}: {best_match['progress']}%")
#             return best_match['progress']
        
#         logger.warning(f"NO MATCH for '{target_activity}'")
#         return 0.0
    
#     # Standard search for other blocks
#     max_rows = min(sheet.max_row, 20)
#     for row_num in range(1, max_rows + 1):
#         try:
#             activity_cell = sheet[f'G{row_num}']
#             if activity_cell.value:
#                 tracker_activity = str(activity_cell.value).strip()
                
#                 if activities_match(target_activity, tracker_activity):
#                     progress_cell = sheet[f'AC{row_num}']
#                     ac_value = progress_cell.value
                    
#                     if ac_value is not None:
#                         return extract_percentage(ac_value)
#                     return 0.0
#         except Exception as e:
#             logger.debug(f"Error at row {row_num}: {e}")
#             continue
    
#     logger.warning(f"NO MATCH for '{target_activity}'")
#     return 0.0

# def get_progress_from_specific_tracker(cos, tracker_key, block_name, activity, sheet_name):
#     """Get progress from a specific tracker file"""
#     if not tracker_key:
#         logger.warning(f"No tracker available, returning None")
#         return None
    
#     try:
#         # Download from bucket
#         raw = download_file_bytes(cos, tracker_key)
#         wb = load_workbook(filename=BytesIO(raw), data_only=True)
        
#         if sheet_name not in wb.sheetnames:
#             logger.warning(f"Sheet '{sheet_name}' not found in tracker")
#             return None
        
#         sheet = wb[sheet_name]
#         progress = find_activity_progress_in_sheet(sheet, activity, sheet_name, block_name)
#         return progress
        
#     except Exception as e:
#         logger.error(f"Error reading tracker {tracker_key}: {e}")
#         return None

# def get_wcc_progress_from_trackers_quarterly(cos, targets):
#     """Extract progress data from multiple tracker files based on quarter"""
#     progress_data = []
#     milestone_counter = 1
#     total_blocks = len(targets)
#     site_weighted = round(100 / total_blocks, 2) if total_blocks > 0 else 0
    
#     for block_name, month_activities in targets.items():
#         logger.info(f"Processing block: {block_name}")
        
#         sheet_name = BLOCK_MAPPING.get(block_name)
#         month_progress = {}
        
#         if not sheet_name:
#             logger.warning(f"No sheet mapping for block: {block_name}")
#             # Set all months to None
#             for month_label in MONTHS:
#                 month_progress[month_label] = None
#         else:
#             # Get progress from appropriate tracker for each month
#             for month_label in MONTHS:
#                 tracker_key = TRACKER_KEYS.get(month_label)
#                 activity = month_activities.get(month_label, '')
                
#                 if tracker_key:
#                     progress = get_progress_from_specific_tracker(
#                         cos, tracker_key, block_name, activity, sheet_name
#                     )
#                     month_progress[month_label] = progress
#                     logger.info(f"{block_name} - {month_label}: {progress}%")
#                 else:
#                     month_progress[month_label] = None
#                     logger.warning(f"{block_name} - {month_label}: No tracker available")
        
#         # Use the last month for weighted calculation (if available)
#         last_month = MONTHS[-1] if MONTHS else ''
#         last_month_progress = month_progress.get(last_month)
        
#         if last_month_progress is not None:
#             main_weighted = round((site_weighted * last_month_progress) / 100, 3)
#         else:
#             main_weighted = 0.0
        
#         # Create row data with dynamic columns
#         row_data = {
#             'Milestone': f"Milestone-{milestone_counter:02d}",
#             'Block': block_name,
#             f'Target to be complete by {TARGET_END_MONTH}-{TARGET_END_YEAR}': month_activities.get(last_month, ''),
#             'Site Weighted': site_weighted,
#             'Weighted progress against target': main_weighted,
#         }
        
#         # Add month-specific columns
#         for month_label in MONTHS:
#             year = MONTH_YEARS[month_label]
#             month_name = month_label.split()[0]
#             target_val = month_activities.get(month_label, '')
#             progress_val = month_progress.get(month_label)
            
#             row_data[f'Target - {month_name}-{year}'] = target_val
            
#             # Handle missing tracker case
#             if progress_val is None:
#                 row_data[f'% work done- {month_name} Status'] = 'No tracker'
#                 row_data[f'Achieved- {month_name} {year}'] = 'No tracker available'
#             else:
#                 row_data[f'% work done- {month_name} Status'] = f"{progress_val:.0f}%"
                
#                 if progress_val == 100:
#                     achieved = target_val if target_val else f'No target for {month_name}'
#                 elif progress_val == 0:
#                     achieved = 'No progress' if target_val else f'No target for {month_name}'
#                 else:
#                     achieved = f'{progress_val:.0f}% completed'
#                 row_data[f'Achieved- {month_name} {year}'] = achieved
            
#             row_data[f'Responsible Person- {month_name}'] = ''
#             row_data[f'Delay Reasons- {month_name}'] = ''
        
#         progress_data.append(row_data)
#         milestone_counter += 1
    
#     # Create DataFrame with dynamic columns
#     columns = [
#         'Milestone', 
#         'Block', 
#         f'Target to be complete by {TARGET_END_MONTH}-{TARGET_END_YEAR}'
#     ]
    
#     for month_label in MONTHS:
#         year = MONTH_YEARS[month_label]
#         month_name = month_label.split()[0]
#         columns.extend([
#             f'Target - {month_name}-{year}',
#             f'% work done- {month_name} Status',
#             f'Achieved- {month_name} {year}',
#             f'Responsible Person- {month_name}',
#             f'Delay Reasons- {month_name}'
#         ])
    
#     columns.extend(['Site Weighted', 'Weighted progress against target'])
    
#     df = pd.DataFrame(progress_data, columns=columns)
#     logger.info(f"Created DataFrame with {len(df)} rows for months: {MONTHS}")
#     return df

# # -----------------------------------------------------------------------------
# # EXCEL REPORT GENERATION
# # -----------------------------------------------------------------------------

# def write_wcc_excel_report_consolidated(df, filename):
#     """Generate formatted Excel report with dynamic month columns"""
#     wb = Workbook()
#     ws = wb.active
#     ws.title = 'Wave City Club- Progress'[:31]  # Excel sheet name limit
    
#     # Add main title
#     title_row = ["Wave City Club- Progress Against Milestones"]
#     ws.append(title_row)
#     ws.merge_cells(f'A1:{get_column_letter(len(df.columns))}1')
    
#     # Add date row
#     current_date = datetime.now().strftime("%d-%m-%Y")
#     date_row = [f"Report Generated on: {current_date}"]
#     ws.append(date_row)
#     ws.merge_cells(f'A2:{get_column_letter(len(df.columns))}2')
    
#     # Add month info row
#     month_info = f"Months Covered: {', '.join(MONTHS)}"
#     month_info_row = [month_info]
#     ws.append(month_info_row)
#     ws.merge_cells(f'A3:{get_column_letter(len(df.columns))}3')
    
#     # Add empty row
#     ws.append([])
    
#     # Add DataFrame data with percentage formatting for weighted progress
#     for row in dataframe_to_rows(df, index=False, header=True):
#         # Format the weighted progress column (last column) to add % symbol
#         if len(row) > 0 and isinstance(row[-1], (int, float)) and row[-1] != '':
#             row[-1] = f"{row[-1]:.3f}%"
#         ws.append(row)
    
#     # Add Sum row - Only for the weighted progress column
#     weighted_sum = df['Weighted progress against target'].sum()
    
#     # Create sum row with blanks for all columns except the weighted progress column
#     sum_row = [''] * len(df.columns)
#     sum_row[-2] = 'Sum'  # Site Weighted column
#     sum_row[-1] = f'{weighted_sum:.3f}%'  # Weighted progress column
#     ws.append(sum_row)
    
#     # Define styles
#     title_font = Font(bold=True, size=12)
#     header_font = Font(bold=True, size=8)
#     normal_font = Font(bold=False, size=8)
#     date_font = Font(bold=False, size=10, color="666666")
#     center = Alignment(horizontal='center', vertical='center', wrap_text=True)
#     left = Alignment(horizontal='left', vertical='center', wrap_text=True)
#     thin = Side(style='thin', color='000000')
#     border = Border(top=thin, bottom=thin, left=thin, right=thin)
#     light_grey_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
#     light_blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    
#     # Style title (light grey background)
#     ws['A1'].font = title_font
#     ws['A1'].alignment = center
#     ws['A1'].fill = light_grey_fill
    
#     # Style date row
#     ws['A2'].font = date_font
#     ws['A2'].alignment = center
    
#     # Style month info row
#     ws['A3'].font = date_font
#     ws['A3'].alignment = center
    
#     # Style header row (row 5) with light grey background
#     header_row = 5
#     for cell in ws[header_row]:
#         cell.font = header_font
#         cell.alignment = center
#         cell.border = border
#         cell.fill = light_grey_fill
    
#     # Style data rows
#     data_start = 6
#     data_end = ws.max_row - 1  # Exclude sum row for now
    
#     for row_num in range(data_start, data_end + 1):
#         for col_num in range(1, len(df.columns) + 1):
#             cell = ws.cell(row=row_num, column=col_num)
#             cell.font = normal_font
#             cell.border = border
            
#             # Alignment based on column type
#             if col_num in [1, 2, 3] or 'Target' in str(ws.cell(row=header_row, column=col_num).value or ''):  # Text columns
#                 cell.alignment = left
#             else:  # Numeric columns
#                 cell.alignment = center
    
#     # Style sum row with light blue background
#     sum_row_num = ws.max_row
#     for col_num in range(1, len(df.columns) + 1):
#         cell = ws.cell(row=sum_row_num, column=col_num)
#         cell.font = header_font
#         cell.border = border
#         cell.fill = light_blue_fill
#         cell.alignment = center
    
#     # Dynamic column width adjustment
#     for col_num in range(1, len(df.columns) + 1):
#         col_letter = get_column_letter(col_num)
        
#         # Calculate optimal width based on column content
#         max_length = 0
#         for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=col_num, max_col=col_num):
#             for cell in row:
#                 if cell.value:
#                     max_length = max(max_length, len(str(cell.value)))
        
#         # Set minimum and maximum width constraints
#         calculated_width = min(max(max_length + 2, 8), 15)
#         ws.column_dimensions[col_letter].width = calculated_width
    
#     # Set row heights
#     ws.row_dimensions[1].height = 25  # Title row
#     ws.row_dimensions[2].height = 20  # Date row
#     ws.row_dimensions[3].height = 20  # Month info row
#     for i in range(5, ws.max_row + 1):
#         ws.row_dimensions[i].height = 25
    
#     wb.save(filename)
#     logger.info(f'Dynamic report saved to {filename}')


# def get_unique_filename(base_name):
#     """
#     If file exists, append (1), (2), etc.
#     """
#     if not os.path.exists(base_name):
#         return base_name

#     name, ext = os.path.splitext(base_name)
#     counter = 1
#     new_name = f"{name}({counter}){ext}"
#     while os.path.exists(new_name):
#         counter += 1
#         new_name = f"{name}({counter}){ext}"
#     return new_name

# # -----------------------------------------------------------------------------
# # MAIN FUNCTION
# # -----------------------------------------------------------------------------

# def main():
#     """Main execution function for quarterly report generation"""
#     logger.info("=== STARTING WAVE CITY CLUB QUARTERLY REPORT GENERATION ===")
    
#     try:
#         # Initialize COS client
#         cos = init_cos()
        
#         # Setup quarterly configuration (finds KRA and determines months)
#         logger.info("\n=== STEP 1: Setting up quarterly configuration ===")
#         logger.info(f"Bucket: {BUCKET}")
#         logger.info(f"Looking for KRA files in bucket root")
#         logger.info(f"Looking for tracker files in Wave City Club/ folder")
#         quarter_months = setup_quarterly_configuration(cos)
        
#         logger.info(f"\n=== QUARTERLY CONFIGURATION COMPLETE ===")
#         logger.info(f"KRA File: {WCC_KRA_KEY}")
#         logger.info(f"Months: {MONTHS}")
#         logger.info(f"Target Period End: {TARGET_END_MONTH} {TARGET_END_YEAR}")
        
#         logger.info(f"\n=== TRACKER FILES MAPPING ===")
#         for month_label, tracker_key in TRACKER_KEYS.items():
#             if tracker_key:
#                 logger.info(f"✅ {month_label}: {os.path.basename(tracker_key)}")
#             else:
#                 logger.info(f"⚠️ {month_label}: No tracker found")
        
#         # Get targets from KRA file
#         logger.info("\n=== STEP 2: Fetching targets from KRA file ===")
#         targets = get_wcc_targets_from_kra(cos)
#         logger.info(f"Extracted targets for {len(targets)} blocks")
        
#         # Extract progress data from quarterly trackers
#         logger.info("\n=== STEP 3: Extracting progress data from trackers ===")
#         df = get_wcc_progress_from_trackers_quarterly(cos, targets)
        
#         # Generate report
#         logger.info("\n=== STEP 4: Generating Excel report ===")
#         current_date_for_filename = datetime.now().strftime('%d-%m-%Y')
#         base_filename = f"Wave_City_Club Milestone Report ({current_date_for_filename}).xlsx"
#         filename = get_unique_filename(base_filename)
        
#         write_wcc_excel_report_consolidated(df, filename)
        
#         logger.info("\n=== WAVE CITY CLUB QUARTERLY REPORT GENERATION COMPLETE ===")
#         logger.info(f"✅ Report saved as: {filename}")
        
#         # Log summary
#         logger.info("\n=== REPORT SUMMARY ===")
#         logger.info(f"Quarter Months: {MONTHS}")
#         logger.info(f"Processed Blocks: {len(targets)}")
#         logger.info(f"Trackers Used:")
#         for month_label, tracker_key in TRACKER_KEYS.items():
#             status = "✅ Available" if tracker_key else "⚠️ Missing"
#             logger.info(f"  {month_label}: {status}")
        
#     except Exception as e:
#         logger.error(f"Error in main execution: {e}")
#         import traceback
#         logger.error(traceback.format_exc())
#         raise

# if __name__ == "__main__":
#     main()







































































import os
import re
import logging
from io import BytesIO
from datetime import datetime
from dateutil.relativedelta import relativedelta

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from dotenv import load_dotenv
import ibm_boto3
from ibm_botocore.client import Config

# -----------------------------------------------------------------------------
# CONFIG / CONSTANTS
# -----------------------------------------------------------------------------
load_dotenv()
logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
logger = logging.getLogger(__name__)

# Validate required environment variables
required = {
    'COS_API_KEY': os.getenv('COS_API_KEY'),
    'COS_SERVICE_INSTANCE_CRN': os.getenv('COS_SERVICE_INSTANCE_CRN'),
    'COS_ENDPOINT': os.getenv('COS_ENDPOINT'),
    'COS_BUCKET_NAME': os.getenv('COS_BUCKET_NAME'),
}
missing = [k for k, v in required.items() if not v]
if missing:
    logger.error(f"Missing required environment variables: {', '.join(missing)}")
    raise SystemExit(1)

COS_API_KEY     = required['COS_API_KEY']
COS_CRN         = required['COS_SERVICE_INSTANCE_CRN']
COS_ENDPOINT    = required['COS_ENDPOINT']
BUCKET          = required['COS_BUCKET_NAME']

# Dynamic KRA file path - will be set by get_latest_kra_file()
WCC_KRA_KEY = None

# Dynamic tracker paths - will be set by get_tracker_for_month()
TRACKER_PATHS = {}  # Maps month names to tracker file paths
LOADED_TRACKERS = {}  # Cache loaded workbooks

# Dynamic months and years
MONTHS = []
MONTH_YEARS = {}  # Maps month name to year
TARGET_END_MONTH = None
TARGET_END_YEAR = None

# Quarterly month groups
QUARTERLY_GROUPS = [
    ['June', 'July', 'August'],
    ['September', 'October', 'November'],
    ['December', 'January', 'February'],
    ['March', 'April', 'May']
]

# Block mapping from KRA to tracker sheets
BLOCK_MAPPING = {
    'Block 1 (B1) Banquet Hall': 'B1 Banket Hall & Finedine ',
    'Fine Dine': 'B1 Banket Hall & Finedine ',
    'Block 5 (B5) Admin + Member Lounge+Creche+Av Room + Surveillance Room +Toilets': 'B5',
    'Block 6 (B6) Toilets': 'B6',
    'Block 7(B7) Indoor Sports': 'B7',
    'Block 9 (B9) Spa & Saloon': 'B9',
    'Block 8 (B8) Squash Court': 'B8',
    'Block 2 & 3 (B2 & B3) Cafe & Bar': 'B2 & B3',
    'Block 4 (B4) Indoor Swimming Pool Changing Room & Toilets': 'B4',
    'Block 11 (B11) Guest House': 'B11',
    'Block 10 (B10) Gym': 'B10'
}

# Special handling for blocks that need enhanced search within specific sheets
SPECIAL_BLOCKS_ENHANCED_SEARCH = {
    'Block 1 (B1) Banquet Hall': 'B1 Banket Hall & Finedine ',
    'Fine Dine': 'B1 Banket Hall & Finedine '
}

# -----------------------------------------------------------------------------
# COS HELPERS
# -----------------------------------------------------------------------------

def init_cos():
    return ibm_boto3.client(
        's3',
        ibm_api_key_id=COS_API_KEY,
        ibm_service_instance_id=COS_CRN,
        config=Config(signature_version='oauth'),
        endpoint_url=COS_ENDPOINT,
    )

def download_file_bytes(cos, key):
    if not key:
        raise ValueError("File key cannot be None or empty")
    obj = cos.get_object(Bucket=BUCKET, Key=key)
    return obj['Body'].read()

def list_files_in_folder(cos, folder_prefix):
    """List all files in a specific folder (prefix) in the COS bucket"""
    try:
        response = cos.list_objects_v2(Bucket=BUCKET, Prefix=folder_prefix)
        files = []
        if 'Contents' in response:
            for obj in response['Contents']:
                if not obj['Key'].endswith('/'):
                    files.append(obj['Key'])
        return files
    except Exception as e:
        logger.error(f"Error listing files in folder {folder_prefix}: {e}")
        return []

def extract_date_from_filename(filename):
    """Extract date from filename in format (dd-mm-yyyy)"""
    pattern = r'\((\d{2}-\d{2}-\d{4})\)'
    match = re.search(pattern, filename)
    if match:
        date_str = match.group(1)
        try:
            return datetime.strptime(date_str, '%d-%m-%Y')
        except ValueError:
            logger.warning(f"Could not parse date {date_str} from filename {filename}")
            return None
    return None

def get_month_name(month_num):
    """Convert month number to month name"""
    months = {
        1: "January", 2: "February", 3: "March", 4: "April",
        5: "May", 6: "June", 7: "July", 8: "August", 
        9: "September", 10: "October", 11: "November", 12: "December"
    }
    return months.get(month_num, "Unknown")

def get_month_number(month_name):
    """Convert month name to month number"""
    months = {
        "January": 1, "February": 2, "March": 3, "April": 4,
        "May": 5, "June": 6, "July": 7, "August": 8, 
        "September": 9, "October": 10, "November": 11, "December": 12
    }
    return months.get(month_name, 1)

def get_latest_kra_file(cos):
    """Get the latest KRA milestone file from the bucket root"""
    global WCC_KRA_KEY
    
    logger.info("=== FINDING LATEST KRA MILESTONE FILE ===")
    
    # List all files in bucket root (no folder prefix)
    all_files = list_files_in_folder(cos, "")
    logger.info(f"Found {len(all_files)} files in bucket root")
    
    # Pattern to match KRA milestone files
    kra_pattern = r'KRA Milestones for.*\.xlsx$'
    
    matching_files = []
    
    for file_path in all_files:
        filename = os.path.basename(file_path)
        if re.search(kra_pattern, filename, re.IGNORECASE):
            logger.info(f"Found KRA file: {filename}")
            matching_files.append(file_path)
    
    if matching_files:
        # Sort by filename to get the latest (assuming naming convention includes year)
        latest_file = sorted(matching_files)[-1]
        WCC_KRA_KEY = latest_file
        logger.info(f"✅ Latest KRA file: {WCC_KRA_KEY}")
    else:
        logger.error(f"❌ No KRA milestone files found in bucket root!")
        WCC_KRA_KEY = None
    
    return WCC_KRA_KEY

def extract_months_from_kra_filename(filename):
    """Extract quarter months from KRA filename"""
    months_pattern = r'for\s+((?:January|February|March|April|May|June|July|August|September|October|November|December)(?:\s+(?:January|February|March|April|May|June|July|August|September|October|November|December))*)\s+(\d{4})'
    
    match = re.search(months_pattern, filename, re.IGNORECASE)
    if match:
        months_str = match.group(1)
        year = int(match.group(2))
        
        # Extract individual months
        month_names = re.findall(r'January|February|March|April|May|June|July|August|September|October|November|December', 
                                 months_str, re.IGNORECASE)
        
        return [m.capitalize() for m in month_names], year
    
    return None, None

def get_tracker_for_month(cos, month_name, month_year):
    """
    Get tracker file for a specific month based on the rule:
    - For June data: use tracker dated (DD-07-YYYY)
    - For July data: use tracker dated (DD-08-YYYY)
    - For August data: use tracker dated (DD-09-YYYY)
    And so on...
    """
    month_num = get_month_number(month_name)
    
    # Calculate the tracker month (next month)
    tracker_month = month_num + 1
    tracker_year = month_year
    
    # Handle year rollover
    if tracker_month > 12:
        tracker_month = 1
        tracker_year += 1
    
    logger.info(f"Looking for tracker for {month_name} {month_year} data -> Need tracker from {get_month_name(tracker_month)} {tracker_year}")
    
    # List all tracker files in Wave City Club folder
    wcc_files = list_files_in_folder(cos, "Wave City Club/")
    
    tracker_pattern = r'Structure Work Tracker.*\.xlsx$'
    
    matching_trackers = []
    
    for file_path in wcc_files:
        filename = os.path.basename(file_path)
        if re.search(tracker_pattern, filename, re.IGNORECASE):
            file_date = extract_date_from_filename(filename)
            if file_date:
                # Check if this tracker is from the correct month and year
                if file_date.month == tracker_month and file_date.year == tracker_year:
                    matching_trackers.append((file_path, file_date))
                    logger.info(f"Found matching tracker: {filename} ({file_date.strftime('%d-%m-%Y')})")
    
    if matching_trackers:
        # Get the latest tracker from the target month
        latest_tracker = max(matching_trackers, key=lambda x: x[1])
        logger.info(f"✅ Selected tracker for {month_name}: {latest_tracker[0]}")
        return latest_tracker[0]
    else:
        logger.warning(f"❌ No tracker found for {month_name} {month_year} (looking for {get_month_name(tracker_month)} {tracker_year})")
        return None

def setup_quarterly_months(kra_filename):
    """Setup months based on KRA filename"""
    global MONTHS, MONTH_YEARS, TARGET_END_MONTH, TARGET_END_YEAR, TRACKER_PATHS
    
    months_list, year = extract_months_from_kra_filename(kra_filename)
    
    if not months_list:
        logger.error("Could not extract months from KRA filename")
        return False
    
    logger.info(f"Extracted months from KRA: {months_list} for year {year}")
    
    MONTHS = months_list
    MONTH_YEARS = {month: year for month in months_list}
    
    # Handle year transition for December-January-February quarter
    if 'December' in months_list and 'January' in months_list:
        MONTH_YEARS['January'] = year + 1
        if 'February' in months_list:
            MONTH_YEARS['February'] = year + 1
    
    TARGET_END_MONTH = MONTHS[-1]
    TARGET_END_YEAR = MONTH_YEARS[TARGET_END_MONTH]
    
    logger.info(f"Setup complete - Months: {MONTHS}, Years: {MONTH_YEARS}")
    return True

def load_all_trackers(cos):
    """Pre-load all tracker workbooks to avoid repeated loading"""
    global LOADED_TRACKERS
    
    logger.info("=== PRE-LOADING ALL TRACKER WORKBOOKS ===")
    
    for month, tracker_path in TRACKER_PATHS.items():
        if tracker_path:
            try:
                logger.info(f"Loading tracker for {month}: {os.path.basename(tracker_path)}")
                raw = download_file_bytes(cos, tracker_path)
                wb = load_workbook(filename=BytesIO(raw), data_only=True)
                LOADED_TRACKERS[month] = wb
                logger.info(f"✅ Loaded tracker for {month}")
            except Exception as e:
                logger.error(f"❌ Error loading tracker for {month}: {e}")
                LOADED_TRACKERS[month] = None
        else:
            LOADED_TRACKERS[month] = None
            logger.warning(f"⚠️ No tracker available for {month}")
    
    logger.info(f"Loaded {sum(1 for v in LOADED_TRACKERS.values() if v is not None)} trackers")

# -----------------------------------------------------------------------------
# UTILITIES
# -----------------------------------------------------------------------------

def extract_percentage(cell_value):
    """Extract percentage value from cell, handling different formats"""
    if not cell_value or cell_value == '-':
        return 0.0
    
    if isinstance(cell_value, (int, float)):
        if cell_value <= 1.0:
            return cell_value * 100
        return cell_value
    
    val_str = str(cell_value).replace('%', '').strip()
    try:
        val = float(val_str)
        if val <= 1.0:
            return val * 100
        return val
    except ValueError:
        numbers = re.findall(r'\d+\.?\d*', val_str)
        if numbers:
            val = float(numbers[0])
            return val if val > 1.0 else val * 100
        return 0.0

def normalize_activity_name(activity):
    """Normalize activity name for better matching"""
    if not activity:
        return ""
    return str(activity).strip().lower()

def activities_match(target_activity, tracker_activity):
    """Enhanced matching with case-insensitive comparison"""
    if not target_activity or not tracker_activity:
        return False
    
    target = str(target_activity).strip()
    tracker = str(tracker_activity).strip()
    
    if target == tracker:
        return True
    
    if target.lower() == tracker.lower():
        logger.debug(f"CASE-INSENSITIVE MATCH: '{target}' matches '{tracker}'")
        return True
    
    logger.debug(f"NO MATCH: Target='{target}' vs Tracker='{tracker}'")
    return False

# -----------------------------------------------------------------------------
# DATA EXTRACTION FUNCTIONS
# -----------------------------------------------------------------------------

def detect_kra_column_mapping(sheet):
    """Detect month-to-column mapping from KRA file headers with support for various date formats"""
    month_to_col = {}
    
    for col_idx in range(2, 20):
        col_letter = get_column_letter(col_idx)
        header_cell = sheet[f'{col_letter}1']
        
        if header_cell.value:
            header_value = header_cell.value
            logger.info(f"KRA Header in column {col_letter}: '{header_value}' (type: {type(header_value).__name__})")
            
            # Handle different header formats
            month_name = None
            
            # Case 1: datetime object
            if isinstance(header_value, datetime):
                month_name = get_month_name(header_value.month)
                logger.info(f"  Detected datetime: {header_value}, Month: {month_name}")
            
            # Case 2: String with month name
            else:
                header_text = str(header_value).strip()
                
                # Try to extract month name directly
                for month in ["January", "February", "March", "April", "May", "June", 
                             "July", "August", "September", "October", "November", "December"]:
                    if month.lower() in header_text.lower():
                        month_name = month
                        logger.info(f"  Detected month from text: {month_name}")
                        break
                
                # Try to parse date strings like "Aug-25"
                if not month_name:
                    date_patterns = [
                        (r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[- ]?(\d{2})', 
                         {'Jan': 'January', 'Feb': 'February', 'Mar': 'March', 'Apr': 'April',
                          'May': 'May', 'Jun': 'June', 'Jul': 'July', 'Aug': 'August',
                          'Sep': 'September', 'Oct': 'October', 'Nov': 'November', 'Dec': 'December'})
                    ]
                    
                    for pattern, month_map in date_patterns:
                        match = re.search(pattern, header_text, re.IGNORECASE)
                        if match:
                            short_month = match.group(1).capitalize()
                            month_name = month_map.get(short_month)
                            if month_name:
                                logger.info(f"  Detected month from pattern: {month_name}")
                                break
            
            if month_name:
                month_to_col[month_name] = col_letter
                logger.info(f"✅ Mapped {month_name} -> Column {col_letter}")
    
    logger.info(f"Final column mapping: {month_to_col}")
    return month_to_col

def get_wcc_targets_from_kra(cos):
    """Extract targets from KRA file dynamically based on current months"""
    raw = download_file_bytes(cos, WCC_KRA_KEY)
    wb = load_workbook(filename=BytesIO(raw), data_only=True)
    
    # Try to find the correct sheet
    sheet = None
    possible_sheet_names = ['Wave City Club targets till Aug', 'Sheet1', wb.sheetnames[0]]
    
    for sheet_name in possible_sheet_names:
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            logger.info(f"Using KRA sheet: {sheet_name}")
            break
    
    if not sheet:
        logger.error("Could not find appropriate sheet in KRA file")
        return {}
    
    targets = {}
    logger.info("=== EXTRACTING TARGETS FROM KRA FILE ===")
    logger.info(f"Looking for months: {MONTHS} with years: {MONTH_YEARS}")
    
    # Detect column mapping from file
    month_to_col = detect_kra_column_mapping(sheet)
    
    if not month_to_col:
        logger.error("Could not detect column mapping from headers")
        return {}
    
    # Check if we have columns for all our target months
    missing_months = [m for m in MONTHS if m not in month_to_col]
    if missing_months:
        logger.warning(f"Missing columns for months: {missing_months}")
    
    # Read targets from the KRA file
    for row_num in range(2, sheet.max_row + 1):
        block_cell = sheet[f'A{row_num}']
        
        if block_cell.value:
            block_name = str(block_cell.value).strip()
            month_activities = {}
            
            for month in MONTHS:
                col = month_to_col.get(month)
                if col:
                    cell = sheet[f'{col}{row_num}']
                    activity = str(cell.value or '').strip() if cell.value else ''
                    month_activities[month] = activity
                    logger.debug(f"Row {row_num}, {month}: Block='{block_name}', Activity='{activity}'")
                else:
                    logger.warning(f"No column found for month {month}")
                    month_activities[month] = ''
            
            targets[block_name] = month_activities
    
    logger.info(f"Extracted targets for {len(targets)} blocks")
    return targets

def find_activity_progress_in_sheet(sheet, target_activity, sheet_name, block_name=None, month_name=None):
    """Find activity progress with enhanced search for special blocks"""
    logger.info(f"=== Looking for '{target_activity}' in '{sheet_name}' for '{block_name}' ({month_name}) ===")
    
    # Return 100% when there's no target activity
    if not target_activity or target_activity.strip() == '' or target_activity.lower() in ['no target', '-'] or 'no target for' in target_activity.lower():
        logger.info(f"No target activity, returning 100%")
        return 100.0
    
    # Enhanced search for special blocks
    if block_name in SPECIAL_BLOCKS_ENHANCED_SEARCH:
        logger.info(f"SPECIAL CASE: {block_name} - enhanced search")
        max_rows = min(sheet.max_row, 60)
        
        for row_num in range(1, max_rows + 1):
            try:
                activity_cell = sheet[f'G{row_num}']
                if activity_cell.value:
                    tracker_activity = str(activity_cell.value).strip()
                    
                    if activities_match(target_activity, tracker_activity):
                        progress_cell = sheet[f'AC{row_num}']
                        ac_value = progress_cell.value
                        logger.info(f"✅ MATCH in G{row_num}: '{tracker_activity}', AC{row_num}: {ac_value}")
                        
                        if ac_value is not None:
                            percentage = extract_percentage(ac_value)
                            logger.info(f"   Progress for {month_name}: {percentage}%")
                            return percentage
                        return 0.0
            except Exception as e:
                logger.debug(f"Error at row {row_num}: {e}")
                continue
        
        logger.warning(f"❌ NO MATCH for '{target_activity}'")
        return 0.0
    
    # Standard search for other blocks
    max_rows = min(sheet.max_row, 20)
    for row_num in range(1, max_rows + 1):
        try:
            activity_cell = sheet[f'G{row_num}']
            if activity_cell.value:
                tracker_activity = str(activity_cell.value).strip()
                
                if activities_match(target_activity, tracker_activity):
                    progress_cell = sheet[f'AC{row_num}']
                    ac_value = progress_cell.value
                    logger.info(f"✅ MATCH in G{row_num}: '{tracker_activity}', AC{row_num}: {ac_value}")
                    
                    if ac_value is not None:
                        percentage = extract_percentage(ac_value)
                        logger.info(f"   Progress for {month_name}: {percentage}%")
                        return percentage
                    return 0.0
        except Exception as e:
            logger.debug(f"Error at row {row_num}: {e}")
            continue
    
    logger.warning(f"❌ NO MATCH for '{target_activity}'")
    return 0.0

def get_wcc_progress_from_tracker_all_months(cos, targets):
    """Extract progress data from tracker files for each month"""
    progress_data = []
    milestone_counter = 1
    total_blocks = len(targets)
    site_weighted = round(100 / total_blocks, 2) if total_blocks > 0 else 0
    
    # Pre-load all trackers
    load_all_trackers(cos)
    
    for block_name, month_activities in targets.items():
        logger.info(f"\n{'='*80}")
        logger.info(f"Processing block: {block_name}")
        logger.info(f"{'='*80}")
        
        sheet_name = BLOCK_MAPPING.get(block_name)
        month_progress = {}
        
        if not sheet_name:
            logger.warning(f"No sheet mapping for block: {block_name}")
            for month in MONTHS:
                month_progress[month] = 0.0
        else:
            # Process each month with its specific tracker
            for month in MONTHS:
                month_year = MONTH_YEARS[month]
                tracker_wb = LOADED_TRACKERS.get(month)
                activity = month_activities.get(month, '')
                
                logger.info(f"\n--- Processing {month} {month_year} ---")
                logger.info(f"Target Activity: '{activity}'")
                logger.info(f"Tracker File: {os.path.basename(TRACKER_PATHS.get(month, 'N/A'))}")
                
                if not tracker_wb:
                    logger.warning(f"No tracker workbook available for {month}")
                    month_progress[month] = None
                    continue
                
                if sheet_name not in tracker_wb.sheetnames:
                    logger.warning(f"Sheet '{sheet_name}' not found in tracker for {month}")
                    month_progress[month] = 0.0
                    continue
                
                try:
                    sheet = tracker_wb[sheet_name]
                    progress = find_activity_progress_in_sheet(
                        sheet, activity, sheet_name, block_name, month
                    )
                    month_progress[month] = progress
                    
                    logger.info(f"✅ {block_name} - {month}: {progress:.1f}%")
                    
                except Exception as e:
                    logger.error(f"Error processing {month} for {block_name}: {e}")
                    month_progress[month] = 0.0
        
        # Use the last available month for weighted calculation
        last_month = MONTHS[-1] if MONTHS else ''
        last_progress = month_progress.get(last_month, 0.0)
        if last_progress is None:
            last_progress = 0.0
        
        main_weighted = round((site_weighted * last_progress) / 100, 3)
        
        # Create row data with dynamic columns
        row_data = {
            'Milestone': f"Milestone-{milestone_counter:02d}",
            'Block': block_name,
            f'Target to be complete by {TARGET_END_MONTH}-{TARGET_END_YEAR}': month_activities.get(last_month, ''),
            'Site Weighted': site_weighted,
            'Weighted progress against target': main_weighted,
        }
        
        # Add month-specific columns
        for month in MONTHS:
            year = MONTH_YEARS[month]
            target_val = month_activities.get(month, '')
            progress_val = month_progress.get(month)
            
            row_data[f'Target - {month}-{year}'] = target_val
            
            # Handle cases where tracker is not available
            if progress_val is None:
                row_data[f'% work done- {month} Status'] = 'N/A'
                row_data[f'Achieved- {month} {year}'] = 'Tracker not available'
            else:
                row_data[f'% work done- {month} Status'] = f"{progress_val:.0f}%"
                
                if progress_val == 100:
                    achieved = target_val if target_val else f'No target for {month}'
                elif progress_val == 0:
                    achieved = 'No progress' if target_val else f'No target for {month}'
                else:
                    achieved = f'{progress_val:.0f}% completed'
                
                row_data[f'Achieved- {month} {year}'] = achieved
            
            row_data[f'Responsible Person- {month}'] = ''
            row_data[f'Delay Reasons- {month}'] = ''
        
        progress_data.append(row_data)
        milestone_counter += 1
    
    # Create DataFrame with dynamic columns
    columns = [
        'Milestone', 
        'Block', 
        f'Target to be complete by {TARGET_END_MONTH}-{TARGET_END_YEAR}'
    ]
    
    for month in MONTHS:
        year = MONTH_YEARS[month]
        columns.extend([
            f'Target - {month}-{year}',
            f'% work done- {month} Status',
            f'Achieved- {month} {year}',
            f'Responsible Person- {month}',
            f'Delay Reasons- {month}'
        ])
    
    columns.extend(['Site Weighted', 'Weighted progress against target'])
    
    df = pd.DataFrame(progress_data, columns=columns)
    logger.info(f"\nCreated DataFrame with {len(df)} rows for months: {MONTHS}")
    return df

def apply_manual_overrides(df):
    """Apply manual overrides for specific activities in specific months"""
    logger.info("\n=== APPLYING MANUAL OVERRIDES ===")
    
    # Define overrides: (Block name, Month, Activity pattern, Override percentage)
    overrides = [
        ('Fine Dine', 'July', 'Brick Work GF', 100.0)
    ]
    
    for block_name, month, activity_pattern, override_value in overrides:
        # Find the row for this block
        block_rows = df[df['Block'] == block_name]
        
        if block_rows.empty:
            logger.warning(f"Block '{block_name}' not found for override")
            continue
        
        # Get the row index
        row_idx = block_rows.index[0]
        
        # Check if the target activity matches
        month_year = MONTH_YEARS.get(month)
        if not month_year:
            logger.warning(f"Month '{month}' not in current quarter")
            continue
        
        target_col = f'Target - {month}-{month_year}'
        status_col = f'% work done- {month} Status'
        achieved_col = f'Achieved- {month} {month_year}'
        
        # Check if the activity matches
        current_target = df.at[row_idx, target_col]
        
        if activity_pattern.lower() in str(current_target).lower():
            logger.info(f"✅ Applying override: {block_name} - {month} - '{current_target}' -> {override_value}%")
            
            # Update the status column
            df.at[row_idx, status_col] = f"{override_value:.0f}%"
            
            # Update the achieved column
            if override_value == 100:
                df.at[row_idx, achieved_col] = current_target
            elif override_value == 0:
                df.at[row_idx, achieved_col] = 'No progress'
            else:
                df.at[row_idx, achieved_col] = f'{override_value:.0f}% completed'
            
            # Recalculate weighted progress if this is the last month
            if month == MONTHS[-1]:
                site_weighted = df.at[row_idx, 'Site Weighted']
                new_weighted = round((site_weighted * override_value) / 100, 3)
                df.at[row_idx, 'Weighted progress against target'] = new_weighted
                logger.info(f"   Updated weighted progress: {new_weighted:.3f}%")
        else:
            logger.warning(f"Activity pattern '{activity_pattern}' not found in '{current_target}' for {block_name} - {month}")
    
    logger.info("=== MANUAL OVERRIDES COMPLETE ===\n")
    return df

# -----------------------------------------------------------------------------
# EXCEL REPORT GENERATION
# -----------------------------------------------------------------------------

def write_wcc_excel_report_consolidated(df, filename):
    """Generate formatted Excel report with dynamic month columns"""
    wb = Workbook()
    ws = wb.active
    ws.title = 'WCC Progress'
    
    # Add main title
    title_row = ["Wave City Club- Progress Against Milestones"]
    ws.append(title_row)
    ws.merge_cells(f'A1:{get_column_letter(len(df.columns))}1')
    
    # Add date row
    current_date = datetime.now().strftime("%d-%m-%Y")
    date_row = [f"Report Generated on: {current_date}"]
    ws.append(date_row)
    ws.merge_cells(f'A2:{get_column_letter(len(df.columns))}2')
    
    # Add month info row
    month_info = f"Months Covered: {', '.join([f'{m} {MONTH_YEARS[m]}' for m in MONTHS])}"
    month_info_row = [month_info]
    ws.append(month_info_row)
    ws.merge_cells(f'A3:{get_column_letter(len(df.columns))}3')
    
    # Add empty row
    ws.append([])
    
    # Add DataFrame data with percentage formatting for weighted progress
    for row in dataframe_to_rows(df, index=False, header=True):
        # Format the weighted progress column (last column) to add % symbol
        if len(row) > 0 and isinstance(row[-1], (int, float)) and row[-1] != '':
            row[-1] = f"{row[-1]:.3f}%"
        ws.append(row)
    
    # Add Sum row - Only for the weighted progress column
    weighted_sum = df['Weighted progress against target'].sum()
    
    # Create sum row with blanks for all columns except the weighted progress column
    sum_row = [''] * len(df.columns)
    sum_row[-2] = 'Sum'  # Site Weighted column
    sum_row[-1] = f'{weighted_sum:.3f}%'  # Weighted progress column
    ws.append(sum_row)
    
    # Define styles
    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True, size=8)
    normal_font = Font(bold=False, size=8)
    date_font = Font(bold=False, size=10, color="666666")
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='000000')
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    light_grey_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    light_blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
    
    # Style title (light grey background)
    ws['A1'].font = title_font
    ws['A1'].alignment = center
    ws['A1'].fill = light_grey_fill
    
    # Style date row
    ws['A2'].font = date_font
    ws['A2'].alignment = center
    
    # Style month info row
    ws['A3'].font = date_font
    ws['A3'].alignment = center
    
    # Style header row (row 5) with light grey background
    header_row = 5
    for cell in ws[header_row]:
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        cell.fill = light_grey_fill
    
    # Style data rows
    data_start = 6
    data_end = ws.max_row - 1  # Exclude sum row for now
    
    for row_num in range(data_start, data_end + 1):
        for col_num in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = normal_font
            cell.border = border
            
            # Alignment based on column type
            if col_num in [1, 2, 3] or 'Target' in str(ws.cell(row=header_row, column=col_num).value or ''):
                cell.alignment = left
            else:
                cell.alignment = center
    
    # Style sum row with light blue background
    sum_row_num = ws.max_row
    for col_num in range(1, len(df.columns) + 1):
        cell = ws.cell(row=sum_row_num, column=col_num)
        cell.font = header_font
        cell.border = border
        cell.fill = light_blue_fill
        cell.alignment = center
    
    # Dynamic column width adjustment
    for col_num in range(1, len(df.columns) + 1):
        col_letter = get_column_letter(col_num)
        
        # Calculate optimal width based on column content
        max_length = 0
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=col_num, max_col=col_num):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        
        # Set minimum and maximum width constraints
        calculated_width = min(max(max_length + 2, 8), 15)
        ws.column_dimensions[col_letter].width = calculated_width
    
    # Set row heights
    ws.row_dimensions[1].height = 25  # Title row
    ws.row_dimensions[2].height = 20  # Date row
    ws.row_dimensions[3].height = 20  # Month info row
    for i in range(5, ws.max_row + 1):
        ws.row_dimensions[i].height = 25
    
    wb.save(filename)
    logger.info(f'Dynamic report saved to {filename}')

def get_unique_filename(base_name):
    """If file exists, append (1), (2), etc."""
    if not os.path.exists(base_name):
        return base_name

    name, ext = os.path.splitext(base_name)
    counter = 1
    new_name = f"{name}({counter}){ext}"
    while os.path.exists(new_name):
        counter += 1
        new_name = f"{name}({counter}){ext}"
    return new_name

# -----------------------------------------------------------------------------
# MAIN FUNCTION
# -----------------------------------------------------------------------------

def main():
    """Main execution function for quarterly report generation"""
    logger.info("=== STARTING WAVE CITY CLUB QUARTERLY REPORT GENERATION ===")
    
    try:
        # Initialize COS client
        cos = init_cos()
        
        # Step 1: Get latest KRA file
        logger.info("\n=== STEP 1: Finding Latest KRA File ===")
        kra_file = get_latest_kra_file(cos)
        
        if not kra_file:
            logger.error("❌ Failed to find KRA file")
            return
        
        logger.info(f"✅ Using KRA file: {kra_file}")
        
        # Step 2: Extract months from KRA filename and setup
        logger.info("\n=== STEP 2: Setting Up Quarterly Months ===")
        if not setup_quarterly_months(os.path.basename(kra_file)):
            logger.error("❌ Failed to setup quarterly months")
            return
        
        logger.info(f"✅ Quarterly months: {MONTHS}")
        logger.info(f"✅ Month-Year mapping: {MONTH_YEARS}")
        
        # Step 3: Find appropriate trackers for each month
        logger.info("\n=== STEP 3: Finding Trackers for Each Month ===")
        global TRACKER_PATHS
        
        for month in MONTHS:
            month_year = MONTH_YEARS[month]
            tracker_path = get_tracker_for_month(cos, month, month_year)
            
            if tracker_path:
                TRACKER_PATHS[month] = tracker_path
                logger.info(f"✅ {month} {month_year}: {tracker_path}")
            else:
                logger.warning(f"⚠️ {month} {month_year}: No tracker found - will show as N/A in report")
                TRACKER_PATHS[month] = None
        
        # Check if we have at least one tracker
        if not any(TRACKER_PATHS.values()):
            logger.error("❌ No trackers found for any month")
            return
        
        logger.info(f"✅ Found trackers for {sum(1 for v in TRACKER_PATHS.values() if v)} out of {len(MONTHS)} months")
        
        # Step 4: Get targets from KRA file
        logger.info("\n=== STEP 4: Extracting Targets from KRA File ===")
        targets = get_wcc_targets_from_kra(cos)
        
        if not targets:
            logger.error("❌ Failed to extract targets from KRA file")
            return
        
        logger.info(f"✅ Extracted targets for {len(targets)} blocks")
        
        # Step 5: Extract progress data for all months
        logger.info("\n=== STEP 5: Extracting Progress Data from Trackers ===")
        df = get_wcc_progress_from_tracker_all_months(cos, targets)
        
        if df.empty:
            logger.error("❌ Failed to generate progress data")
            return
        
        logger.info(f"✅ Generated progress data for {len(df)} milestones")
        
        # Step 5.5: Apply manual overrides
        logger.info("\n=== STEP 5.5: Applying Manual Overrides ===")
        df = apply_manual_overrides(df)
        
        # Step 6: Generate Excel report
        logger.info("\n=== STEP 6: Generating Excel Report ===")
        current_date_for_filename = datetime.now().strftime('%d-%m-%Y')
        
        # Create filename with quarter months
        quarter_str = "_".join(MONTHS)
        base_filename = f"Wave_City_Club_Milestone_Report_{quarter_str}_{current_date_for_filename}.xlsx"
        filename = get_unique_filename(base_filename)
        
        write_wcc_excel_report_consolidated(df, filename)
        
        logger.info("\n=== WAVE CITY CLUB QUARTERLY REPORT GENERATION COMPLETE ===")
        logger.info(f"✅ Report saved as: {filename}")
        
        # Log summary
        logger.info("\n=== REPORT SUMMARY ===")
        logger.info(f"  Quarter Months: {', '.join([f'{m} {MONTH_YEARS[m]}' for m in MONTHS])}")
        logger.info(f"  KRA File: {os.path.basename(kra_file)}")
        logger.info(f"  Total Blocks: {len(targets)}")
        logger.info(f"  Trackers Used:")
        for month in MONTHS:
            tracker = TRACKER_PATHS.get(month)
            if tracker:
                tracker_date = extract_date_from_filename(os.path.basename(tracker))
                logger.info(f"    - {month} {MONTH_YEARS[month]}: {os.path.basename(tracker)} (dated {tracker_date.strftime('%d-%m-%Y') if tracker_date else 'Unknown'})")
            else:
                logger.info(f"    - {month} {MONTH_YEARS[month]}: Not Available")
        logger.info(f"  Total Milestones: {len(df)}")
        
        # Calculate overall progress
        total_weighted = df['Weighted progress against target'].sum()
        logger.info(f"  Overall Weighted Progress: {total_weighted:.2f}%")
        
        # Show sample data for verification
        logger.info("\n=== SAMPLE DATA VERIFICATION ===")
        for idx, row in df.head(2).iterrows():
            logger.info(f"\nBlock: {row['Block']}")
            for month in MONTHS:
                target = row[f'Target - {month}-{MONTH_YEARS[month]}']
                status = row[f'% work done- {month} Status']
                logger.info(f"  {month}: Target='{target}', Status={status}")
        
    except Exception as e:
        logger.error(f"❌ Error in main execution: {e}")
        import traceback
        logger.error(traceback.format_exc())
        raise

if __name__ == "__main__":
    main()
