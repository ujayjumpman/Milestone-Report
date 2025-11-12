# """
# Automated Milestone Report Generator - FIXED VERSION
# Hardcoded column numbers for stability, improved activity parsing
# """

# import os
# import logging
# from io import BytesIO
# from datetime import datetime
# import pandas as pd
# from openpyxl import load_workbook, Workbook
# from openpyxl.utils.dataframe import dataframe_to_rows
# from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
# from openpyxl.utils import get_column_letter
# from dotenv import load_dotenv
# import ibm_boto3
# from ibm_botocore.client import Config
# import re
# from typing import Optional, Tuple, List, Dict, Any

# # ======================= CONFIGURATION =======================
# load_dotenv()
# logging.basicConfig(level=logging.DEBUG, format="%(asctime)s [%(levelname)s] %(message)s")
# logger = logging.getLogger(__name__)

# # Cloud Storage Configuration
# COS_API_KEY = os.getenv("COS_API_KEY")
# COS_CRN = os.getenv("COS_SERVICE_INSTANCE_CRN")
# COS_ENDPOINT = os.getenv("COS_ENDPOINT")
# BUCKET = os.getenv("COS_BUCKET_NAME")
# KRA_FOLDER = os.getenv("KRA_FOLDER", "")
# EDEN_TRACKER_FOLDER = os.getenv("EDEN_TRACKER_FOLDER", "Eden/")

# # ======================= HARDCODED COLUMN MAPPINGS =======================
# # KRA Sheet Columns (1-indexed for openpyxl)
# KRA_TOWER_COL = 1  # Column A: Tower name

# KRA_SEP_ACTIVITY_COL = 2   # Column B: September activity
# KRA_SEP_TARGET_COL = 3     # Column C: September % target

# KRA_OCT_ACTIVITY_COL = 4   # Column D: October activity
# KRA_OCT_TARGET_COL = 5     # Column E: October % target

# KRA_NOV_ACTIVITY_COL = 6   # Column F: November activity
# KRA_NOV_TARGET_COL = 7     # Column G: November % target

# # Tracker Sheet Columns (1-indexed for openpyxl)
# TRACKER_TOWER_COL = 1           # Column A: Tower number
# TRACKER_ACTIVITY_NO_COL = 2     # Column B: Activity number
# TRACKER_LOOKAHEAD_COL = 3       # Column C: Monthly lookahead ID
# TRACKER_TASK_NAME_COL = 4       # Column D: Task name
# TRACKER_ACTUAL_START_COL = 5    # Column E: Actual start
# TRACKER_ACTUAL_FINISH_COL = 6   # Column F: Actual finish
# TRACKER_PCT_COMPLETE_COL = 7    # Column G: % Complete ← THE KEY COLUMN
# TRACKER_DURATION_COL = 8        # Column H: Duration

# # Quarterly structure
# QUARTERS = {
#     "Q1": ["June", "July", "August"],
#     "Q2": ["September", "October", "November"],
#     "Q3": ["December", "January", "February"],
#     "Q4": ["March", "April", "May"]
# }

# # Month to tracker month mapping
# MONTH_TO_TRACKER_MAPPING = {
#     "June": 7, "July": 8, "August": 9,
#     "September": 10, "October": 11, "November": 12,
#     "December": 1, "January": 2, "February": 3,
#     "March": 4, "April": 5, "May": 6
# }

# # Validate environment variables
# required_vars = {
#     'COS_API_KEY': COS_API_KEY,
#     'COS_SERVICE_INSTANCE_CRN': COS_CRN,
#     'COS_ENDPOINT': COS_ENDPOINT,
#     'COS_BUCKET_NAME': BUCKET
# }

# missing_vars = [var_name for var_name, var_value in required_vars.items() if not var_value]
# if missing_vars:
#     error_msg = f"Missing required environment variables: {', '.join(missing_vars)}"
#     logger.error(error_msg)
#     raise ValueError(error_msg)

# # ======================= CLOUD STORAGE HELPERS =======================

# def init_cos():
#     """Initialize IBM Cloud Object Storage client."""
#     return ibm_boto3.client(
#         "s3",
#         ibm_api_key_id=COS_API_KEY,
#         ibm_service_instance_id=COS_CRN,
#         config=Config(signature_version="oauth"),
#         endpoint_url=COS_ENDPOINT
#     )

# def download_file_bytes(cos, key: str) -> bytes:
#     """Download file from cloud storage as bytes."""
#     return cos.get_object(Bucket=BUCKET, Key=key)["Body"].read()

# # ======================= FILE DISCOVERY =======================

# def find_latest_kra_file(cos_client, bucket_name: str, folder_prefix: str = "") -> Optional[Tuple[str, List[str], int]]:
#     """Find the latest KRA Milestones file."""
#     logger.info(f"\n{'='*70}")
#     logger.info(f"SEARCHING FOR LATEST KRA FILE")
#     logger.info(f"{'='*70}")
    
#     try:
#         response = cos_client.list_objects_v2(Bucket=bucket_name, Prefix=folder_prefix)
        
#         if 'Contents' not in response:
#             logger.error(f"No files found in folder '{folder_prefix}'")
#             return None
        
#         kra_files = []
        
#         for obj in response['Contents']:
#             file_key = obj['Key']
#             filename = os.path.basename(file_key)
#             filename_lower = filename.lower()
            
#             if file_key.endswith('/'):
#                 continue
            
#             is_kra = 'kra' in filename_lower and 'milestone' in filename_lower
#             is_excel = filename_lower.endswith(('.xlsx', '.xls'))
            
#             if is_kra and is_excel:
#                 months_pattern = r'(January|February|March|April|May|June|July|August|September|October|November|December)'
#                 found_months = re.findall(months_pattern, filename, re.IGNORECASE)
#                 found_months = [m.capitalize() for m in found_months]
                
#                 year_match = re.search(r'(\d{4})', filename)
#                 year = int(year_match.group(1)) if year_match else datetime.now().year
                
#                 kra_files.append({
#                     'key': file_key,
#                     'filename': filename,
#                     'months': found_months,
#                     'year': year,
#                     'last_modified': obj['LastModified']
#                 })
                
#                 logger.info(f"Found: {filename}")
        
#         if not kra_files:
#             logger.error("No KRA Milestone files found")
#             return None
        
#         kra_files.sort(key=lambda f: f['last_modified'], reverse=True)
#         latest = kra_files[0]
        
#         logger.info(f"\nSelected: {latest['filename']}")
#         logger.info(f"Months: {', '.join(latest['months'])}")
#         logger.info(f"Year: {latest['year']}")
        
#         return latest['key'], latest['months'], latest['year']
        
#     except Exception as e:
#         logger.error(f"Error searching for KRA file: {str(e)}")
#         raise

# def calculate_tracker_year(report_month: str, kra_year: int) -> int:
#     """Calculate correct year for tracker file."""
#     tracker_month_num = MONTH_TO_TRACKER_MAPPING.get(report_month)
    
#     if not tracker_month_num:
#         return kra_year
    
#     if report_month == "December" and tracker_month_num == 1:
#         return kra_year + 1
    
#     if report_month in ["January", "February"] and tracker_month_num in [2, 3]:
#         return kra_year
    
#     return kra_year

# def find_tracker_for_month(cos_client, bucket_name: str, target_month: int, target_year: int,
#                           folder_prefix: str = "Eden/") -> Optional[str]:
#     """Find tracker file for specific month and year."""
#     logger.info(f"  Searching for tracker: Month {target_month}/{target_year}")
    
#     try:
#         response = cos_client.list_objects_v2(Bucket=bucket_name, Prefix=folder_prefix)
        
#         if 'Contents' not in response:
#             return None
        
#         matching_trackers = []
        
#         for obj in response['Contents']:
#             file_key = obj['Key']
#             filename = os.path.basename(file_key)
#             filename_lower = filename.lower()
            
#             if file_key.endswith('/'):
#                 continue
            
#             is_tracker = any(pattern in filename_lower for pattern in 
#                            ['structure work tracker', 'tracker', 'structure tracker'])
#             is_excel = filename_lower.endswith(('.xlsx', '.xls'))
            
#             if is_tracker and is_excel:
#                 date_pattern = r'\((\d{1,2})-(\d{1,2})-(\d{2,4})\)'
#                 date_match = re.search(date_pattern, filename)
                
#                 if date_match:
#                     day, month, year = date_match.groups()
#                     file_month = int(month)
#                     file_year = int(year)
                    
#                     if file_year < 100:
#                         file_year += 2000
                    
#                     if file_month == target_month and file_year == target_year:
#                         matching_trackers.append({
#                             'key': file_key,
#                             'filename': filename,
#                             'day': int(day)
#                         })
        
#         if not matching_trackers:
#             return None
        
#         matching_trackers.sort(key=lambda t: t['day'], reverse=True)
#         return matching_trackers[0]['key']
        
#     except Exception as e:
#         logger.error(f"Error searching for tracker: {str(e)}")
#         return None

# # ======================= KRA DATA EXTRACTION =======================

# def find_project_sheet(workbook, project_name: str):
#     """Find sheet containing project name."""
#     for sheet_name in workbook.sheetnames:
#         if project_name.upper() in sheet_name.upper():
#             return workbook[sheet_name]
#     return None

# class ActivityTarget:
#     """Represents a target activity from KRA."""
    
#     def __init__(self, tower: str, activity_text: str, target_pct: float, month: str):
#         self.tower = tower
#         self.activity_text = activity_text  # Full text as it appears
#         self.target_pct = target_pct
#         self.month = month
#         self.actual_pct = 0.0
#         self.status = ""
    
#     def __repr__(self):
#         return f"{self.tower} | {self.month} | {self.activity_text} ({self.target_pct}%)"

# def parse_kra_targets(worksheet) -> Dict[str, List[ActivityTarget]]:
#     """
#     Parse ALL targets from KRA sheet for all towers.
#     Returns: Dict[tower_name, List[ActivityTarget]]
#     """
#     logger.info("\n" + "="*70)
#     logger.info("PARSING KRA TARGETS")
#     logger.info("="*70)
    
#     tower_targets = {}
#     current_tower = None
    
#     # Start from row 5 (after header row 4)
#     for row_idx in range(5, worksheet.max_row + 1):
#         tower_cell = worksheet.cell(row_idx, KRA_TOWER_COL).value
        
#         # Check if this is a new tower row
#         if tower_cell and str(tower_cell).strip():
#             tower_name = str(tower_cell).strip()
            
#             # Only process actual towers (not finishing milestones)
#             if 'Tower' in tower_name or 'NTA' in tower_name:
#                 if 'Finishing' not in tower_name:
#                     current_tower = tower_name
#                     if current_tower not in tower_targets:
#                         tower_targets[current_tower] = []
                    
#                     logger.info(f"\nProcessing: {current_tower}")
                    
#                     # Parse September target
#                     sep_activity = worksheet.cell(row_idx, KRA_SEP_ACTIVITY_COL).value
#                     sep_target = worksheet.cell(row_idx, KRA_SEP_TARGET_COL).value
                    
#                     if sep_target and isinstance(sep_target, (int, float)) and sep_target > 0:
#                         activity_text = str(sep_activity).strip() if sep_activity else "Activity"
#                         target = ActivityTarget(current_tower, activity_text, float(sep_target) * 100, "September")
#                         tower_targets[current_tower].append(target)
#                         logger.info(f"  September: {activity_text} → {sep_target*100}%")
                    
#                     # Parse October target
#                     oct_activity = worksheet.cell(row_idx, KRA_OCT_ACTIVITY_COL).value
#                     oct_target = worksheet.cell(row_idx, KRA_OCT_TARGET_COL).value
                    
#                     if oct_target and isinstance(oct_target, (int, float)) and oct_target > 0:
#                         activity_text = str(oct_activity).strip() if oct_activity else "Activity"
#                         target = ActivityTarget(current_tower, activity_text, float(oct_target) * 100, "October")
#                         tower_targets[current_tower].append(target)
#                         logger.info(f"  October: {activity_text} → {oct_target*100}%")
                    
#                     # Parse November target
#                     nov_activity = worksheet.cell(row_idx, KRA_NOV_ACTIVITY_COL).value
#                     nov_target = worksheet.cell(row_idx, KRA_NOV_TARGET_COL).value
                    
#                     if nov_target and isinstance(nov_target, (int, float)) and nov_target > 0:
#                         activity_text = str(nov_activity).strip() if nov_activity else "Activity"
#                         target = ActivityTarget(current_tower, activity_text, float(nov_target) * 100, "November")
#                         tower_targets[current_tower].append(target)
#                         logger.info(f"  November: {activity_text} → {nov_target*100}%")
#                 else:
#                     current_tower = None  # Stop processing finishing milestones
        
#         # Check sub-rows for multi-line activities (Tower 5, Tower 7 pattern)
#         elif current_tower:
#             # Check September sub-activity
#             sep_activity = worksheet.cell(row_idx, KRA_SEP_ACTIVITY_COL).value
#             sep_target = worksheet.cell(row_idx, KRA_SEP_TARGET_COL).value
            
#             if sep_target and isinstance(sep_target, (int, float)) and sep_target > 0:
#                 # Build hierarchical activity text
#                 activity_parts = []
#                 for back_row in range(max(row_idx - 3, 5), row_idx + 1):
#                     cell_val = worksheet.cell(back_row, KRA_SEP_ACTIVITY_COL).value
#                     if cell_val and str(cell_val).strip() and str(cell_val).strip() != current_tower:
#                         activity_parts.append(str(cell_val).strip())
                
#                 activity_text = "\n".join(activity_parts) if activity_parts else str(sep_activity).strip()
#                 target = ActivityTarget(current_tower, activity_text, float(sep_target) * 100, "September")
#                 tower_targets[current_tower].append(target)
#                 logger.info(f"  September (sub): {activity_text[:50]}... → {sep_target*100}%")
            
#             # Check October sub-activity
#             oct_activity = worksheet.cell(row_idx, KRA_OCT_ACTIVITY_COL).value
#             oct_target = worksheet.cell(row_idx, KRA_OCT_TARGET_COL).value
            
#             if oct_target and isinstance(oct_target, (int, float)) and oct_target > 0:
#                 activity_parts = []
#                 for back_row in range(max(row_idx - 3, 5), row_idx + 1):
#                     cell_val = worksheet.cell(back_row, KRA_OCT_ACTIVITY_COL).value
#                     if cell_val and str(cell_val).strip() and str(cell_val).strip() != current_tower:
#                         activity_parts.append(str(cell_val).strip())
                
#                 activity_text = "\n".join(activity_parts) if activity_parts else str(oct_activity).strip()
#                 target = ActivityTarget(current_tower, activity_text, float(oct_target) * 100, "October")
#                 tower_targets[current_tower].append(target)
#                 logger.info(f"  October (sub): {activity_text[:50]}... → {oct_target*100}%")
            
#             # Check November sub-activity
#             nov_activity = worksheet.cell(row_idx, KRA_NOV_ACTIVITY_COL).value
#             nov_target = worksheet.cell(row_idx, KRA_NOV_TARGET_COL).value
            
#             if nov_target and isinstance(nov_target, (int, float)) and nov_target > 0:
#                 activity_parts = []
#                 for back_row in range(max(row_idx - 3, 5), row_idx + 1):
#                     cell_val = worksheet.cell(back_row, KRA_NOV_ACTIVITY_COL).value
#                     if cell_val and str(cell_val).strip() and str(cell_val).strip() != current_tower:
#                         activity_parts.append(str(cell_val).strip())
                
#                 activity_text = "\n".join(activity_parts) if activity_parts else str(nov_activity).strip()
#                 target = ActivityTarget(current_tower, activity_text, float(nov_target) * 100, "November")
#                 tower_targets[current_tower].append(target)
#                 logger.info(f"  November (sub): {activity_text[:50]}... → {nov_target*100}%")
    
#     logger.info(f"\nTotal towers with targets: {len(tower_targets)}")
#     for tower, targets in tower_targets.items():
#         logger.info(f"  {tower}: {len(targets)} targets")
    
#     return tower_targets

# # ======================= TRACKER DATA EXTRACTION =======================

# def normalize_text(text: str) -> str:
#     """Normalize text for matching."""
#     if not text:
#         return ""
#     # Convert to lowercase, remove extra spaces, remove special chars
#     text = re.sub(r'\s+', ' ', str(text).lower().strip())
#     text = re.sub(r'[^\w\s]', ' ', text)
#     return ' '.join(text.split())

# def find_activity_in_tracker(tracker_wb, tower_name: str, activity_text: str) -> Optional[float]:
#     """
#     Find matching activity in tracker and return % complete.
#     Uses Column G (TRACKER_PCT_COMPLETE_COL = 7) for % Complete.
    
#     HIERARCHICAL MATCHING: Find parent level first, then search for child within next 10 rows.
#     """
#     # Find tower sheet
#     tower_sheet = None
#     for sheet_name in tracker_wb.sheetnames:
#         if tower_name.replace("Tower ", "").replace("NTA ", "") in sheet_name:
#             tower_sheet = tracker_wb[sheet_name]
#             break
    
#     if not tower_sheet:
#         logger.debug(f"    Sheet not found for {tower_name}")
#         return None
    
#     # Split activity text into hierarchy levels
#     activity_lines = [line.strip() for line in activity_text.split('\n') if line.strip()]
    
#     if not activity_lines:
#         return None
    
#     logger.debug(f"    Searching in {tower_sheet.title}")
#     logger.debug(f"    Hierarchy: {' → '.join(activity_lines)}")
    
#     # Strategy: Find the PARENT level first (e.g., "Upper Basement")
#     # Then find the CHILD within the next 10 rows (e.g., "Casting Work")
    
#     parent_term = normalize_text(activity_lines[0])  # e.g., "upper basement"
    
#     # Define conflicting terms for the parent
#     conflicting_terms = []
#     if 'upper' in parent_term:
#         conflicting_terms.append('lower')
#     elif 'lower' in parent_term:
#         conflicting_terms.append('upper')
#     elif 'ground' in parent_term:
#         conflicting_terms.extend(['1st', '2nd', '3rd', '4th'])
#     elif '1st' in parent_term:
#         conflicting_terms.extend(['ground', '2nd', '3rd', '4th'])
#     elif '2nd' in parent_term:
#         conflicting_terms.extend(['ground', '1st', '3rd', '4th'])
#     elif '3rd' in parent_term:
#         conflicting_terms.extend(['ground', '1st', '2nd', '4th'])
    
#     logger.debug(f"    Parent term: '{parent_term}'")
#     logger.debug(f"    Conflicting terms: {conflicting_terms}")
    
#     # STEP 1: Find the parent row
#     parent_row = None
#     for row_idx in range(3, tower_sheet.max_row + 1):
#         task_name = tower_sheet.cell(row_idx, TRACKER_TASK_NAME_COL).value
        
#         if not task_name:
#             continue
        
#         task_normalized = normalize_text(task_name)
        
#         # Check if this row matches the parent
#         if parent_term in task_normalized:
#             # Make sure it doesn't contain conflicting terms
#             has_conflict = any(conflict in task_normalized for conflict in conflicting_terms)
            
#             if not has_conflict:
#                 parent_row = row_idx
#                 logger.debug(f"    Found parent at row {row_idx}: {task_name.strip()}")
#                 break
    
#     if not parent_row:
#         logger.debug(f"    Parent '{parent_term}' not found")
#         return None
    
#     # STEP 2: If we have child levels, search within next 10 rows
#     if len(activity_lines) > 1:
#         # Get search terms from remaining levels
#         child_terms = [normalize_text(line) for line in activity_lines[1:]]
        
#         logger.debug(f"    Child terms: {child_terms}")
        
#         best_match = None
#         best_match_score = 0
#         best_match_row = None
        
#         # Search within next 10 rows after parent
#         for row_idx in range(parent_row + 1, min(parent_row + 11, tower_sheet.max_row + 1)):
#             task_name = tower_sheet.cell(row_idx, TRACKER_TASK_NAME_COL).value
            
#             if not task_name:
#                 continue
            
#             task_normalized = normalize_text(task_name)
            
#             # Calculate match score: prioritize matching MORE terms (deeper hierarchy)
#             # Weight later terms more heavily (they're more specific, closer to leaf nodes)
#             match_count = 0
#             for idx, term in enumerate(child_terms):
#                 if term in task_normalized:
#                     # Weight: first child term = 1, second = 2, third = 3, etc.
#                     # This ensures "Casting Work" (term 2, weight 2) beats "Column/Shear Wall" (term 1, weight 1)
#                     match_count += (idx + 1)
            
#             # Only update if we have a BETTER match (higher score)
#             if match_count > best_match_score:
#                 # Get % Complete from Column G
#                 pct_value = tower_sheet.cell(row_idx, TRACKER_PCT_COMPLETE_COL).value
                
#                 if pct_value is not None:
#                     try:
#                         if isinstance(pct_value, (int, float)):
#                             pct_complete = float(pct_value) * 100
#                         else:
#                             pct_complete = float(str(pct_value).replace('%', ''))
                        
#                         if 0 <= pct_complete <= 100:
#                             best_match_score = match_count
#                             best_match = pct_complete
#                             best_match_row = row_idx
#                             logger.debug(f"    Child match at row {row_idx}: {task_name.strip()[:50]} = {pct_complete:.1f}% (score: {match_count})")
#                     except (ValueError, TypeError):
#                         pass
        
#         if best_match is not None:
#             logger.debug(f"    ✓ SELECTED: row {best_match_row}, {best_match:.1f}%")
#             return best_match
#         else:
#             logger.debug(f"    Child terms not found within 10 rows of parent")
#             return None
    
#     else:
#         # Single level activity - return parent row's % Complete
#         pct_value = tower_sheet.cell(parent_row, TRACKER_PCT_COMPLETE_COL).value
        
#         if pct_value is not None:
#             try:
#                 if isinstance(pct_value, (int, float)):
#                     pct_complete = float(pct_value) * 100
#                 else:
#                     pct_complete = float(str(pct_value).replace('%', ''))
                
#                 if 0 <= pct_complete <= 100:
#                     logger.debug(f"    ✓ SELECTED parent row: {pct_complete:.1f}%")
#                     return pct_complete
#             except (ValueError, TypeError):
#                 pass
        
#         return None
#     """
#     Find matching activity in tracker and return % complete.
#     Uses Column G (TRACKER_PCT_COMPLETE_COL = 7) for % Complete.
    
#     Improved hierarchical matching: First line terms are CRITICAL and must all match.
#     """
#     # Find tower sheet
#     tower_sheet = None
#     for sheet_name in tracker_wb.sheetnames:
#         if tower_name.replace("Tower ", "").replace("NTA ", "") in sheet_name:
#             tower_sheet = tracker_wb[sheet_name]
#             break
    
#     if not tower_sheet:
#         logger.debug(f"    Sheet not found for {tower_name}")
#         return None
    
#     # Split activity text into lines (hierarchy levels)
#     activity_lines = [line.strip() for line in activity_text.split('\n') if line.strip()]
    
#     # CRITICAL: First line must match EXACTLY (as a phrase)
#     # "Upper Basement" should NOT match "Lower Basement"
#     critical_line = ""
#     if activity_lines:
#         critical_line = normalize_text(activity_lines[0])
    
#     # Get all search terms from all lines for overall scoring
#     all_search_terms = []
#     for line in activity_lines:
#         normalized = normalize_text(line)
#         words = [w for w in normalized.split() if len(w) > 2]
#         all_search_terms.extend(words)
    
#     logger.debug(f"    Searching in {tower_sheet.title}")
#     logger.debug(f"    Activity: {' → '.join(activity_lines)}")
#     logger.debug(f"    CRITICAL phrase (must match exactly): '{critical_line}'")
#     logger.debug(f"    All search terms: {all_search_terms}")
    
#     # Identify conflicting location terms to explicitly exclude
#     conflicting_terms = []
#     search_text_full = ' '.join([normalize_text(line) for line in activity_lines])
    
#     # Define conflicting pairs
#     conflict_pairs = [
#         ('upper', 'lower'),
#         ('lower', 'upper'),
#         ('ground', '1st'),
#         ('ground', '2nd'),
#         ('ground', '3rd'),
#         ('1st', '2nd'),
#         ('1st', '3rd'),
#         ('2nd', '1st'),
#         ('2nd', '3rd'),
#         ('3rd', '1st'),
#         ('3rd', '2nd'),
#     ]
    
#     # Find which conflicts to check for
#     for our_term, their_term in conflict_pairs:
#         if our_term in search_text_full:
#             conflicting_terms.append(their_term)
    
#     logger.debug(f"    Will reject if contains: {conflicting_terms}")
    
#     # Search through tracker rows
#     best_match = None
#     best_match_score = 0
#     best_match_row = None
    
#     for row_idx in range(3, tower_sheet.max_row + 1):
#         task_name = tower_sheet.cell(row_idx, TRACKER_TASK_NAME_COL).value
        
#         if not task_name:
#             continue
        
#         task_normalized = normalize_text(task_name)
        
#         # STEP 0: Check for conflicting location terms
#         has_conflict = False
#         for conflict_term in conflicting_terms:
#             if conflict_term in task_normalized:
#                 logger.debug(f"    REJECTED row {row_idx}: contains conflicting term '{conflict_term}': {task_name[:50]}...")
#                 has_conflict = True
#                 break
        
#         if has_conflict:
#             continue  # Skip this row entirely
        
#         # STEP 1: Check if the COMPLETE critical line/phrase is present
#         # "upper basement" must be in the task, not just "basement"
#         if critical_line and critical_line not in task_normalized:
#             # Critical phrase not found, skip this row
#             continue
        
#         # STEP 2: Count all matching terms (for overall match quality)
#         total_match_count = sum(1 for term in all_search_terms if term in task_normalized)
        
#         # Calculate match percentage
#         match_percentage = total_match_count / len(all_search_terms) if all_search_terms else 0
        
#         # Need at least 70% of ALL terms to match
#         if match_percentage >= 0.7 and total_match_count > best_match_score:
#             # Get % Complete from Column G
#             pct_value = tower_sheet.cell(row_idx, TRACKER_PCT_COMPLETE_COL).value
            
#             if pct_value is not None:
#                 try:
#                     # Handle different formats
#                     if isinstance(pct_value, (int, float)):
#                         pct_complete = float(pct_value) * 100
#                     else:
#                         pct_complete = float(str(pct_value).replace('%', ''))
                    
#                     if 0 <= pct_complete <= 100:
#                         best_match_score = total_match_count
#                         best_match = pct_complete
#                         best_match_row = row_idx
#                         logger.debug(f"    Match: row {row_idx}, score={total_match_count}/{len(all_search_terms)} ({match_percentage*100:.0f}%): {task_name[:60]}... = {pct_complete:.1f}%")
#                 except (ValueError, TypeError):
#                     pass
    
#     if best_match is not None:
#         logger.debug(f"    ✓ SELECTED: row {best_match_row}, {best_match:.1f}%")
#     else:
#         logger.debug(f"    ✗ NO MATCH (need: exact critical phrase + 70%+ overall)")
    
#     return best_match

# # ======================= REPORT GENERATION =======================

# def sort_towers(tower_name: str) -> tuple:
#     """
#     Custom sort key for towers.
#     Returns tuple: (priority, numeric_value, original_name)
#     - Towers get priority 0
#     - NTAs get priority 1
#     This ensures Towers appear before NTAs, both sorted numerically
#     """
#     if tower_name.startswith('Tower'):
#         # Extract number from "Tower 4", "Tower 5", etc.
#         match = re.search(r'Tower\s*(\d+)', tower_name)
#         if match:
#             return (0, int(match.group(1)), tower_name)
#         return (0, 999, tower_name)
#     elif tower_name.startswith('NTA'):
#         # Extract number from "NTA 01", "NTA 02", etc.
#         match = re.search(r'NTA\s*(\d+)', tower_name)
#         if match:
#             return (1, int(match.group(1)), tower_name)
#         return (1, 999, tower_name)
#     else:
#         # Any other format
#         return (2, 999, tower_name)

# def generate_report(tower_targets: Dict[str, List[ActivityTarget]], 
#                    tracker_workbooks: Dict[str, Any], months: List[str], year: int) -> pd.DataFrame:
#     """
#     Generate milestone report DataFrame.
    
#     Args:
#         tower_targets: All targets from KRA by tower
#         tracker_workbooks: Dict of available tracker workbooks {month: workbook}
#         months: List of quarter months
#         year: KRA year
#     """
#     logger.info("\n" + "="*70)
#     logger.info("GENERATING REPORT")
#     logger.info("="*70)
    
#     report_rows = []
    
#     # Sort towers: Towers first (4, 5, 6, 7...), then NTAs (01, 02, 03...)
#     sorted_tower_names = sorted(tower_targets.keys(), key=sort_towers)
    
#     for tower_name in sorted_tower_names:
#         logger.info(f"\nProcessing: {tower_name}")
        
#         row_data = {'Tower': tower_name}
        
#         # Process each month
#         for month in months:
#             month_targets = [t for t in tower_targets[tower_name] if t.month == month]
            
#             # Check if tracker is available for this month
#             tracker_wb = tracker_workbooks.get(month)
            
#             if not month_targets:
#                 # No targets for this month - leave completely blank
#                 row_data[f"Activity- {month} {year}"] = ""
#                 row_data[f"% Complete- {month}"] = ""
#                 row_data[f"Status- {month}"] = ""
#                 row_data[f"Responsible- {month}"] = ""
#                 row_data[f"Delay- {month}"] = ""
#                 continue
            
#             # We have targets - always show activity text
#             activities_text = []
#             for target in month_targets:
#                 activities_text.append(target.activity_text)
            
#             row_data[f"Activity- {month} {year}"] = "\n".join(activities_text)
            
#             # If tracker is available, populate data
#             if tracker_wb:
#                 total_actual = 0
#                 matched = 0
                
#                 for target in month_targets:
#                     # Find in tracker
#                     actual_pct = find_activity_in_tracker(tracker_wb, tower_name, target.activity_text)
                    
#                     if actual_pct is not None:
#                         target.actual_pct = actual_pct
#                         total_actual += actual_pct
                        
#                         if abs(actual_pct - target.target_pct) < 1:
#                             target.status = "Matched"
#                             matched += 1
#                         elif actual_pct >= target.target_pct:
#                             target.status = "Achieved"
#                             matched += 1
#                         else:
#                             target.status = "Not Matched"
#                     else:
#                         target.status = "Not Found"
                    
#                     logger.info(f"  {month}: {target.activity_text[:40]} | Target={target.target_pct:.0f}%, Actual={target.actual_pct:.0f}%, Status={target.status}")
                
#                 avg_actual = total_actual / len(month_targets) if month_targets else 0
                
#                 if matched == len(month_targets) and matched > 0:
#                     status = "Achieved"
#                 elif matched > 0:
#                     status = "Partial"
#                 else:
#                     status = "Not Achieved"
                
#                 row_data[f"% Complete- {month}"] = f"{avg_actual:.0f}%"
#                 row_data[f"Status- {month}"] = status
#             else:
#                 # Tracker not available - leave data columns blank
#                 logger.info(f"  {month}: Tracker not available - data columns left blank")
#                 row_data[f"% Complete- {month}"] = ""
#                 row_data[f"Status- {month}"] = ""
            
#             # Responsible and Delay columns always blank (manual entry)
#             row_data[f"Responsible- {month}"] = ""
#             row_data[f"Delay- {month}"] = ""
        
#         # Add summary columns
#         last_month = months[-1]
#         last_targets = [t for t in tower_targets[tower_name] if t.month == last_month]
#         row_data[f"Target till {last_month}"] = "\n".join([t.activity_text for t in last_targets])
        
#         weightage = 100 if not tower_name.startswith('NTA') else 50
#         row_data['Weightage'] = weightage
#         row_data['Weighted %'] = "0%"  # Placeholder
        
#         report_rows.append(row_data)
    
#     return pd.DataFrame(report_rows)

# def format_report(worksheet, dataframe):
#     """Apply formatting to report."""
#     header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
#     header_font = Font(bold=True, color="FFFFFF")
    
#     # Format title
#     worksheet.cell(1, 1).font = Font(bold=True, size=14)
#     worksheet.cell(2, 1).font = Font(italic=True, size=10)
    
#     # Format headers
#     for col in range(1, worksheet.max_column + 1):
#         cell = worksheet.cell(4, col)
#         cell.fill = header_fill
#         cell.font = header_font
#         cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
#     # Format data
#     thin_border = Border(
#         left=Side(style='thin'), right=Side(style='thin'),
#         top=Side(style='thin'), bottom=Side(style='thin')
#     )
    
#     for row in range(5, worksheet.max_row + 1):
#         for col in range(1, worksheet.max_column + 1):
#             cell = worksheet.cell(row, col)
#             cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
#             cell.border = thin_border
    
#     # Column widths
#     for col_idx, column in enumerate(dataframe.columns, start=1):
#         col_letter = get_column_letter(col_idx)
#         if 'Activity' in column or 'Target' in column:
#             worksheet.column_dimensions[col_letter].width = 40
#         else:
#             worksheet.column_dimensions[col_letter].width = 15
    
#     worksheet.row_dimensions[4].height = 50

# # ======================= MAIN =======================

# def main():
#     """Main execution."""
#     try:
#         logger.info("\n" + "="*70)
#         logger.info("MILESTONE REPORT GENERATOR v2.0 - FIXED")
#         logger.info("="*70)
        
#         # Step 1: Find KRA
#         logger.info("\nSTEP 1: Finding latest KRA file")
#         cos = init_cos()
        
#         kra_result = find_latest_kra_file(cos, BUCKET, KRA_FOLDER)
#         if not kra_result:
#             logger.error("Could not find KRA file")
#             return
        
#         kra_key, quarter_months, kra_year = kra_result
        
#         # Step 2: Load KRA
#         logger.info("\nSTEP 2: Loading KRA and parsing targets")
#         kra_bytes = download_file_bytes(cos, kra_key)
#         kra_wb = load_workbook(filename=BytesIO(kra_bytes), data_only=True)
        
#         kra_ws = find_project_sheet(kra_wb, "EDEN")
#         if not kra_ws:
#             logger.error("EDEN sheet not found")
#             return
        
#         tower_targets = parse_kra_targets(kra_ws)
        
#         if not tower_targets:
#             logger.error("No targets found in KRA")
#             return
        
#         # Step 3: Load all available trackers
#         logger.info("\nSTEP 3: Loading tracker files")
#         tracker_workbooks = {}
        
#         for month in quarter_months:
#             tracker_month_num = MONTH_TO_TRACKER_MAPPING.get(month)
#             if not tracker_month_num:
#                 logger.warning(f"  No mapping for {month}")
#                 continue
            
#             tracker_year = calculate_tracker_year(month, kra_year)
#             logger.info(f"\n  {month} data requires tracker: {tracker_month_num:02d}/{tracker_year}")
            
#             tracker_key = find_tracker_for_month(cos, BUCKET, tracker_month_num, tracker_year, EDEN_TRACKER_FOLDER)
            
#             if tracker_key:
#                 logger.info(f"    ✓ Found: {os.path.basename(tracker_key)}")
#                 tracker_bytes = download_file_bytes(cos, tracker_key)
#                 tracker_wb = load_workbook(filename=BytesIO(tracker_bytes), data_only=True)
#                 tracker_workbooks[month] = tracker_wb
#                 logger.info(f"    ✓ Loaded with sheets: {tracker_wb.sheetnames}")
#             else:
#                 logger.warning(f"    ✗ Not found - {month} column will show activities but no completion data")
        
#         logger.info(f"\n  Summary: {len(tracker_workbooks)}/{len(quarter_months)} trackers loaded")
        
#         # Step 4: Generate report
#         logger.info("\nSTEP 4: Generating report")
#         report_df = generate_report(tower_targets, tracker_workbooks, quarter_months, kra_year)
        
#         # Step 5: Save
#         logger.info("\nSTEP 5: Saving report")
#         output_file = f"Eden_Milestone_Report_{'_'.join(quarter_months)}_{kra_year}.xlsx"
        
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Progress Report"
        
#         ws.append(["Eden- Progress Against Milestones"])
#         ws.append([f"Report Generated: {datetime.now().strftime('%B %d, %Y')}"])
#         ws.append([])
        
#         for r in dataframe_to_rows(report_df, index=False, header=True):
#             ws.append(r)
        
#         format_report(ws, report_df)
#         wb.save(output_file)
        
#         logger.info(f"\n{'='*70}")
#         logger.info("REPORT COMPLETE")
#         logger.info(f"{'='*70}")
#         logger.info(f"File: {output_file}")
#         logger.info(f"Towers: {len(report_df)}")
#         logger.info(f"Months with data: {list(tracker_workbooks.keys())}")
#         logger.info(f"Months with blank data: {[m for m in quarter_months if m not in tracker_workbooks]}")
#         logger.info(f"{'='*70}\n")
        
#     except Exception as e:
#         logger.error(f"Error: {str(e)}", exc_info=True)
#         raise

# if __name__ == "__main__":
#     main()




































"""
Automated Milestone Report Generator - FIXED VERSION
Hardcoded column numbers for stability, improved activity parsing
MODIFIED: NTA-05 October activity hardcoded to 80%
"""

import os
import logging
from io import BytesIO
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import ibm_boto3
from ibm_botocore.client import Config
import re
from typing import Optional, Tuple, List, Dict, Any

# ======================= CONFIGURATION =======================
load_dotenv()
logging.basicConfig(level=logging.DEBUG, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

# Cloud Storage Configuration
COS_API_KEY = os.getenv("COS_API_KEY")
COS_CRN = os.getenv("COS_SERVICE_INSTANCE_CRN")
COS_ENDPOINT = os.getenv("COS_ENDPOINT")
BUCKET = os.getenv("COS_BUCKET_NAME")
KRA_FOLDER = os.getenv("KRA_FOLDER", "")
EDEN_TRACKER_FOLDER = os.getenv("EDEN_TRACKER_FOLDER", "Eden/")

# ======================= HARDCODED COLUMN MAPPINGS =======================
# KRA Sheet Columns (1-indexed for openpyxl)
KRA_TOWER_COL = 1  # Column A: Tower name

KRA_SEP_ACTIVITY_COL = 2   # Column B: September activity
KRA_SEP_TARGET_COL = 3     # Column C: September % target

KRA_OCT_ACTIVITY_COL = 4   # Column D: October activity
KRA_OCT_TARGET_COL = 5     # Column E: October % target

KRA_NOV_ACTIVITY_COL = 6   # Column F: November activity
KRA_NOV_TARGET_COL = 7     # Column G: November % target

# Tracker Sheet Columns (1-indexed for openpyxl)
TRACKER_TOWER_COL = 1           # Column A: Tower number
TRACKER_ACTIVITY_NO_COL = 2     # Column B: Activity number
TRACKER_LOOKAHEAD_COL = 3       # Column C: Monthly lookahead ID
TRACKER_TASK_NAME_COL = 4       # Column D: Task name
TRACKER_ACTUAL_START_COL = 5    # Column E: Actual start
TRACKER_ACTUAL_FINISH_COL = 6   # Column F: Actual finish
TRACKER_PCT_COMPLETE_COL = 7    # Column G: % Complete ← THE KEY COLUMN
TRACKER_DURATION_COL = 8        # Column H: Duration

# Quarterly structure
QUARTERS = {
    "Q1": ["June", "July", "August"],
    "Q2": ["September", "October", "November"],
    "Q3": ["December", "January", "February"],
    "Q4": ["March", "April", "May"]
}

# Month to tracker month mapping
MONTH_TO_TRACKER_MAPPING = {
    "June": 7, "July": 8, "August": 9,
    "September": 10, "October": 11, "November": 12,
    "December": 1, "January": 2, "February": 3,
    "March": 4, "April": 5, "May": 6
}

# Validate environment variables
required_vars = {
    'COS_API_KEY': COS_API_KEY,
    'COS_SERVICE_INSTANCE_CRN': COS_CRN,
    'COS_ENDPOINT': COS_ENDPOINT,
    'COS_BUCKET_NAME': BUCKET
}

missing_vars = [var_name for var_name, var_value in required_vars.items() if not var_value]
if missing_vars:
    error_msg = f"Missing required environment variables: {', '.join(missing_vars)}"
    logger.error(error_msg)
    raise ValueError(error_msg)

# ======================= CLOUD STORAGE HELPERS =======================

def init_cos():
    """Initialize IBM Cloud Object Storage client."""
    return ibm_boto3.client(
        "s3",
        ibm_api_key_id=COS_API_KEY,
        ibm_service_instance_id=COS_CRN,
        config=Config(signature_version="oauth"),
        endpoint_url=COS_ENDPOINT
    )

def download_file_bytes(cos, key: str) -> bytes:
    """Download file from cloud storage as bytes."""
    return cos.get_object(Bucket=BUCKET, Key=key)["Body"].read()

# ======================= FILE DISCOVERY =======================

def find_latest_kra_file(cos_client, bucket_name: str, folder_prefix: str = "") -> Optional[Tuple[str, List[str], int]]:
    """Find the latest KRA Milestones file."""
    logger.info(f"\n{'='*70}")
    logger.info(f"SEARCHING FOR LATEST KRA FILE")
    logger.info(f"{'='*70}")
    
    try:
        response = cos_client.list_objects_v2(Bucket=bucket_name, Prefix=folder_prefix)
        
        if 'Contents' not in response:
            logger.error(f"No files found in folder '{folder_prefix}'")
            return None
        
        kra_files = []
        
        for obj in response['Contents']:
            file_key = obj['Key']
            filename = os.path.basename(file_key)
            filename_lower = filename.lower()
            
            if file_key.endswith('/'):
                continue
            
            is_kra = 'kra' in filename_lower and 'milestone' in filename_lower
            is_excel = filename_lower.endswith(('.xlsx', '.xls'))
            
            if is_kra and is_excel:
                months_pattern = r'(January|February|March|April|May|June|July|August|September|October|November|December)'
                found_months = re.findall(months_pattern, filename, re.IGNORECASE)
                found_months = [m.capitalize() for m in found_months]
                
                year_match = re.search(r'(\d{4})', filename)
                year = int(year_match.group(1)) if year_match else datetime.now().year
                
                kra_files.append({
                    'key': file_key,
                    'filename': filename,
                    'months': found_months,
                    'year': year,
                    'last_modified': obj['LastModified']
                })
                
                logger.info(f"Found: {filename}")
        
        if not kra_files:
            logger.error("No KRA Milestone files found")
            return None
        
        kra_files.sort(key=lambda f: f['last_modified'], reverse=True)
        latest = kra_files[0]
        
        logger.info(f"\nSelected: {latest['filename']}")
        logger.info(f"Months: {', '.join(latest['months'])}")
        logger.info(f"Year: {latest['year']}")
        
        return latest['key'], latest['months'], latest['year']
        
    except Exception as e:
        logger.error(f"Error searching for KRA file: {str(e)}")
        raise

def calculate_tracker_year(report_month: str, kra_year: int) -> int:
    """Calculate correct year for tracker file."""
    tracker_month_num = MONTH_TO_TRACKER_MAPPING.get(report_month)
    
    if not tracker_month_num:
        return kra_year
    
    if report_month == "December" and tracker_month_num == 1:
        return kra_year + 1
    
    if report_month in ["January", "February"] and tracker_month_num in [2, 3]:
        return kra_year
    
    return kra_year

def find_tracker_for_month(cos_client, bucket_name: str, target_month: int, target_year: int,
                          folder_prefix: str = "Eden/") -> Optional[str]:
    """Find tracker file for specific month and year."""
    logger.info(f"  Searching for tracker: Month {target_month}/{target_year}")
    
    try:
        response = cos_client.list_objects_v2(Bucket=bucket_name, Prefix=folder_prefix)
        
        if 'Contents' not in response:
            return None
        
        matching_trackers = []
        
        for obj in response['Contents']:
            file_key = obj['Key']
            filename = os.path.basename(file_key)
            filename_lower = filename.lower()
            
            if file_key.endswith('/'):
                continue
            
            is_tracker = any(pattern in filename_lower for pattern in 
                           ['structure work tracker', 'tracker', 'structure tracker'])
            is_excel = filename_lower.endswith(('.xlsx', '.xls'))
            
            if is_tracker and is_excel:
                date_pattern = r'\((\d{1,2})-(\d{1,2})-(\d{2,4})\)'
                date_match = re.search(date_pattern, filename)
                
                if date_match:
                    day, month, year = date_match.groups()
                    file_month = int(month)
                    file_year = int(year)
                    
                    if file_year < 100:
                        file_year += 2000
                    
                    if file_month == target_month and file_year == target_year:
                        matching_trackers.append({
                            'key': file_key,
                            'filename': filename,
                            'day': int(day)
                        })
        
        if not matching_trackers:
            return None
        
        matching_trackers.sort(key=lambda t: t['day'], reverse=True)
        return matching_trackers[0]['key']
        
    except Exception as e:
        logger.error(f"Error searching for tracker: {str(e)}")
        return None

# ======================= KRA DATA EXTRACTION =======================

def find_project_sheet(workbook, project_name: str):
    """Find sheet containing project name."""
    for sheet_name in workbook.sheetnames:
        if project_name.upper() in sheet_name.upper():
            return workbook[sheet_name]
    return None

class ActivityTarget:
    """Represents a target activity from KRA."""
    
    def __init__(self, tower: str, activity_text: str, target_pct: float, month: str):
        self.tower = tower
        self.activity_text = activity_text  # Full text as it appears
        self.target_pct = target_pct
        self.month = month
        self.actual_pct = 0.0
        self.status = ""
    
    def __repr__(self):
        return f"{self.tower} | {self.month} | {self.activity_text} ({self.target_pct}%)"

def parse_kra_targets(worksheet) -> Dict[str, List[ActivityTarget]]:
    """
    Parse ALL targets from KRA sheet for all towers.
    Returns: Dict[tower_name, List[ActivityTarget]]
    """
    logger.info("\n" + "="*70)
    logger.info("PARSING KRA TARGETS")
    logger.info("="*70)
    
    tower_targets = {}
    current_tower = None
    
    # Start from row 5 (after header row 4)
    for row_idx in range(5, worksheet.max_row + 1):
        tower_cell = worksheet.cell(row_idx, KRA_TOWER_COL).value
        
        # Check if this is a new tower row
        if tower_cell and str(tower_cell).strip():
            tower_name = str(tower_cell).strip()
            
            # Only process actual towers (not finishing milestones)
            if 'Tower' in tower_name or 'NTA' in tower_name:
                if 'Finishing' not in tower_name:
                    current_tower = tower_name
                    if current_tower not in tower_targets:
                        tower_targets[current_tower] = []
                    
                    logger.info(f"\nProcessing: {current_tower}")
                    
                    # Parse September target
                    sep_activity = worksheet.cell(row_idx, KRA_SEP_ACTIVITY_COL).value
                    sep_target = worksheet.cell(row_idx, KRA_SEP_TARGET_COL).value
                    
                    if sep_target and isinstance(sep_target, (int, float)) and sep_target > 0:
                        activity_text = str(sep_activity).strip() if sep_activity else "Activity"
                        target = ActivityTarget(current_tower, activity_text, float(sep_target) * 100, "September")
                        tower_targets[current_tower].append(target)
                        logger.info(f"  September: {activity_text} → {sep_target*100}%")
                    
                    # Parse October target
                    oct_activity = worksheet.cell(row_idx, KRA_OCT_ACTIVITY_COL).value
                    oct_target = worksheet.cell(row_idx, KRA_OCT_TARGET_COL).value
                    
                    if oct_target and isinstance(oct_target, (int, float)) and oct_target > 0:
                        activity_text = str(oct_activity).strip() if oct_activity else "Activity"
                        target = ActivityTarget(current_tower, activity_text, float(oct_target) * 100, "October")
                        tower_targets[current_tower].append(target)
                        logger.info(f"  October: {activity_text} → {oct_target*100}%")
                    
                    # Parse November target
                    nov_activity = worksheet.cell(row_idx, KRA_NOV_ACTIVITY_COL).value
                    nov_target = worksheet.cell(row_idx, KRA_NOV_TARGET_COL).value
                    
                    if nov_target and isinstance(nov_target, (int, float)) and nov_target > 0:
                        activity_text = str(nov_activity).strip() if nov_activity else "Activity"
                        target = ActivityTarget(current_tower, activity_text, float(nov_target) * 100, "November")
                        tower_targets[current_tower].append(target)
                        logger.info(f"  November: {activity_text} → {nov_target*100}%")
                else:
                    current_tower = None  # Stop processing finishing milestones
        
        # Check sub-rows for multi-line activities (Tower 5, Tower 7 pattern)
        elif current_tower:
            # Check September sub-activity
            sep_activity = worksheet.cell(row_idx, KRA_SEP_ACTIVITY_COL).value
            sep_target = worksheet.cell(row_idx, KRA_SEP_TARGET_COL).value
            
            if sep_target and isinstance(sep_target, (int, float)) and sep_target > 0:
                # Build hierarchical activity text
                activity_parts = []
                for back_row in range(max(row_idx - 3, 5), row_idx + 1):
                    cell_val = worksheet.cell(back_row, KRA_SEP_ACTIVITY_COL).value
                    if cell_val and str(cell_val).strip() and str(cell_val).strip() != current_tower:
                        activity_parts.append(str(cell_val).strip())
                
                activity_text = "\n".join(activity_parts) if activity_parts else str(sep_activity).strip()
                target = ActivityTarget(current_tower, activity_text, float(sep_target) * 100, "September")
                tower_targets[current_tower].append(target)
                logger.info(f"  September (sub): {activity_text[:50]}... → {sep_target*100}%")
            
            # Check October sub-activity
            oct_activity = worksheet.cell(row_idx, KRA_OCT_ACTIVITY_COL).value
            oct_target = worksheet.cell(row_idx, KRA_OCT_TARGET_COL).value
            
            if oct_target and isinstance(oct_target, (int, float)) and oct_target > 0:
                activity_parts = []
                for back_row in range(max(row_idx - 3, 5), row_idx + 1):
                    cell_val = worksheet.cell(back_row, KRA_OCT_ACTIVITY_COL).value
                    if cell_val and str(cell_val).strip() and str(cell_val).strip() != current_tower:
                        activity_parts.append(str(cell_val).strip())
                
                activity_text = "\n".join(activity_parts) if activity_parts else str(oct_activity).strip()
                target = ActivityTarget(current_tower, activity_text, float(oct_target) * 100, "October")
                tower_targets[current_tower].append(target)
                logger.info(f"  October (sub): {activity_text[:50]}... → {oct_target*100}%")
            
            # Check November sub-activity
            nov_activity = worksheet.cell(row_idx, KRA_NOV_ACTIVITY_COL).value
            nov_target = worksheet.cell(row_idx, KRA_NOV_TARGET_COL).value
            
            if nov_target and isinstance(nov_target, (int, float)) and nov_target > 0:
                activity_parts = []
                for back_row in range(max(row_idx - 3, 5), row_idx + 1):
                    cell_val = worksheet.cell(back_row, KRA_NOV_ACTIVITY_COL).value
                    if cell_val and str(cell_val).strip() and str(cell_val).strip() != current_tower:
                        activity_parts.append(str(cell_val).strip())
                
                activity_text = "\n".join(activity_parts) if activity_parts else str(nov_activity).strip()
                target = ActivityTarget(current_tower, activity_text, float(nov_target) * 100, "November")
                tower_targets[current_tower].append(target)
                logger.info(f"  November (sub): {activity_text[:50]}... → {nov_target*100}%")
    
    logger.info(f"\nTotal towers with targets: {len(tower_targets)}")
    for tower, targets in tower_targets.items():
        logger.info(f"  {tower}: {len(targets)} targets")
    
    return tower_targets

# ======================= TRACKER DATA EXTRACTION =======================

def normalize_text(text: str) -> str:
    """Normalize text for matching."""
    if not text:
        return ""
    # Convert to lowercase, remove extra spaces, remove special chars
    text = re.sub(r'\s+', ' ', str(text).lower().strip())
    text = re.sub(r'[^\w\s]', ' ', text)
    return ' '.join(text.split())

def find_activity_in_tracker(tracker_wb, tower_name: str, activity_text: str, month: str = None) -> Optional[float]:
    """
    Find matching activity in tracker and return % complete.
    Uses Column G (TRACKER_PCT_COMPLETE_COL = 7) for % Complete.
    
    SPECIAL OVERRIDE: NTA-05 October "Lower Basement, Column/Shear Wall, Checking & Casting Work" = 80%
    
    HIERARCHICAL MATCHING: Find parent level first, then search for child within next 10 rows.
    """
    # ============= HARDCODED OVERRIDE FOR NTA-05 OCTOBER =============
    tower_normalized = tower_name.replace(" ", "").upper()
    is_nta05 = tower_normalized in ["NTA05", "NTA5", "NTA-05", "NTA-5"]
    is_october = month == "October"
    
    if is_nta05 and is_october:
        activity_normalized = normalize_text(activity_text)
        # Check if this is the specific activity
        has_lower_basement = "lower" in activity_normalized and "basement" in activity_normalized
        has_column_shear = ("column" in activity_normalized or "shear" in activity_normalized)
        has_checking_casting = ("checking" in activity_normalized and "casting" in activity_normalized) or "checking casting" in activity_normalized
        
        if has_lower_basement and has_column_shear and has_checking_casting:
            logger.info(f"    ✓✓✓ HARDCODED OVERRIDE: NTA-05 October activity = 80% ✓✓✓")
            return 80.0
    # ================================================================
    
    # Find tower sheet
    tower_sheet = None
    for sheet_name in tracker_wb.sheetnames:
        if tower_name.replace("Tower ", "").replace("NTA ", "") in sheet_name:
            tower_sheet = tracker_wb[sheet_name]
            break
    
    if not tower_sheet:
        logger.debug(f"    Sheet not found for {tower_name}")
        return None
    
    # Split activity text into hierarchy levels
    activity_lines = [line.strip() for line in activity_text.split('\n') if line.strip()]
    
    if not activity_lines:
        return None
    
    logger.debug(f"    Searching in {tower_sheet.title}")
    logger.debug(f"    Hierarchy: {' → '.join(activity_lines)}")
    
    # Strategy: Find the PARENT level first (e.g., "Upper Basement")
    # Then find the CHILD within the next 10 rows (e.g., "Casting Work")
    
    parent_term = normalize_text(activity_lines[0])  # e.g., "upper basement"
    
    # Define conflicting terms for the parent
    conflicting_terms = []
    if 'upper' in parent_term:
        conflicting_terms.append('lower')
    elif 'lower' in parent_term:
        conflicting_terms.append('upper')
    elif 'ground' in parent_term:
        conflicting_terms.extend(['1st', '2nd', '3rd', '4th'])
    elif '1st' in parent_term:
        conflicting_terms.extend(['ground', '2nd', '3rd', '4th'])
    elif '2nd' in parent_term:
        conflicting_terms.extend(['ground', '1st', '3rd', '4th'])
    elif '3rd' in parent_term:
        conflicting_terms.extend(['ground', '1st', '2nd', '4th'])
    
    logger.debug(f"    Parent term: '{parent_term}'")
    logger.debug(f"    Conflicting terms: {conflicting_terms}")
    
    # STEP 1: Find the parent row
    parent_row = None
    for row_idx in range(3, tower_sheet.max_row + 1):
        task_name = tower_sheet.cell(row_idx, TRACKER_TASK_NAME_COL).value
        
        if not task_name:
            continue
        
        task_normalized = normalize_text(task_name)
        
        # Check if this row matches the parent
        if parent_term in task_normalized:
            # Make sure it doesn't contain conflicting terms
            has_conflict = any(conflict in task_normalized for conflict in conflicting_terms)
            
            if not has_conflict:
                parent_row = row_idx
                logger.debug(f"    Found parent at row {row_idx}: {task_name.strip()}")
                break
    
    if not parent_row:
        logger.debug(f"    Parent '{parent_term}' not found")
        return None
    
    # STEP 2: If we have child levels, search within next 10 rows
    if len(activity_lines) > 1:
        # Get search terms from remaining levels
        child_terms = [normalize_text(line) for line in activity_lines[1:]]
        
        logger.debug(f"    Child terms: {child_terms}")
        
        best_match = None
        best_match_score = 0
        best_match_row = None
        
        # Search within next 10 rows after parent
        for row_idx in range(parent_row + 1, min(parent_row + 11, tower_sheet.max_row + 1)):
            task_name = tower_sheet.cell(row_idx, TRACKER_TASK_NAME_COL).value
            
            if not task_name:
                continue
            
            task_normalized = normalize_text(task_name)
            
            # Calculate match score: prioritize matching MORE terms (deeper hierarchy)
            # Weight later terms more heavily (they're more specific, closer to leaf nodes)
            match_count = 0
            for idx, term in enumerate(child_terms):
                if term in task_normalized:
                    # Weight: first child term = 1, second = 2, third = 3, etc.
                    # This ensures "Casting Work" (term 2, weight 2) beats "Column/Shear Wall" (term 1, weight 1)
                    match_count += (idx + 1)
            
            # Only update if we have a BETTER match (higher score)
            if match_count > best_match_score:
                # Get % Complete from Column G
                pct_value = tower_sheet.cell(row_idx, TRACKER_PCT_COMPLETE_COL).value
                
                if pct_value is not None:
                    try:
                        if isinstance(pct_value, (int, float)):
                            pct_complete = float(pct_value) * 100
                        else:
                            pct_complete = float(str(pct_value).replace('%', ''))
                        
                        if 0 <= pct_complete <= 100:
                            best_match_score = match_count
                            best_match = pct_complete
                            best_match_row = row_idx
                            logger.debug(f"    Child match at row {row_idx}: {task_name.strip()[:50]} = {pct_complete:.1f}% (score: {match_count})")
                    except (ValueError, TypeError):
                        pass
        
        if best_match is not None:
            logger.debug(f"    ✓ SELECTED: row {best_match_row}, {best_match:.1f}%")
            return best_match
        else:
            logger.debug(f"    Child terms not found within 10 rows of parent")
            return None
    
    else:
        # Single level activity - return parent row's % Complete
        pct_value = tower_sheet.cell(parent_row, TRACKER_PCT_COMPLETE_COL).value
        
        if pct_value is not None:
            try:
                if isinstance(pct_value, (int, float)):
                    pct_complete = float(pct_value) * 100
                else:
                    pct_complete = float(str(pct_value).replace('%', ''))
                
                if 0 <= pct_complete <= 100:
                    logger.debug(f"    ✓ SELECTED parent row: {pct_complete:.1f}%")
                    return pct_complete
            except (ValueError, TypeError):
                pass
        
        return None

# ======================= REPORT GENERATION =======================

def sort_towers(tower_name: str) -> tuple:
    """
    Custom sort key for towers.
    Returns tuple: (priority, numeric_value, original_name)
    - Towers get priority 0
    - NTAs get priority 1
    This ensures Towers appear before NTAs, both sorted numerically
    """
    if tower_name.startswith('Tower'):
        # Extract number from "Tower 4", "Tower 5", etc.
        match = re.search(r'Tower\s*(\d+)', tower_name)
        if match:
            return (0, int(match.group(1)), tower_name)
        return (0, 999, tower_name)
    elif tower_name.startswith('NTA'):
        # Extract number from "NTA 01", "NTA 02", etc.
        match = re.search(r'NTA\s*(\d+)', tower_name)
        if match:
            return (1, int(match.group(1)), tower_name)
        return (1, 999, tower_name)
    else:
        # Any other format
        return (2, 999, tower_name)

def generate_report(tower_targets: Dict[str, List[ActivityTarget]], 
                   tracker_workbooks: Dict[str, Any], months: List[str], year: int) -> pd.DataFrame:
    """
    Generate milestone report DataFrame.
    
    Args:
        tower_targets: All targets from KRA by tower
        tracker_workbooks: Dict of available tracker workbooks {month: workbook}
        months: List of quarter months
        year: KRA year
    """
    logger.info("\n" + "="*70)
    logger.info("GENERATING REPORT")
    logger.info("="*70)
    
    report_rows = []
    
    # Sort towers: Towers first (4, 5, 6, 7...), then NTAs (01, 02, 03...)
    sorted_tower_names = sorted(tower_targets.keys(), key=sort_towers)
    
    for tower_name in sorted_tower_names:
        # Skip the standalone "NTA" row if it exists
        if tower_name.strip().upper() == "NTA":
            logger.info(f"\nSkipping: {tower_name} (not a valid tower)")
            continue
            
        logger.info(f"\nProcessing: {tower_name}")
        
        row_data = {'Tower': tower_name}
        
        # Process each month
        for month in months:
            month_targets = [t for t in tower_targets[tower_name] if t.month == month]
            
            # Check if tracker is available for this month
            tracker_wb = tracker_workbooks.get(month)
            
            if not month_targets:
                # No targets for this month - leave completely blank
                row_data[f"Activity- {month} {year}"] = ""
                row_data[f"% Complete- {month}"] = ""
                row_data[f"Status- {month}"] = ""
                row_data[f"Weightage- {month}"] = ""
                row_data[f"Weighted %- {month}"] = ""
                continue
            
            # We have targets - always show activity text
            activities_text = []
            for target in month_targets:
                activities_text.append(target.activity_text)
            
            row_data[f"Activity- {month} {year}"] = "\n".join(activities_text)
            
            # If tracker is available, populate data
            if tracker_wb:
                total_actual = 0
                matched = 0
                
                for target in month_targets:
                    # Find in tracker - PASS THE MONTH to enable hardcoded override
                    actual_pct = find_activity_in_tracker(tracker_wb, tower_name, target.activity_text, month)
                    
                    if actual_pct is not None:
                        # NEW LOGIC: If tracker % > target %, activity is fully completed (100%)
                        if actual_pct > target.target_pct:
                            target.actual_pct = 100.0
                            target.status = "Achieved"
                            matched += 1
                            logger.info(f"  {month}: {target.activity_text[:40]} | Target={target.target_pct:.0f}%, Tracker={actual_pct:.0f}% → COMPLETED 100%")
                        else:
                            target.actual_pct = actual_pct
                            
                            if abs(actual_pct - target.target_pct) < 1:
                                target.status = "Matched"
                                matched += 1
                            elif actual_pct >= target.target_pct:
                                target.status = "Achieved"
                                matched += 1
                            else:
                                target.status = "Not Matched"
                            
                            logger.info(f"  {month}: {target.activity_text[:40]} | Target={target.target_pct:.0f}%, Actual={target.actual_pct:.0f}%, Status={target.status}")
                        
                        total_actual += target.actual_pct
                    else:
                        target.status = "Not Found"
                        logger.info(f"  {month}: {target.activity_text[:40]} | Target={target.target_pct:.0f}%, Actual=Not Found")
                
                avg_actual = total_actual / len(month_targets) if month_targets else 0
                
                if matched == len(month_targets) and matched > 0:
                    status = "Achieved"
                elif matched > 0:
                    status = "Partial"
                else:
                    status = "Not Achieved"
                
                row_data[f"% Complete- {month}"] = f"{avg_actual:.0f}%"
                row_data[f"Status- {month}"] = status
                
                # Weightage and Weighted % for this month
                weightage = 100  # Always 100 for all towers
                weighted_pct = (avg_actual / weightage) * 100 if weightage > 0 else 0
                row_data[f"Weightage- {month}"] = weightage
                row_data[f"Weighted %- {month}"] = f"{weighted_pct:.0f}%"
            else:
                # Tracker not available - leave data columns blank
                logger.info(f"  {month}: Tracker not available - data columns left blank")
                row_data[f"% Complete- {month}"] = ""
                row_data[f"Status- {month}"] = ""
                row_data[f"Weightage- {month}"] = ""
                row_data[f"Weighted %- {month}"] = ""
        
        # Add summary columns at the end
        last_month = months[-1]
        last_targets = [t for t in tower_targets[tower_name] if t.month == last_month]
        row_data[f"Target till {last_month}"] = "\n".join([t.activity_text for t in last_targets])
        
        # Add single Responsible and Delay columns at the end (manual entry)
        row_data['Responsible'] = ""
        row_data['Delay Reason'] = ""
        
        report_rows.append(row_data)
    
    # Add summary row with averages at the end
    summary_row = {'Tower': 'AVERAGE WEIGHTED %'}
    
    for month in months:
        # Calculate average of Weighted % for this month across all towers
        weighted_values = []
        
        for row in report_rows:
            weighted_val = row.get(f"Weighted %- {month}", "")
            if weighted_val and weighted_val != "":
                try:
                    # Remove % sign and convert to float
                    val = float(str(weighted_val).replace('%', ''))
                    weighted_values.append(val)
                except (ValueError, TypeError):
                    pass
        
        # Calculate average
        if weighted_values:
            avg_weighted = sum(weighted_values) / len(weighted_values)
            summary_row[f"Weighted %- {month}"] = f"{avg_weighted:.1f}%"
        else:
            summary_row[f"Weighted %- {month}"] = ""
        
        # Leave other columns blank for summary row
        summary_row[f"Activity- {month} {year}"] = ""
        summary_row[f"% Complete- {month}"] = ""
        summary_row[f"Status- {month}"] = ""
        summary_row[f"Weightage- {month}"] = ""
    
    # Add empty values for end columns
    summary_row[f"Target till {months[-1]}"] = ""
    summary_row['Responsible'] = ""
    summary_row['Delay Reason'] = ""
    
    report_rows.append(summary_row)
    
    return pd.DataFrame(report_rows)

def format_report(worksheet, dataframe):
    """Apply formatting to report."""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    summary_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    summary_font = Font(bold=True, size=11)
    
    # Format title
    worksheet.cell(1, 1).font = Font(bold=True, size=14)
    worksheet.cell(2, 1).font = Font(italic=True, size=10)
    
    # Format headers
    for col in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(4, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Format data
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Last row is the summary row
    summary_row_idx = worksheet.max_row
    
    for row in range(5, worksheet.max_row + 1):
        is_summary_row = (row == summary_row_idx)
        
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row, col)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.border = thin_border
            
            # Special formatting for summary row
            if is_summary_row:
                cell.fill = summary_fill
                cell.font = summary_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Column widths
    for col_idx, column in enumerate(dataframe.columns, start=1):
        col_letter = get_column_letter(col_idx)
        if 'Activity' in column or 'Target' in column:
            worksheet.column_dimensions[col_letter].width = 40
        else:
            worksheet.column_dimensions[col_letter].width = 15
    
    worksheet.row_dimensions[4].height = 50
    worksheet.row_dimensions[summary_row_idx].height = 30  # Make summary row slightly taller

# ======================= MAIN =======================

def main():
    """Main execution."""
    try:
        logger.info("\n" + "="*70)
        logger.info("MILESTONE REPORT GENERATOR v2.0 - FIXED")
        logger.info("SPECIAL: NTA-05 October hardcoded to 80%")
        logger.info("="*70)
        
        # Step 1: Find KRA
        logger.info("\nSTEP 1: Finding latest KRA file")
        cos = init_cos()
        
        kra_result = find_latest_kra_file(cos, BUCKET, KRA_FOLDER)
        if not kra_result:
            logger.error("Could not find KRA file")
            return
        
        kra_key, quarter_months, kra_year = kra_result
        
        # Step 2: Load KRA
        logger.info("\nSTEP 2: Loading KRA and parsing targets")
        kra_bytes = download_file_bytes(cos, kra_key)
        kra_wb = load_workbook(filename=BytesIO(kra_bytes), data_only=True)
        
        kra_ws = find_project_sheet(kra_wb, "EDEN")
        if not kra_ws:
            logger.error("EDEN sheet not found")
            return
        
        tower_targets = parse_kra_targets(kra_ws)
        
        if not tower_targets:
            logger.error("No targets found in KRA")
            return
        
        # Step 3: Load all available trackers
        logger.info("\nSTEP 3: Loading tracker files")
        tracker_workbooks = {}
        
        for month in quarter_months:
            tracker_month_num = MONTH_TO_TRACKER_MAPPING.get(month)
            if not tracker_month_num:
                logger.warning(f"  No mapping for {month}")
                continue
            
            tracker_year = calculate_tracker_year(month, kra_year)
            logger.info(f"\n  {month} data requires tracker: {tracker_month_num:02d}/{tracker_year}")
            
            tracker_key = find_tracker_for_month(cos, BUCKET, tracker_month_num, tracker_year, EDEN_TRACKER_FOLDER)
            
            if tracker_key:
                logger.info(f"    ✓ Found: {os.path.basename(tracker_key)}")
                tracker_bytes = download_file_bytes(cos, tracker_key)
                tracker_wb = load_workbook(filename=BytesIO(tracker_bytes), data_only=True)
                tracker_workbooks[month] = tracker_wb
                logger.info(f"    ✓ Loaded with sheets: {tracker_wb.sheetnames}")
            else:
                logger.warning(f"    ✗ Not found - {month} column will show activities but no completion data")
        
        logger.info(f"\n  Summary: {len(tracker_workbooks)}/{len(quarter_months)} trackers loaded")
        
        # Step 4: Generate report
        logger.info("\nSTEP 4: Generating report")
        report_df = generate_report(tower_targets, tracker_workbooks, quarter_months, kra_year)
        
        # Step 5: Save
        logger.info("\nSTEP 5: Saving report")
        output_file = f"Eden_Milestone_Report_{'_'.join(quarter_months)}_{kra_year}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Progress Report"
        
        ws.append(["Eden- Progress Against Milestones"])
        ws.append([f"Report Generated: {datetime.now().strftime('%B %d, %Y')}"])
        ws.append([])
        
        for r in dataframe_to_rows(report_df, index=False, header=True):
            ws.append(r)
        
        format_report(ws, report_df)
        wb.save(output_file)
        
        logger.info(f"\n{'='*70}")
        logger.info("REPORT COMPLETE")
        logger.info(f"{'='*70}")
        logger.info(f"File: {output_file}")
        logger.info(f"Towers: {len(report_df)}")
        logger.info(f"Months with data: {list(tracker_workbooks.keys())}")
        logger.info(f"Months with blank data: {[m for m in quarter_months if m not in tracker_workbooks]}")
        logger.info(f"{'='*70}\n")
        
    except Exception as e:
        logger.error(f"Error: {str(e)}", exc_info=True)
        raise

if __name__ == "__main__":
    main()
