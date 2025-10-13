import os
import re
import logging
from io import BytesIO
from datetime import datetime
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from dotenv import load_dotenv
import ibm_boto3
from ibm_botocore.client import Config

# =============== CONFIG / CONSTANTS ===============
load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

COS_API_KEY = os.getenv("COS_API_KEY")
COS_CRN = os.getenv("COS_SERVICE_INSTANCE_CRN")
COS_ENDPOINT = os.getenv("COS_ENDPOINT")
BUCKET = os.getenv("COS_BUCKET_NAME")
EWS_LIG_KRA_KEY = os.getenv("KRA_FILE_PATH")

# Dynamic tracker path
EWS_LIG_STRUCTURE_KEY = None

# Dynamic months
MONTHS = []
MONTH_TO_NUM = {}
TRACKER_DATE = None
PROCESSING_MONTHS = []

KRA_SHEET = "EW-LI P4 Targets Till August "
TRACKER_SHEET = "Revised Baseline 45daysNGT+Rai"

# Tower positions
TOWER1_POUR_COLS = ['D', 'H', 'L', 'P']
TOWER1_ROW_START, TOWER1_ROW_END = 5, 22

TOWER3_POUR_COLS = ['D', 'H', 'L', 'P']
TOWER3_ROW_START, TOWER3_ROW_END = 27, 44

TOWER2_POUR_COLS = ['U', 'Y', 'AC', 'AG']
TOWER2_ROW_START, TOWER2_ROW_END = 5, 22

# Styles
YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
GREY = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")


# =============== UTILITY FUNCTIONS ===============

def list_files_in_folder(cos, folder_prefix):
    try:
        response = cos.list_objects_v2(Bucket=BUCKET, Prefix=folder_prefix)
        files = []
        if 'Contents' in response:
            for obj in response['Contents']:
                if not obj['Key'].endswith('/'):
                    files.append(obj['Key'])
        return files
    except Exception as e:
        logger.error(f"Error listing files in folder {folder_prefix}: {e}")
        return []


def extract_date_from_filename(filename):
    """Extract date from filename in multiple formats like (01-07-2025), (2025-07-01), (01.07.2025)"""
    pattern = r'\(([^)]+)\)'  # anything inside ()
    match = re.search(pattern, filename)
    if match:
        date_str = match.group(1)
        # Try pandas flexible parser
        dt = pd.to_datetime(date_str, dayfirst=True, errors="coerce")
        if pd.notna(dt):
            return dt.to_pydatetime()

        # Fallback strict formats
        for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y", "%d-%b-%y", "%d-%b-%Y"):
            try:
                return datetime.strptime(date_str.strip(), fmt)
            except ValueError:
                continue

        logger.warning(f"❌ Could not parse date {date_str} from filename {filename}")
    return None


def get_month_name(month_num):
    months = {
        1: "January", 2: "February", 3: "March", 4: "April",
        5: "May", 6: "June", 7: "July", 8: "August",
        9: "September", 10: "October", 11: "November", 12: "December"
    }
    return months.get(month_num, "Unknown")


def setup_dynamic_months_and_targets(tracker_date):
    """
    Dynamically setup MONTHS, MONTH_TO_NUM, PROCESSING_MONTHS, and tower target cells.
    - Always considers previous 3 months relative to tracker_date.
    - Handles year transitions.
    - Month labels include year (e.g., 'June 2025').
    """
    global MONTHS, MONTH_TO_NUM, TRACKER_DATE, PROCESSING_MONTHS
    global TOWER1_TARGETS_CELLS, TOWER2_TARGETS_CELLS, TOWER3_TARGETS_CELLS

    TRACKER_DATE = tracker_date
    tracker_month = tracker_date.month
    tracker_year = tracker_date.year

    logger.info(f"📊 Tracker date: {tracker_date:%d-%m-%Y}")

    # Determine previous 3 months
    MONTHS_DATA = []
    for offset in range(3, 0, -1):  # last 3 months
        month_num = tracker_month - offset
        year = tracker_year
        if month_num < 1:
            month_num += 12
            year -= 1
        MONTHS_DATA.append((month_num, year))

    # Month labels with year
    MONTHS = [f"{get_month_name(m)}" for m, y in MONTHS_DATA]

    # Mapping month name to number (without year)
    MONTH_TO_NUM = {get_month_name(m): m for m in range(1, 13)}

    # Processing months (all previous 3 months)
    PROCESSING_MONTHS = MONTHS[:]

    logger.info(f"Processing quarterly months: {MONTHS}")

    # Base columns for 3 months per quarter
    base_cols = ['B', 'C', 'D']

    # Dynamic target cells for towers
    TOWER1_TARGETS_CELLS = {month: f"{base_cols[i]}4" for i, month in enumerate(MONTHS)}
    TOWER3_TARGETS_CELLS = {month: f"{base_cols[i]}12" for i, month in enumerate(MONTHS)}
    TOWER2_TARGETS_CELLS = {month: f"{base_cols[i]}19" for i, month in enumerate(MONTHS)}

    logger.info(f"TOWER1_TARGETS_CELLS: {TOWER1_TARGETS_CELLS}")
    logger.info(f"TOWER2_TARGETS_CELLS: {TOWER2_TARGETS_CELLS}")
    logger.info(f"TOWER3_TARGETS_CELLS: {TOWER3_TARGETS_CELLS}")


def get_latest_tracker_paths(cos):
    global EWS_LIG_STRUCTURE_KEY

    logger.info("=== FINDING LATEST EWS LIG P4 TRACKER FILE ===")

    ews_files = list_files_in_folder(cos, "EWS LIG P4/")
    logger.info(f"Found {len(ews_files)} files in EWS LIG P4 folder")

    tracker_pattern = r'Structure Work Tracker.*\.xlsx$'
    matching_files = []

    for file_path in ews_files:
        filename = os.path.basename(file_path)
        if re.search(tracker_pattern, filename, re.IGNORECASE):
            file_date = extract_date_from_filename(filename)
            if file_date:
                matching_files.append((file_path, file_date))
                logger.info(f"Found tracker: {filename} (date {file_date:%d-%m-%Y})")
            else:
                logger.warning(f"Matching file without date: {filename}")

    if not matching_files:
        raise Exception("❌ No valid EWS LIG tracker files found")

    latest_file, latest_date = max(matching_files, key=lambda x: x[1])
    EWS_LIG_STRUCTURE_KEY = latest_file
    logger.info(f"✅ Using tracker: {EWS_LIG_STRUCTURE_KEY} ({latest_date:%d-%m-%Y})")

    setup_dynamic_months_and_targets(latest_date)
    return EWS_LIG_STRUCTURE_KEY


def get_previous_months():
    return PROCESSING_MONTHS


# =============== CORE FUNCTIONS ===============

def detect_tracker_year(sheet, pour_cols, row_start, row_end):
    years_found = set()
    for col in pour_cols:
        for row in range(row_start, row_end + 1):
            cell_value = sheet[f"{col}{row}"].value
            dt = parse_date(cell_value)
            if dt:
                years_found.add(dt.year)
    return TRACKER_DATE.year if TRACKER_DATE else (max(years_found) if years_found else datetime.now().year)


def init_cos():
    return ibm_boto3.client(
        "s3",
        ibm_api_key_id=COS_API_KEY,
        ibm_service_instance_id=COS_CRN,
        config=Config(signature_version="oauth"),
        endpoint_url=COS_ENDPOINT,
    )


def download_file_bytes(cos, key):
    obj = cos.get_object(Bucket=BUCKET, Key=key)
    return obj["Body"].read()


def get_targets_from_kra(wb, sheet_name, cell_map):
    sheet = wb[sheet_name]
    targets = {}
    for month in MONTHS:
        if month in cell_map:
            value = sheet[cell_map[month]].value
            try:
                targets[month] = int(str(value).strip().split()[0]) if value else 0
            except Exception:
                targets[month] = 0
        else:
            targets[month] = 0
    return targets


# ---- UNIVERSAL DATE PARSER ----
def parse_date(val):
    """Try to parse any Excel cell value into datetime."""
    if isinstance(val, datetime):
        return val

    if isinstance(val, str) and val.strip():
        dt = pd.to_datetime(val.strip(), dayfirst=True, errors="coerce")
        if pd.notna(dt):
            return dt.to_pydatetime()

        for fmt in [
            "%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d.%m.%Y",
            "%d-%b-%y", "%d-%b-%Y", "%d/%b/%Y", "%d %b %Y"
        ]:
            try:
                return datetime.strptime(val.strip(), fmt)
            except ValueError:
                continue
    return None


# ---- POUR COUNTING ----
def count_pours_for_month(sheet, pour_cols, row_start, row_end, month_num, year):
    count = 0
    for col in pour_cols:
        for row in range(row_start, row_end + 1):
            dt = parse_date(sheet[f"{col}{row}"].value)
            if dt and dt.month == month_num and dt.year == year:
                count += 1
    return count


def count_pours(sheet, pour_cols, row_start, row_end, months, year):
    results = {}
    for m in months:
        month_num = MONTH_TO_NUM[m]
        results[m] = count_pours_for_month(sheet, pour_cols, row_start, row_end, month_num, year)
    return results


# ---- DATAFRAME BUILDER ----
def build_structure_dataframe(tower_name, targets, completed):
    prev_months = get_previous_months()
    weightage = 100

    cum_targets, cum_completed = {}, {}
    for i, m in enumerate(MONTHS):
        cum_targets[m] = sum(targets[mm] for mm in MONTHS[:i + 1])
        cum_completed[m] = sum(completed.get(mm, 0) for mm in MONTHS[:i + 1] if mm in prev_months)

    def pct(m):
        if m not in prev_months:
            return ""
        t, d = cum_targets[m], cum_completed[m]
        return f"{min(round((d / t) * 100, 2), 100)}%" if t else "0.0%"

    total_target = sum(targets.values())
    target_parts = [f"{targets[m]} {m}" for m in MONTHS if targets[m] > 0]
    target_text = f"{total_target} Pours ({', '.join(target_parts)})"

    row = {
        "Milestone": f"{tower_name} Structure",
        "Target Till": target_text,
        "Weightage": weightage,
        "Weighted Delay against Targets": "",
        "Total achieved": "",
        "Delay Reasons": "",
    }

    for m in MONTHS:
        row[f"% Work Done against Target-Till {m}"] = pct(m)
    for m in MONTHS:
        row[f"Target achieved in {m}"] = f"{completed[m]} out of {targets[m]}" if m in prev_months else ""

    row["Total achieved"] = f"{sum(completed[m] for m in prev_months)} out of {total_target}"

    if prev_months:
        last_m = prev_months[-1]
        pct_val = pct(last_m).replace("%", "")
        row["Weighted Delay against Targets"] = f"{round((float(pct_val) * weightage) / 100, 2)}%" if pct_val else "0.0%"

    all_cols = ["Milestone", "Target Till"] + [f"% Work Done against Target-Till {m}" for m in MONTHS] + [
        "Weightage", "Weighted Delay against Targets"] + [f"Target achieved in {m}" for m in MONTHS] + [
                   "Total achieved", "Delay Reasons"]
    return pd.DataFrame([row], columns=all_cols)


# ---- REPORT WRITER ----
def write_excel_report(dfs, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "EWS-LIG Milestones"

    current_date = datetime.now().strftime("%d-%m-%Y")
    ws.append(["EWS-LIG Milestones Report"])
    ws.append([f"Report Generated on: {current_date}"])
    ws.append([])

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="000000")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    max_cols = len(dfs[0][1].columns) if dfs else 12
    ws.merge_cells(f"A1:{get_column_letter(max_cols)}1")
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws["A1"].fill = GREY

    for title, df, total_label in dfs:
        ws.append([title])
        title_row = ws.max_row
        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=len(df.columns))
        for cell in ws[title_row]:
            cell.fill = GREY
            cell.font = bold
            cell.alignment = center
            cell.border = border

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        header_row = title_row + 1
        for cell in ws[header_row]:
            cell.font = bold
            cell.alignment = center
            cell.border = border
        for r in range(header_row + 1, ws.max_row + 1):
            for c in ws[r]:
                c.alignment = left if c.col_idx in (1, 2) else center
                c.border = border

        total_delay = sum(float(str(v).strip("%")) for v in df["Weighted Delay against Targets"] if v and str(v).strip())
        total_row = [""] * len(df.columns)
        total_row[0] = total_label
        for idx, col in enumerate(df.columns, start=1):
            if col == "Weighted Delay against Targets":
                total_row[idx - 1] = f"{round(total_delay, 2)}%"
        ws.append(total_row)
        for c in ws[ws.max_row]:
            c.font = bold
            c.fill = YELLOW
            c.alignment = left if c.col_idx == 1 else center
            c.border = border

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)
    wb.save(filename)
    logger.info(f"Report saved to {filename}")

def get_unique_filename(base_name):
    """
    If file exists, append (1), (2), etc.
    """
    if not os.path.exists(base_name):
        return base_name

    name, ext = os.path.splitext(base_name)
    counter = 1
    new_name = f"{name}({counter}){ext}"
    while os.path.exists(new_name):
        counter += 1
        new_name = f"{name}({counter}){ext}"
    return new_name


# =============== MAIN ===============
def main():
    logger.info("=== STARTING EWS-LIG REPORT GENERATION ===")
    cos = init_cos()
    get_latest_tracker_paths(cos)

    kra_raw = download_file_bytes(cos, EWS_LIG_KRA_KEY)
    kra_wb = load_workbook(filename=BytesIO(kra_raw), data_only=True)
    tracker_raw = download_file_bytes(cos, EWS_LIG_STRUCTURE_KEY)
    tracker_wb = load_workbook(filename=BytesIO(tracker_raw), data_only=True)
    sheet = tracker_wb[TRACKER_SHEET]

    tracker_year = detect_tracker_year(sheet, TOWER1_POUR_COLS, TOWER1_ROW_START, TOWER1_ROW_END)
    logger.info(f"Using tracker year: {tracker_year}")

    dfs = []
    for tname, cellmap, cols, r1, r2 in [
        ("Tower 1", TOWER1_TARGETS_CELLS, TOWER1_POUR_COLS, TOWER1_ROW_START, TOWER1_ROW_END),
        ("Tower 3", TOWER3_TARGETS_CELLS, TOWER3_POUR_COLS, TOWER3_ROW_START, TOWER3_ROW_END),
        ("Tower 2", TOWER2_TARGETS_CELLS, TOWER2_POUR_COLS, TOWER2_ROW_START, TOWER2_ROW_END),
    ]:
        targets = get_targets_from_kra(kra_wb, KRA_SHEET, cellmap)
        completed = count_pours(sheet, cols, r1, r2, MONTHS, tracker_year)
        df = build_structure_dataframe(tname, targets, completed)
        dfs.append((f"{tname} Structure Progress Against Milestones", df, f"Total Delay {tname} Structure"))
        logger.info(f"{tname}: Targets={targets}, Completed={completed}")

    base_filename = f"EWS_LIG_Milestone_Report ({TRACKER_DATE:%d-%m-%Y}).xlsx"
    filename = get_unique_filename(base_filename)

    write_excel_report(dfs, filename)
    logger.info("=== REPORT GENERATION COMPLETE ===")


if __name__ == "__main__":
    main()
